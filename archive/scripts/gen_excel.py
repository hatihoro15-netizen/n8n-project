import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HF = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HFL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TF = Font(name="Arial", bold=True, size=16, color="1F4E79")
SF = Font(name="Arial", bold=True, size=13, color="1F4E79")
DF = Font(name="Arial", size=11)
DB = Font(name="Arial", size=11, bold=True)
AF = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WF = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GF = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RF = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YF = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
UF = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
TB = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
CT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)

def ahs(ws, row, mc):
    for col in range(1, mc+1):
        c = ws.cell(row=row, column=col)
        c.font = HF; c.fill = HFL; c.alignment = CT; c.border = TB

def ads(ws, sr, er, mc, cc=None):
    if cc is None: cc = []
    for r in range(sr, er+1):
        fl = AF if (r-sr)%2==1 else WF
        for c in range(1, mc+1):
            cl = ws.cell(row=r, column=c)
            cl.font = DF; cl.fill = fl; cl.border = TB
            cl.alignment = CT if c in cc else LW

def att(ws, t, mc, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1); c.value = t; c.font = TF
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 40

def ast(ws, t, mc, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1); c.value = t; c.font = SF
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30

def aw(ws, mc, mn=12, mx=50):
    for col in range(1, mc+1):
        ml = mn
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    l = max(len(str(ln)) for ln in str(cell.value).split(chr(10)))
                    a = int(l * 1.3)
                    if a > ml: ml = min(a, mx)
        ws.column_dimensions[get_column_letter(col)].width = ml + 2

# SHEET 1
ws1 = wb.active
ws1.title = "전체 시스템 개요"
att(ws1, "N8N YouTube 완전 자동화 시스템 - 전체 개요", 2, row=1)
for c, h in enumerate(["항목", "내용"], 1):
    ws1.cell(row=3, column=c, value=h)
ahs(ws1, 3, 2)
d1 = [
    ("프로젝트명", "N8N YouTube 완전 자동화 시스템"),
    ("프로젝트 목표", "AI를 활용한 YouTube 영상 기획-제작-업로드-커뮤니티 관리 완전 자동화"),
    ("n8n 서버 주소", "https://n8n.srv1345711.hstgr.cloud"),
    ("GitHub 저장소", "https://github.com/hatihoro15-netizen/n8n-project"),
    ("운영 채널 수", "3개 (루믹스 솔루션, 온카스터디, 슬롯)"),
    ("총 워크플로우 수", "5개 (루믹스 숏폼/롱폼, 온카스터디 숏폼/롱폼, 슬롯 쇼츠)"),
    ("자동화 범위", "주제 생성 > 나레이션 > 이미지 > 영상 > 편집 > 업로드 > 댓글"),
    ("핵심 AI 엔진", "Google Gemini 2.5 Flash, fal.ai (이미지/TTS/BGM), Kie.ai (영상)"),
    ("영상 편집", "Shotstack API (타임라인 조립 + 렌더링)"),
    ("데이터 관리", "Google Sheets (채널별 시트 관리)"),
    ("일일 생산량", "숏폼 6개 + 롱폼 2개 = 총 8개/일"),
    ("월간 생산량", "숏폼 180개 + 롱폼 60개 = 총 240개/월"),
    ("예상 월 운영비", "약 /월 (API 비용 기준)"),
    ("서버 호스팅", "Hostinger VPS (n8n 자체 호스팅)"),
    ("프로젝트 시작일", "2025년"),
    ("현재 상태", "루믹스 크레덴셜 완료, 온카스터디/슬롯 워크플로우 생성 진행중"),
]
for i, (k, v) in enumerate(d1):
    ws1.cell(row=4+i, column=1, value=k)
    ws1.cell(row=4+i, column=2, value=v)
ads(ws1, 4, 4+len(d1)-1, 2, cc=[1])
for r in range(4, 4+len(d1)):
    ws1.cell(row=r, column=1).font = DB
aw(ws1, 2, mn=18, mx=70)
ws1.freeze_panes = "A4"
print("Sheet 1 done")

# SHEET 2
ws2 = wb.create_sheet("부서별 역할 분담")
att(ws2, "부서별 역할 분담", 6, row=1)
for c, h in enumerate(["부서명", "담당자", "역할", "자동화 여부", "사용 도구", "주요 업무 상세"], 1):
    ws2.cell(row=3, column=c, value=h)
ahs(ws2, 3, 6)
d2 = [
    ("콘텐츠 기획팀", "AI + 관리자", "AI 주제 생성, 프롬프트 최적화", "반자동", "Gemini, n8n", "채널별 톤앤매너, 프롬프트 최적화, 콘텐츠 방향, 간접광고 전략, 피드 최적화"),
    ("영상 제작팀", "n8n 자동화", "워크플로우/품질 관리", "완전 자동", "fal.ai, Kie.ai, Shotstack", "이미지(FLUX2Pro), TTS(ElevenLabs), 영상(Kling2.6), BGM(Beatoven), 업스케일(AuraSR)"),
    ("채널 운영팀", "자동화+관리자", "업로드/댓글/커뮤니티", "반자동", "YouTube API, n8n", "자동 업로드(OAuth2), 첫 댓글, 커뮤니티 탭, 스케줄, 브랜딩"),
    ("데이터 분석팀", "Sheets+관리자", "데이터/성과 분석", "반자동", "Sheets, Analytics", "실시간 기록, 성과 추적, CTR 분석, 구독자 모니터링, A/B 테스트"),
    ("기술 인프라팀", "관리자", "서버/API/에러 관리", "수동+모니터링", "n8n, Hostinger VPS", "서버 모니터링, API 사용량, 에러 대응, 크레덴셜 보안, 비용 추적"),
    ("마케팅/SEO팀", "관리자+AI", "썸네일/SEO/해시태그", "반자동", "AI이미지, SEO도구", "썸네일 최적화, SEO 키워드, 해시태그 전략, 크로스 프로모션, 경쟁분석"),
]
for i, row in enumerate(d2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=4+i, column=c, value=val)
ads(ws2, 4, 9, 6, cc=[2, 4])
for r in range(4, 10):
    ws2.cell(row=r, column=1).font = DB
aw(ws2, 6, mn=14, mx=55)
ws2.column_dimensions["F"].width = 55
ws2.freeze_panes = "A4"
print("Sheet 2 done")
