import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HF = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HFL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TF = Font(name="Arial", bold=True, size=16, color="1F4E79")
SF = Font(name="Arial", bold=True, size=13, color="1F4E79")
DF = Font(name="Arial", size=11)
DB = Font(name="Arial", size=11, bold=True)
AF = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WF = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GF = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RF = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YF = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
UPF = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
TB = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
CT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)

def ahs(ws, row, mc):
    for col in range(1, mc+1):
        c = ws.cell(row=row, column=col)
        c.font = HF; c.fill = HFL; c.alignment = CT; c.border = TB

def ads(ws, sr, er, mc, cc=None):
    if cc is None: cc = []
    for r in range(sr, er+1):
        fl = AF if (r-sr)%2==1 else WF
        for c in range(1, mc+1):
            cl = ws.cell(row=r, column=c)
            cl.font = DF; cl.fill = fl; cl.border = TB
            cl.alignment = CT if c in cc else LW

def att(ws, t, mc, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1); c.value = t; c.font = TF
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 40

def ast(ws, t, mc, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1); c.value = t; c.font = SF
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30

def aw(ws, mc, mn=12, mx=50):
    for col in range(1, mc+1):
        ml = mn
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    l = max(len(str(ln)) for ln in str(cell.value).split(chr(10)))
                    a = int(l * 1.3)
                    if a > ml: ml = min(a, mx)
        ws.column_dimensions[get_column_letter(col)].width = ml + 2

# SHEET 1
ws1 = wb.active
ws1.title = "전체 시스템 개요"
att(ws1, "N8N YouTube 완전 자동화 시스템 - 전체 개요", 2, row=1)
for c, h in enumerate(["항목", "내용"], 1):
    ws1.cell(row=3, column=c, value=h)
ahs(ws1, 3, 2)
d1 = [
    ("프로젝트명", "N8N YouTube 완전 자동화 시스템"),
    ("프로젝트 목표", "AI를 활용한 YouTube 영상 기획-제작-업로드-커뮤니티 관리 완전 자동화"),
    ("n8n 서버 주소", "https://n8n.srv1345711.hstgr.cloud"),
    ("GitHub 저장소", "https://github.com/hatihoro15-netizen/n8n-project"),
    ("운영 채널 수", "3개 (루믹스 솔루션, 온카스터디, 슬롯)"),
    ("총 워크플로우 수", "5개 (루믹스 숏폼/롱폼, 온카스터디 숏폼/롱폼, 슬롯 쇼츠)"),
    ("자동화 범위", "주제 생성 > 나레이션 > 이미지 > 영상 > 편집 > 업로드 > 댓글"),
    ("핵심 AI 엔진", "Google Gemini 2.5 Flash, fal.ai, Kie.ai"),
    ("영상 편집", "Shotstack API (타임라인 조립 + 렌더링)"),
    ("데이터 관리", "Google Sheets (채널별 시트 관리)"),
    ("일일 생산량", "숏폼 6개 + 롱폼 2개 = 총 8개/일"),
    ("월간 생산량", "숏폼 180개 + 롱폼 60개 = 총 240개/월"),
    ("예상 월 운영비", "약 $357/월 (API 비용 기준)"),
    ("서버 호스팅", "Hostinger VPS (n8n 자체 호스팅)"),
    ("프로젝트 시작일", "2025년"),
    ("현재 상태", "루믹스 크레덴셜 완료, 온카스터디/슬롯 진행중"),
]
for i, (k, v) in enumerate(d1):
    ws1.cell(row=4+i, column=1, value=k)
    ws1.cell(row=4+i, column=2, value=v)
ads(ws1, 4, 4+len(d1)-1, 2, cc=[1])
for r in range(4, 4+len(d1)):
    ws1.cell(row=r, column=1).font = DB
aw(ws1, 2, mn=18, mx=70)
ws1.freeze_panes = "A4"
print("Sheet 1 done")

# SHEET 2
ws2 = wb.create_sheet("부서별 역할 분담")
att(ws2, "부서별 역할 분담", 6, row=1)
for c, h in enumerate(["부서명", "담당자", "역할", "자동화 여부", "사용 도구", "주요 업무 상세"], 1):
    ws2.cell(row=3, column=c, value=h)
ahs(ws2, 3, 6)
d2 = [
    ("콘텐츠 기획팀", "AI + 관리자", "AI 주제 생성, 프롬프트 최적화", "반자동", "Gemini, n8n", "- 채널별 톤앤매너 관리\n- AI 프롬프트 최적화 및 업데이트\n- 콘텐츠 방향/트렌드 반영\n- 간접광고 삽입 전략\n- 피드 최적화 키워드 관리"),
    ("영상 제작팀", "n8n 자동화", "워크플로우/품질 관리", "완전 자동", "fal.ai, Kie.ai, Shotstack", "- 이미지 품질 관리 (FLUX 2 Pro)\n- TTS 음성 (ElevenLabs 한국어)\n- AI 영상 (Kling 2.6 via Kie.ai)\n- BGM 볼륨 밸런스 (Beatoven)\n- Shotstack 템플릿 관리\n- 업스케일 (AuraSR 2배)"),
    ("채널 운영팀", "자동화+관리자", "업로드/댓글/커뮤니티", "반자동", "YouTube API, n8n", "- 영상 자동 업로드 (OAuth2)\n- 첫 번째 댓글 자동 작성\n- 커뮤니티 탭 관리\n- 업로드 스케줄 관리\n- 채널 브랜딩 일관성"),
    ("데이터 분석팀", "Sheets+관리자", "데이터/성과 분석", "반자동", "Sheets, Analytics", "- 실시간 데이터 기록\n- 성과 추적 (조회/좋아요/댓글)\n- CTR 분석 및 개선\n- 구독자 모니터링\n- A/B 테스트\n- 월간/주간 리포트"),
    ("기술 인프라팀", "관리자", "서버/API/에러 관리", "수동+모니터링", "n8n, Hostinger VPS", "- 서버 모니터링/유지보수\n- API 사용량 모니터링\n- 에러 대응/재시도 로직\n- 크레덴셜 보안 관리\n- 월간 비용 추적\n- SSL 인증서 관리"),
    ("마케팅/SEO팀", "관리자+AI", "썸네일/SEO/해시태그", "반자동", "AI이미지, SEO도구", "- 썸네일 자동 생성 최적화\n- 제목/설명 SEO 키워드\n- 해시태그 전략 수립\n- 크로스 프로모션\n- 트렌드 키워드 반영\n- 경쟁 채널 분석"),
]
for i, row in enumerate(d2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=4+i, column=c, value=val)
ads(ws2, 4, 9, 6, cc=[2, 4])
for r in range(4, 10):
    ws2.cell(row=r, column=1).font = DB
aw(ws2, 6, mn=14, mx=55)
ws2.column_dimensions["F"].width = 55
ws2.freeze_panes = "A4"
print("Sheet 2 done")

# SHEET 3
ws3 = wb.create_sheet("채널별 워크플로우")
att(ws3, "채널별 워크플로우 현황", 8, row=1)
for c, h in enumerate(["채널", "워크플로우명", "워크플로우 ID", "형식", "콘텐츠 방향", "스케줄", "영상 길이", "상태"], 1):
    ws3.cell(row=3, column=c, value=h)
ahs(ws3, 3, 8)
d3 = [
    ("루믹스 솔루션", "숏폼 v3", "9YOHS8N1URWlzGWj", "쇼츠 (9:16)", "IT 솔루션 간접광고 + 피드 최적화", "12시간마다 (08:00, 20:00)", "30-50초", "크레덴셜 완료"),
    ("루믹스 솔루션", "롱폼 v1", "dsP2aQ1YRyeFRRNA", "일반영상 (16:9)", "루믹스 솔루션 홍보/장점 소개", "24시간마다 (14:00)", "2-3분", "크레덴셜 완료"),
    ("온카스터디", "숏폼 v1", "(새로 생성중)", "쇼츠 (9:16)", "간접광고 다양한 스타일 + 피드 최적화", "12시간마다 (08:00, 20:00)", "30-50초", "생성중"),
    ("온카스터디", "롱폼 v1", "(새로 생성중)", "일반영상 (16:9)", "온카스터디 홍보/장점 소개", "24시간마다 (14:00)", "2-3분", "생성중"),
    ("슬롯", "쇼츠 v1", "(새로 생성중)", "쇼츠 (9:16)", "안전 게임 정보/건전한 슬롯", "12시간마다 (08:00, 20:00)", "30-50초", "생성중"),
]
for i, row in enumerate(d3):
    for c, val in enumerate(row, 1):
        ws3.cell(row=4+i, column=c, value=val)
ads(ws3, 4, 8, 8, cc=[3, 4, 6, 7, 8])
for r in range(4, 9):
    ws3.cell(row=r, column=1).font = DB
    st = ws3.cell(row=r, column=8)
    if st.value == "크레덴셜 완료": st.fill = GF
    elif st.value == "생성중": st.fill = YF
aw(ws3, 8, mn=14, mx=45)
ws3.freeze_panes = "A4"
print("Sheet 3 done")

# SHEET 4
ws4 = wb.create_sheet("API 크레덴셜 관리")
att(ws4, "API 및 크레덴셜 관리 현황", 7, row=1)
for c, h in enumerate(["서비스명", "용도", "크레덴셜명", "크레덴셜 ID", "타입", "사용 워크플로우", "월 예상 비용"], 1):
    ws4.cell(row=3, column=c, value=h)
ahs(ws4, 3, 7)
d4 = [
    ("fal.ai", "이미지/TTS/업스케일/BGM", "Header Auth account", "R0m2RD0rtE8IKRt6", "httpHeaderAuth", "전체 워크플로우", "$50-$100"),
    ("Kie.ai", "영상생성 (Kling 2.6)", "Kie.ai", "34ktW72w0p8fCfUQ", "httpHeaderAuth", "숏폼 전체 (3채널)", "$30-$60"),
    ("Shotstack", "영상편집/렌더링", "Shotstack", "3oEYwvtDnmFeylkp", "httpHeaderAuth", "전체 워크플로우", "$25-$50"),
    ("Pexels", "스톡영상 검색", "Pexels", "1vPRgFSX7u4ecIy4", "httpHeaderAuth", "롱폼 전체 (2채널)", "무료"),
    ("Google Gemini", "AI 주제/나레이션/대본", "Gemini(PaLM) Api account", "IKP349r08J9Hoz5E", "googlePalmApi", "전체 워크플로우", "무료 (일정량)"),
    ("Google Sheets", "데이터 기록/관리", "Google Sheets account", "CWBUyXUqCU9p5VAg", "googleSheetsOAuth2Api", "전체 워크플로우", "무료"),
    ("YouTube (루믹스)", "영상 업로드", "YouTube account", "kRKBMYWf6cB72qUi", "youTubeOAuth2Api", "루믹스만", "무료"),
    ("YouTube (온카스터디)", "영상 업로드", "미생성", "-", "youTubeOAuth2Api", "온카스터디만", "무료"),
    ("YouTube (슬롯)", "영상 업로드", "미생성", "-", "youTubeOAuth2Api", "슬롯만", "무료"),
]
for i, row in enumerate(d4):
    for c, val in enumerate(row, 1):
        ws4.cell(row=4+i, column=c, value=val)
ads(ws4, 4, 12, 7, cc=[4, 5, 7])
for r in range(4, 13):
    ws4.cell(row=r, column=1).font = DB
    cr = ws4.cell(row=r, column=3)
    if cr.value == "미생성": cr.fill = RF
aw(ws4, 7, mn=14, mx=45)
ws4.freeze_panes = "A4"
print("Sheet 4 done")

# SHEET 5
ws5 = wb.create_sheet("영상 제작 파이프라인")
att(ws5, "영상 제작 파이프라인 상세", 5, row=1)
ast(ws5, "[ 숏폼 파이프라인 (쇼츠 30-50초) ]", 5, row=3)
for c, h in enumerate(["단계", "프로세스명", "상세 설명", "사용 서비스", "예상 소요"], 1):
    ws5.cell(row=4, column=c, value=h)
ahs(ws5, 4, 5)
sp = [
    ("1", "스케줄 트리거", "12시간 간격 자동 시작 (08:00, 20:00 KST)", "n8n Schedule", "즉시"),
    ("2", "AI 주제 생성", "Gemini 2.5 Flash: 제목, 나레이션(5파트), 캡션, 댓글, BGM프롬프트", "Gemini 2.5 Flash", "3-5초"),
    ("3", "Sheets 기록", "주제 데이터 시트 저장, Status: 생성중", "Google Sheets", "1-2초"),
    ("4", "나레이션 5파트 분할", "Gemini로 나레이션 5개 파트 균등 분배 (각 6-10초)", "Gemini 2.5 Flash", "2-3초"),
    ("5", "병렬 처리 시작", "TTS + 이미지 + BGM 동시 생성", "n8n Parallel", "-"),
    ("6", "TTS 생성 (x5)", "ElevenLabs via fal.ai 한국어 음성 5파트", "ElevenLabs/fal.ai", "10-15초"),
    ("7", "이미지 생성 (x5)", "FLUX 2 Pro via fal.ai, 1080x1920 세로 5장", "FLUX 2 Pro/fal.ai", "15-20초"),
    ("8", "이미지 업스케일 (x5)", "AuraSR via fal.ai 2배 해상도 (2160x3840)", "AuraSR/fal.ai", "10-15초"),
    ("9", "AI 영상 생성 (x5)", "Kling 2.6 via Kie.ai 이미지->영상 (각 5초)", "Kling 2.6/Kie.ai", "60-120초"),
    ("10", "BGM 생성", "Beatoven via fal.ai 배경음악 약 40초", "Beatoven/fal.ai", "20-30초"),
    ("11", "Shotstack 타임라인", "영상+나레이션+자막+BGM 타임라인 구성", "Shotstack API", "2-3초"),
    ("12", "Shotstack 렌더링", "MP4 1080x1920 렌더링 + 폴링 대기", "Shotstack API", "30-60초"),
    ("13", "YouTube 업로드+완료", "업로드 + 첫 댓글 + 시트 상태 업데이트", "YouTube+Sheets", "10-20초"),
]
for i, row in enumerate(sp):
    for c, val in enumerate(row, 1):
        ws5.cell(row=5+i, column=c, value=val)
ads(ws5, 5, 17, 5, cc=[1, 4, 5])

tr = 18
ws5.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
ws5.cell(row=tr, column=1, value="총 예상 소요 시간 (숏폼 1개)").font = DB
ws5.cell(row=tr, column=1).alignment = CT
for c in range(1, 6): ws5.cell(row=tr, column=c).border = TB
ws5.cell(row=tr, column=4, value="전체").font = DB; ws5.cell(row=tr, column=4).alignment = CT
ws5.cell(row=tr, column=5, value="약 3-5분").font = DB; ws5.cell(row=tr, column=5).alignment = CT

lf = 20
ast(ws5, "[ 롱폼 파이프라인 (일반영상 2-3분) ]", 5, row=lf)
for c, h in enumerate(["단계", "프로세스명", "상세 설명", "사용 서비스", "예상 소요"], 1):
    ws5.cell(row=lf+1, column=c, value=h)
ahs(ws5, lf+1, 5)
lp = [
    ("1", "스케줄 트리거", "24시간 간격 자동 시작 (14:00 KST)", "n8n Schedule", "즉시"),
    ("2", "AI 대본 생성", "Gemini 2.5 Flash: 10개 씬 대본 + Pexels 키워드", "Gemini 2.5 Flash", "5-8초"),
    ("3", "Sheets 기록", "대본 데이터 시트 저장, Status: 생성중", "Google Sheets", "1-2초"),
    ("4", "10파트 분리", "대본을 10개 씬으로 분리 (각 12-18초)", "n8n Code Node", "1초"),
    ("5", "병렬 처리 시작", "TTS x10 + Pexels x10 + BGM 동시 진행", "n8n Parallel", "-"),
    ("6", "TTS 생성 (x10)", "10개 씬 TTS 변환 (ElevenLabs 한국어)", "ElevenLabs/fal.ai", "15-25초"),
    ("7", "Pexels 스톡영상 (x10)", "씬별 키워드로 스톡영상 검색 (16:9)", "Pexels API", "5-10초"),
    ("8", "BGM 생성", "2-3분 배경음악 생성", "Beatoven/fal.ai", "30-45초"),
    ("9", "Shotstack 타임라인", "16:9, 페이드 트랜지션, 하단 자막바", "Shotstack API", "3-5초"),
    ("10", "Shotstack 렌더링", "MP4 1920x1080 (재시도 최대 3회)", "Shotstack API", "60-120초"),
    ("11", "YouTube 업로드+댓글", "업로드 + 제목/설명/태그 + 첫 댓글", "YouTube API", "10-20초"),
    ("12", "시트 상태 업데이트", "Status: 업로드완료, URL 기록", "Google Sheets", "1-2초"),
]
for i, row in enumerate(lp):
    for c, val in enumerate(row, 1):
        ws5.cell(row=lf+2+i, column=c, value=val)
ads(ws5, lf+2, lf+2+len(lp)-1, 5, cc=[1, 4, 5])

tr2 = lf + 2 + len(lp)
ws5.merge_cells(start_row=tr2, start_column=1, end_row=tr2, end_column=3)
ws5.cell(row=tr2, column=1, value="총 예상 소요 시간 (롱폼 1개)").font = DB
ws5.cell(row=tr2, column=1).alignment = CT
for c in range(1, 6): ws5.cell(row=tr2, column=c).border = TB
ws5.cell(row=tr2, column=4, value="전체").font = DB; ws5.cell(row=tr2, column=4).alignment = CT
ws5.cell(row=tr2, column=5, value="약 2-4분").font = DB; ws5.cell(row=tr2, column=5).alignment = CT
aw(ws5, 5, mn=14, mx=50)
ws5.freeze_panes = "A5"
print("Sheet 5 done")

# SHEET 6
ws6 = wb.create_sheet("월간 비용 분석")
att(ws6, "월간 비용 분석", 7, row=1)
ast(ws6, "[ 월간 생산량 기준 ]", 7, row=3)
for c, h in enumerate(["구분", "채널 수", "일일 생산량", "월간 생산량"], 1):
    ws6.cell(row=4, column=c, value=h)
ahs(ws6, 4, 4)
vd = [("숏폼", "3개 채널", "6개/일", "180개/월"), ("롱폼", "2개 채널", "2개/일", "60개/월"), ("합계", "3개 채널", "8개/일", "240개/월")]
for i, row in enumerate(vd):
    for c, val in enumerate(row, 1):
        ws6.cell(row=5+i, column=c, value=val)
ads(ws6, 5, 7, 4, cc=[2, 3, 4])
for c in range(1, 5): ws6.cell(row=7, column=c).font = DB

cs = 9
ast(ws6, "[ 영상 1개당 비용 ]", 7, row=cs)
for c, h in enumerate(["서비스", "단가", "숏폼 사용량", "숏폼 단가", "롱폼 사용량", "롱폼 단가", "비고"], 1):
    ws6.cell(row=cs+1, column=c, value=h)
ahs(ws6, cs+1, 7)
cd = [
    ("이미지 (FLUX 2 Pro)", "$0.03/장", "5장", "$0.15", "-", "-", "숏폼 전용"),
    ("업스케일 (AuraSR)", "$0.01/장", "5장", "$0.05", "-", "-", "숏폼 전용"),
    ("AI 영상 (Kling 2.6)", "$0.28/개", "5개", "$1.40", "-", "-", "숏폼 전용"),
    ("TTS (ElevenLabs)", "$0.02/개", "5개", "$0.10", "10개", "$0.20", "파트 수만큼"),
    ("BGM (Beatoven)", "$0.05/곡", "1곡", "$0.05", "1곡", "$0.05", "영상당 1곡"),
    ("Shotstack 렌더링", "$0.10/건", "1건", "$0.10", "1건", "$0.15", "롱폼 약간 비쌈"),
    ("Pexels 스톡영상", "무료", "-", "-", "10개", "$0.00", "무료 API"),
    ("Gemini AI", "무료", "1회", "$0.00", "1회", "$0.00", "Flash 무료"),
    ("YouTube 업로드", "무료", "1회", "$0.00", "1회", "$0.00", "API 무료"),
]
for i, row in enumerate(cd):
    for c, val in enumerate(row, 1):
        ws6.cell(row=cs+2+i, column=c, value=val)
ads(ws6, cs+2, cs+2+len(cd)-1, 7, cc=[2, 3, 4, 5, 6])

sr = cs + 2 + len(cd)
ws6.cell(row=sr, column=1, value="영상 1개 총 비용").font = DB
ws6.cell(row=sr, column=1).alignment = CT
for c in range(1, 8): ws6.cell(row=sr, column=c).border = TB
ws6.cell(row=sr, column=4, value="$1.85").font = DB; ws6.cell(row=sr, column=4).alignment = CT
ws6.cell(row=sr, column=6, value="$0.40").font = DB; ws6.cell(row=sr, column=6).alignment = CT

ms = sr + 2
ast(ws6, "[ 월간 총 비용 ]", 7, row=ms)
for c, h in enumerate(["항목", "단가", "월간 수량", "월간 비용", "비고", "", ""], 1):
    ws6.cell(row=ms+1, column=c, value=h)
ahs(ws6, ms+1, 7)
md = [
    ("숏폼 (3채널)", "$1.85/개", "180개", "$333.00", "3채널 x 60개"),
    ("롱폼 (2채널)", "$0.40/개", "60개", "$24.00", "2채널 x 30개"),
    ("Gemini AI", "무료", "240회", "$0.00", "Flash 무료"),
    ("n8n 서버", "호스팅비", "1개월", "별도", "Hostinger VPS"),
]
for i, row in enumerate(md):
    for c, val in enumerate(row, 1):
        ws6.cell(row=ms+2+i, column=c, value=val)
ads(ws6, ms+2, ms+5, 5, cc=[2, 3, 4])

gt = ms + 6
ws6.cell(row=gt, column=1, value="월간 API 비용 합계").font = Font(name="Arial", bold=True, size=12, color="1F4E79")
ws6.cell(row=gt, column=1).alignment = CT
ws6.cell(row=gt, column=4, value="약 $357/월").font = Font(name="Arial", bold=True, size=12, color="FF0000")
ws6.cell(row=gt, column=4).alignment = CT
for c in range(1, 8): ws6.cell(row=gt, column=c).border = TB
aw(ws6, 7, mn=14, mx=45)
ws6.freeze_panes = "A5"
print("Sheet 6 done")

# SHEET 7
ws7 = wb.create_sheet("Google Sheets 구조")
att(ws7, "Google Sheets 데이터 구조", 4, row=1)
ast(ws7, "[ Google Sheets 목록 ]", 4, row=3)
for c, h in enumerate(["시트명", "Google Sheets ID", "대상 채널", "상태"], 1):
    ws7.cell(row=4, column=c, value=h)
ahs(ws7, 4, 4)
sd = [
    ("n8n LUMIX", "1qPH9TG4M0Hv4V63_LqAb6X2gNl_ksbf2T_oIkJJ06Ag", "루믹스 솔루션", "사용중"),
    ("n8n 온카스터디", "(미생성 - 동일 구조로 생성 필요)", "온카스터디", "미생성"),
    ("n8n 슬롯", "(미생성 - 동일 구조로 생성 필요)", "슬롯", "미생성"),
]
for i, row in enumerate(sd):
    for c, val in enumerate(row, 1):
        ws7.cell(row=5+i, column=c, value=val)
ads(ws7, 5, 7, 4, cc=[3, 4])
for r in range(5, 8):
    st = ws7.cell(row=r, column=4)
    if st.value == "사용중": st.fill = GF
    elif st.value == "미생성": st.fill = RF

cst = 9
ast(ws7, "[ 컬럼 구조 (모든 시트 공통) ]", 4, row=cst)
for c, h in enumerate(["컬럼명", "설명", "데이터 타입", "예시"], 1):
    ws7.cell(row=cst+1, column=c, value=h)
ahs(ws7, cst+1, 4)
cld = [
    ("Subject", "영상 제목 (AI 생성)", "텍스트", "당신의 IT 환경, 루믹스 솔루션으로 혁신하세요"),
    ("Narration", "나레이션 전문 (숏폼5/롱폼10 파트)", "텍스트(긴)", "Part1: 매일 반복되는 업무에..."),
    ("Caption", "YouTube 설명 (SEO 키워드)", "텍스트", "#루믹스솔루션 #IT자동화"),
    ("댓글", "첫 번째 자동 댓글", "텍스트", "궁금한 점 있으시면 댓글로!"),
    ("BGM prompt", "BGM 생성 분위기 (영어)", "텍스트(영어)", "Uplifting corporate music"),
    ("Status", "현재 처리 상태", "선택값", "준비/생성중/생성완료/업로드완료/에러"),
    ("Publish", "발행 상태", "선택값", "draft/published/scheduled"),
    ("업로드 URL", "YouTube 영상 URL", "URL", "https://youtu.be/xxxxxxxxx"),
    ("생성일시", "자동 기록 시간", "날짜/시간", "2025-01-15 14:30:00"),
    ("워크플로우ID", "워크플로우 식별", "텍스트", "9YOHS8N1URWlzGWj"),
    ("영상형식", "숏폼/롱폼 구분", "선택값", "숏폼 / 롱폼"),
]
for i, row in enumerate(cld):
    for c, val in enumerate(row, 1):
        ws7.cell(row=cst+2+i, column=c, value=val)
ads(ws7, cst+2, cst+2+len(cld)-1, 4, cc=[3])
for r in range(cst+2, cst+2+len(cld)):
    ws7.cell(row=r, column=1).font = DB
aw(ws7, 4, mn=16, mx=55)
ws7.freeze_panes = "A5"
print("Sheet 7 done")

# SHEET 8
ws8 = wb.create_sheet("남은 작업 체크리스트")
att(ws8, "남은 작업 체크리스트", 6, row=1)
for c, h in enumerate(["번호", "작업 내용", "담당", "상태", "우선순위", "비고"], 1):
    ws8.cell(row=3, column=c, value=h)
ahs(ws8, 3, 6)
d8 = [
    ("1", "온카스터디 Google Sheets 생성", "수동(관리자)", "미완료", "높음", "LUMIX와 동일 구조"),
    ("2", "슬롯 Google Sheets 생성", "수동(관리자)", "미완료", "높음", "LUMIX와 동일 구조"),
    ("3", "온카스터디 YouTube OAuth2 생성", "수동(관리자)", "미완료", "높음", "Cloud Console OAuth2"),
    ("4", "슬롯 YouTube OAuth2 생성", "수동(관리자)", "미완료", "높음", "Cloud Console OAuth2"),
    ("5", "온카스터디 숏폼 워크플로우", "자동(Claude)", "진행중", "높음", "루믹스 숏폼 v3 복제"),
    ("6", "온카스터디 롱폼 워크플로우", "자동(Claude)", "진행중", "높음", "루믹스 롱폼 v1 복제"),
    ("7", "슬롯 쇼츠 워크플로우", "자동(Claude)", "진행중", "높음", "루믹스 숏폼 v3 복제"),
    ("8", "크레덴셜 연결", "자동(Claude)", "진행중", "높음", "Sheets/Gemini/YouTube"),
    ("9", "루믹스 프롬프트 업데이트", "자동(Claude)", "미완료", "중간", "피드최적화 반영"),
    ("10", "전체 워크플로우 테스트", "수동+자동", "미완료", "높음", "개별 테스트 필요"),
    ("11", "업스케일 비교 테스트", "수동", "미완료", "중간", "AuraSR on/off 비교"),
    ("12", "영상 품질/프롬프트 튜닝", "수동", "미완료", "중간", "시청 후 프롬프트 개선"),
    ("13", "스케줄 최적화", "수동", "미완료", "낮음", "최적 업로드 시간대"),
    ("14", "에러 알림 시스템", "자동", "미완료", "중간", "Slack/이메일 알림"),
    ("15", "비용 모니터링 대시보드", "자동", "미완료", "낮음", "API 사용량 추적"),
    ("16", "썸네일 자동 생성", "자동", "미완료", "낮음", "AI 커스텀 썸네일"),
]
for i, row in enumerate(d8):
    for c, val in enumerate(row, 1):
        ws8.cell(row=4+i, column=c, value=val)
ads(ws8, 4, 4+len(d8)-1, 6, cc=[1, 3, 4, 5])
for r in range(4, 4+len(d8)):
    st = ws8.cell(row=r, column=4)
    if st.value == "진행중": st.fill = YF
    elif st.value == "미완료": st.fill = RF
    pr = ws8.cell(row=r, column=5)
    if pr.value == "높음": pr.font = Font(name="Arial", size=11, bold=True, color="FF0000")
    elif pr.value == "중간": pr.font = Font(name="Arial", size=11, color="FF8C00")
    elif pr.value == "낮음": pr.font = Font(name="Arial", size=11, color="008000")
aw(ws8, 6, mn=10, mx=50)
ws8.column_dimensions["B"].width = 40
ws8.column_dimensions["F"].width = 35
ws8.freeze_panes = "A4"
print("Sheet 8 done")

# SHEET 9
ws9 = wb.create_sheet("스케줄 계획")
att(ws9, "주간 업로드 스케줄 (한국 시간 기준)", 8, row=1)
ast(ws9, "[ 일일 업로드 스케줄 ]", 8, row=3)
for c, h in enumerate(["시간(KST)", "루믹스(숏폼)", "루믹스(롱폼)", "온카(숏폼)", "온카(롱폼)", "슬롯(쇼츠)", "합계", "비고"], 1):
    ws9.cell(row=4, column=c, value=h)
ahs(ws9, 4, 8)
s9 = [
    ("08:00", "업로드", "-", "업로드", "-", "업로드", "3개", "아침 출근시간, 모바일 피크"),
    ("14:00", "-", "업로드", "-", "업로드", "-", "2개", "점심 이후, 롱폼 적합"),
    ("20:00", "업로드", "-", "업로드", "-", "업로드", "3개", "저녁 황금시간, 최고 트래픽"),
]
for i, row in enumerate(s9):
    for c, val in enumerate(row, 1):
        ws9.cell(row=5+i, column=c, value=val)
ads(ws9, 5, 7, 8, cc=[1, 2, 3, 4, 5, 6, 7])
for r in range(5, 8):
    for c in range(2, 7):
        cl = ws9.cell(row=r, column=c)
        if cl.value == "업로드":
            cl.fill = UPF
            cl.font = Font(name="Arial", size=11, bold=True)

dt = 8
dtv = ["일일 합계", "2개", "1개", "2개", "1개", "2개", "8개/일", "총 8개/일"]
for c, val in enumerate(dtv, 1):
    cl = ws9.cell(row=dt, column=c, value=val)
    cl.font = DB; cl.alignment = CT; cl.border = TB
ws9.cell(row=dt, column=7).font = Font(name="Arial", bold=True, size=12, color="FF0000")

wss = dt + 2
ast(ws9, "[ 주간/월간 생산 요약 ]", 8, row=wss)
for c, h in enumerate(["구분", "루믹스 숏폼", "루믹스 롱폼", "온카 숏폼", "온카 롱폼", "슬롯 쇼츠", "합계", ""], 1):
    ws9.cell(row=wss+1, column=c, value=h)
ahs(ws9, wss+1, 8)
smd = [
    ("일일", "2개", "1개", "2개", "1개", "2개", "8개", ""),
    ("주간(7일)", "14개", "7개", "14개", "7개", "14개", "56개", ""),
    ("월간(30일)", "60개", "30개", "60개", "30개", "60개", "240개", ""),
]
for i, row in enumerate(smd):
    for c, val in enumerate(row, 1):
        ws9.cell(row=wss+2+i, column=c, value=val)
ads(ws9, wss+2, wss+4, 7, cc=[1, 2, 3, 4, 5, 6, 7])
for c in range(1, 8): ws9.cell(row=wss+4, column=c).font = DB

ost = wss + 6
ast(ws9, "[ 업로드 시간대 분석 ]", 8, row=ost)
for c, h in enumerate(["시간대", "대상", "선택 이유", "기대 효과", "", "", "", ""], 1):
    ws9.cell(row=ost+1, column=c, value=h)
ahs(ws9, ost+1, 8)
od = [
    ("오전 08:00", "숏폼(쇼츠)", "출근길 모바일 시청 피크", "초기 조회수 확보, 알고리즘 부스트"),
    ("오후 14:00", "롱폼(일반)", "점심 이후 여유시간", "시청시간 극대화, 높은 완시율"),
    ("오후 20:00", "숏폼(쇼츠)", "퇴근 후 황금시간대", "최대 노출, 추천 확률 증가"),
]
for i, row in enumerate(od):
    for c, val in enumerate(row, 1):
        ws9.cell(row=ost+2+i, column=c, value=val)
    ws9.merge_cells(start_row=ost+2+i, start_column=4, end_row=ost+2+i, end_column=8)
ads(ws9, ost+2, ost+4, 4, cc=[1, 2])
aw(ws9, 8, mn=14, mx=40)
ws9.freeze_panes = "A5"
print("Sheet 9 done")

# Save
out = "/Users/gimdongseog/n8n-project/n8n_기획서_최종.xlsx"
wb.save(out)
import os
sz = os.path.getsize(out)
print(f"Saved: {out}")
print(f"Size: {sz:,} bytes ({sz/1024:.1f} KB)")
print(f"Sheets: {wb.sheetnames}")
print("DONE!")
