from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# 색상 정의
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
NORMAL_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(vertical='center', wrap_text=True)

AGENT_COLORS = {
    "PM": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "DA": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "PL": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "QA": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "PR": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "DL": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
}
PRIORITY_COLORS = {
    "높음": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "중간": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "낮음": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_row(ws, row, cols, fill=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT_WRAP if col > 1 else CENTER
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ========== 시트 1: 프로젝트 개요 ==========
ws1 = wb.active
ws1.title = "프로젝트 개요"
ws1.sheet_properties.tabColor = "2F5496"

ws1.merge_cells('A1:D1')
ws1['A1'] = "n8n AI 멀티 에이전트 영상 제작 시스템"
ws1['A1'].font = TITLE_FONT
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A3:D3')
ws1['A3'] = "프로젝트 개요"
ws1['A3'].font = SUBTITLE_FONT

overview = [
    ["항목", "내용", "", ""],
    ["목적", "n8n AI 에이전트가 유튜브 영상을 기획→제작→업로드→분석까지 완전 자율 수행", "", ""],
    ["핵심 개념", "에이전트 = 부서. AI가 서로 의견을 주고받으며 자율적으로 동작", "", ""],
    ["최종 목표", "클릭 0번 — 내가 안 해도 알아서 돌아가는 시스템 (모니터링만)", "", ""],
    ["현재 상태", "n8n으로 숏폼 영상 제작→유튜브 업로드 자동화 완료 (클릭 1번)", "", ""],
    ["다음 단계", "클릭 1번 → 클릭 0번: AI가 기획/판단/분석까지 자율적으로 수행하도록 확장", "", ""],
    ["채널 구분", "온카스터디/루믹스/슬롯/쇼츠 전용 — 각각 별도 유튜브 채널로 운영", "", ""],
]
for i, row_data in enumerate(overview, 4):
    for j, val in enumerate(row_data, 1):
        ws1.cell(row=i, column=j, value=val)
    if i == 4:
        style_header(ws1, i, 2)
    else:
        style_row(ws1, i, 2)

ws1.merge_cells('A13:D13')
ws1['A13'] = "운영 채널 (각각 별도 유튜브 채널)"
ws1['A13'].font = SUBTITLE_FONT

biz_headers = ["#", "채널 (별도 유튜브)", "콘텐츠 유형", "목적"]
for j, h in enumerate(biz_headers, 1):
    ws1.cell(row=14, column=j, value=h)
style_header(ws1, 14, 4)

biz_data = [
    ["1", "온카스터디 채널", "롱폼 + 숏폼", "커뮤니티 홍보 및 알리기 (별도 채널)"],
    ["2", "슬롯 채널", "롱폼 + 숏폼", "슬롯 관련 콘텐츠만 + 온카스터디 광고 (별도 채널)"],
    ["3", "루믹스 솔루션 채널", "롱폼 + 숏폼", "솔루션 판매/광고 — 알리기 (별도 채널)"],
    ["4", "쇼츠 전용 채널", "숏폼 (쇼츠)", "조회수/체류/클릭 분석 → 다양한 영상 → 마지막에 광고 삽입 (별도 채널)"],
]
for i, row_data in enumerate(biz_data, 15):
    for j, val in enumerate(row_data, 1):
        ws1.cell(row=i, column=j, value=val)
    style_row(ws1, i, 4)

ws1.merge_cells('A21:D21')
ws1['A21'] = "전체 사업 구조"
ws1['A21'].font = SUBTITLE_FONT

ws1.merge_cells('A22:D27')
structure_text = """핵심 목적: 온카스터디 / 루믹스 솔루션을 알리고 판매하는 것
각 채널은 별도 유튜브 채널로 운영 (분리 관리)

[쇼츠 전용 채널] 다양한 소재 → 조회수/체류/클릭 데이터 수집 → 마지막에 광고 삽입
[슬롯 채널] 슬롯 관련 콘텐츠 전용 → 온카스터디 광고 삽입 (슬롯 외 광고 없음)
[온카스터디 채널] 커뮤니티 직접 홍보 콘텐츠
[루믹스 솔루션 채널] 솔루션 소개/판매 콘텐츠

현재 상태: n8n 숏폼 자동화 구축 완료 (클릭 1번 → 영상 제작 → 업로드)"""
ws1['A22'] = structure_text
ws1['A22'].font = NORMAL_FONT
ws1['A22'].alignment = Alignment(wrap_text=True, vertical='top')

set_col_widths(ws1, [10, 25, 20, 50])

# ========== 시트 2: 에이전트(부서) 구성 ==========
ws2 = wb.create_sheet("에이전트 구성")
ws2.sheet_properties.tabColor = "4472C4"

ws2.merge_cells('A1:F1')
ws2['A1'] = "에이전트(부서) 상세 역할"
ws2['A1'].font = TITLE_FONT
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 40

headers = ["에이전트", "약칭", "역할", "AI 모델", "핵심 행동", "기존 대비 추가되는 것"]
for j, h in enumerate(headers, 1):
    ws2.cell(row=3, column=j, value=h)
style_header(ws2, 3, 6)

agents = [
    ["총괄 매니저", "PM", "전체 파이프라인 관리, 에이전트 간 조율", "GPT-4 / Claude", "작업 지시, 우선순위 결정, 최종 승인/반려", "NEW — 자동 실행 오케스트레이터"],
    ["데이터 분석팀", "DA", "유튜브 성과 분석, 트렌드 파악, 시청자 행동 분석", "GPT-4", "조회수/CTR/체류시간/클릭 분석 → 리포트 생성", "NEW — 어떤 영상이 잘 되는지 AI가 판단"],
    ["기획팀", "PL", "영상 아이디어 생성, 대본/구성안 작성", "Claude / GPT-4", "DA 리포트 기반 기획안 2~3개 생성 (서로 다른 방향)", "NEW — 사람 대신 AI가 기획"],
    ["검수/QA팀", "QA", "기획안 검토, 피드백, 품질 관리", "Claude", "기획안 평가 → 통과/수정요청/반려, 의견 주고받기", "NEW — AI끼리 토론하며 품질 향상"],
    ["제작팀", "PR", "영상 실제 제작 (스크립트→영상)", "AI 영상 도구", "TTS + 이미지/영상 + 자막 + BGM + 광고 삽입", "기존 워크플로우 활용 (이미 구축됨)"],
    ["배포팀", "DL", "유튜브 업로드, 썸네일, 태그, 설명", "GPT-4", "제목/설명/태그 최적화 → YouTube API 업로드", "기존 워크플로우 활용 (이미 구축됨)"],
]
for i, row_data in enumerate(agents, 4):
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=j, value=val)
    agent_code = row_data[1]
    fill = AGENT_COLORS.get(agent_code)
    style_row(ws2, i, 6, fill)

ws2.merge_cells('A12:F12')
ws2['A12'] = "현재 vs 목표"
ws2['A12'].font = SUBTITLE_FONT

compare_headers = ["구분", "현재 (클릭 1번)", "목표 (클릭 0번)", "", "", ""]
for j, h in enumerate(compare_headers, 1):
    ws2.cell(row=13, column=j, value=h)
style_header(ws2, 13, 3)

compares = [
    ["기획", "내가 직접 주제/대본 정함", "DA가 분석 → PL이 기획 → QA가 검수 (자동)"],
    ["제작", "클릭 1번 → 자동 제작", "승인된 기획안 → 자동 제작 (기존 유지)"],
    ["업로드", "클릭 1번 → 자동 업로드", "제작 완료 → 자동 업로드 (기존 유지)"],
    ["분석", "내가 직접 확인", "DA가 자동 분석 → 다음 기획에 반영"],
    ["판단", "내가 어떤 영상 만들지 결정", "AI가 데이터 기반으로 자율 판단"],
]
for i, row_data in enumerate(compares, 14):
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=j, value=val)
    style_row(ws2, i, 3)

ws2.merge_cells('A21:F21')
ws2['A21'] = "에이전트 간 흐름"
ws2['A21'].font = SUBTITLE_FONT

ws2.merge_cells('A22:F27')
flow_text = """DA (데이터 분석) → PL (기획) ←→ QA (검수/토론) → PR (제작*) → DL (배포*)
     ↑                                                                        │
     └─────────────────── 성과 피드백 (무한 루프) ─────────────────┘

* PR(제작), DL(배포) = 기존에 이미 만든 워크플로우 연결
* 새로 만들 것 = DA(분석) + PL(기획) + QA(검수) + PM(총괄)
* 기획(PL) ↔ 검수(QA): 피드백 루프 최대 3회 수정/재검토"""
ws2['A22'] = flow_text
ws2['A22'].font = NORMAL_FONT
ws2['A22'].alignment = Alignment(wrap_text=True, vertical='top')

set_col_widths(ws2, [15, 8, 30, 18, 40, 35])

# ========== 시트 3: 워크플로우 목록 ==========
ws3 = wb.create_sheet("워크플로우 목록")
ws3.sheet_properties.tabColor = "70AD47"

ws3.merge_cells('A1:G1')
ws3['A1'] = "n8n 워크플로우 전체 목록"
ws3['A1'].font = TITLE_FONT
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

wf_headers = ["코드", "워크플로우명", "담당 에이전트", "트리거", "설명", "상태", "우선순위"]
for j, h in enumerate(wf_headers, 1):
    ws3.cell(row=3, column=j, value=h)
style_header(ws3, 3, 7)

workflows = [
    ["DA-01", "일일 성과 수집", "데이터 분석 (DA)", "Cron (매일 06:00)", "전체 채널 조회수/CTR/체류시간/클릭 데이터 수집 → Sheets", "신규", "높음"],
    ["DA-02", "트렌드/시청자 행동 분석", "데이터 분석 (DA)", "Cron (매일 07:00)", "어떤 피드에서 유입? 어떤 영상에 오래 머무름? 패턴 분석", "신규", "높음"],
    ["DA-03", "AI 분석 리포트", "데이터 분석 (DA)", "DA-01,02 완료 후", "수집 데이터 → AI 분석 → '다음에 이런 영상 만들어라' 리포트", "신규", "높음"],
    ["PL-01", "기획안 자동 생성", "기획 (PL)", "DA-03 완료 후", "분석 리포트 기반 기획안 2~3개 생성 (서로 다른 방향)", "신규", "높음"],
    ["PL-02", "기획안 수정 (피드백 반영)", "기획 (PL)", "QA 피드백 수신 시", "QA 피드백 반영하여 기획안 수정 후 재제출", "신규", "높음"],
    ["QA-01", "기획 검수 + 토론", "검수/QA (QA)", "PL-01 완료 후", "기획안 평가 (점수+피드백), 승인/수정/반려 — AI끼리 의견 교환", "신규", "높음"],
    ["PR-01", "영상 제작", "제작 (PR)", "QA 승인 후", "스크립트→TTS→이미지/영상→자막→BGM→합성", "기존 있음", "높음"],
    ["PR-02", "광고 삽입", "제작 (PR)", "PR-01 완료 후", "채널별 규칙에 따라 광고 자동 삽입", "수정 필요", "높음"],
    ["DL-01", "메타데이터 + 업로드", "배포 (DL)", "PR-02 완료 후", "AI 제목/설명/태그 최적화 → YouTube 업로드", "기존 있음", "높음"],
    ["DL-02", "업로드 알림", "배포 (DL)", "DL-01 완료 후", "Slack/카톡으로 완료 알림", "신규", "중간"],
    ["PM-01", "일일 자동 파이프라인", "총괄 (PM)", "Cron (매일 08:00)", "DA→PL→QA→PR→DL 전체 자동 실행 관리", "신규", "높음"],
    ["PM-02", "에러 감지 + 재시도", "총괄 (PM)", "에러 발생 시", "에러 감지 → 자동 재시도 또는 알림", "신규", "높음"],
]

for i, row_data in enumerate(workflows, 4):
    for j, val in enumerate(row_data, 1):
        ws3.cell(row=i, column=j, value=val)
    code = row_data[0].split("-")[0]
    fill = AGENT_COLORS.get(code)
    style_row(ws3, i, 7, fill)
    # 우선순위 색상
    priority = row_data[6]
    if priority in PRIORITY_COLORS:
        ws3.cell(row=i, column=7).fill = PRIORITY_COLORS[priority]
        ws3.cell(row=i, column=7).font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    # 상태 표시
    status = row_data[5]
    status_colors = {
        "신규": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "기존 있음": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "수정 필요": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }
    if status in status_colors:
        ws3.cell(row=i, column=6).fill = status_colors[status]

set_col_widths(ws3, [10, 25, 18, 22, 55, 12, 10])

# ========== 시트 4: 채널별 콘텐츠 전략 ==========
ws4 = wb.create_sheet("콘텐츠 전략")
ws4.sheet_properties.tabColor = "FFC000"

ws4.merge_cells('A1:F1')
ws4['A1'] = "채널별 콘텐츠 전략"
ws4['A1'].font = TITLE_FONT
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 40

ch_headers = ["채널", "숏폼", "롱폼", "목적", "광고 삽입", "발행 빈도"]
for j, h in enumerate(ch_headers, 1):
    ws4.cell(row=3, column=j, value=h)
style_header(ws4, 3, 6)

channels = [
    ["온카스터디 채널 (별도)", "O", "O", "커뮤니티 직접 홍보/알리기", "자체 홍보 콘텐츠", "주 3~5회"],
    ["슬롯 채널 (별도)", "O", "O", "온카스터디 광고 목적", "슬롯 관련 콘텐츠만 (타 광고 X)", "주 3~5회"],
    ["루믹스 솔루션 채널 (별도)", "O", "O", "솔루션 판매/광고 — 알리기", "자체 판매/홍보 콘텐츠", "주 3~5회"],
    ["쇼츠 전용 채널 (별도)", "O", "-", "데이터 수집 + 광고 삽입", "다양한 소재 → 마지막에 온카/루믹스 광고", "매일 1~3회"],
]
for i, row_data in enumerate(channels, 4):
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=j, value=val)
    style_row(ws4, i, 6)

# 쇼츠 전용 전략 상세
ws4.merge_cells('A10:F10')
ws4['A10'] = "쇼츠 전용 채널 — 상세 전략"
ws4['A10'].font = SUBTITLE_FONT

shorts_headers = ["단계", "내용", "AI 에이전트", "데이터 포인트", "", ""]
for j, h in enumerate(shorts_headers, 1):
    ws4.cell(row=11, column=j, value=h)
style_header(ws4, 11, 4)

shorts_strategy = [
    ["1. 분석", "사람들이 어떤 피드에서 유입하는지 파악", "DA", "유입 소스, 검색어, 추천 알고리즘"],
    ["2. 분석", "어떤 영상에 오래 머무르는지 확인", "DA", "평균 시청 시간, 시청 유지율"],
    ["3. 분석", "어떤 영상을 클릭하는지 확인", "DA", "CTR, 썸네일 클릭율, 노출수"],
    ["4. 기획", "분석 결과 기반 다양한 소재로 영상 기획", "PL", "잘 되는 패턴 + 새로운 실험"],
    ["5. 제작", "다양한 쇼츠 영상 제작", "PR", ""],
    ["6. 광고", "영상 마지막에 온카스터디/루믹스 광고 삽입", "PR", "광고 소재 풀에서 자동 선택"],
    ["7. 배포", "업로드 + 성과 추적", "DL→DA", "업로드 후 24h/48h/7d 성과 수집"],
]
for i, row_data in enumerate(shorts_strategy, 12):
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=j, value=val)
    style_row(ws4, i, 4)

# 광고 삽입 규칙
ws4.merge_cells('A21:F21')
ws4['A21'] = "채널별 광고 삽입 규칙"
ws4['A21'].font = SUBTITLE_FONT

ad_headers = ["채널", "광고 규칙"]
for j, h in enumerate(ad_headers, 1):
    ws4.cell(row=22, column=j, value=h)
style_header(ws4, 22, 2)

ad_rules = [
    ["온카스터디", "자체 홍보 — 커뮤니티 관련 콘텐츠 자체가 광고"],
    ["슬롯 채널", "슬롯 관련 콘텐츠만 + 온카스터디 광고 삽입 (타 광고 없음)"],
    ["루믹스 솔루션", "자체 판매/홍보 — 솔루션 소개 콘텐츠 자체가 광고"],
    ["쇼츠 전용", "다양한 소재 영상 → 마지막에 온카스터디 or 루믹스 광고 삽입"],
]
for i, row_data in enumerate(ad_rules, 23):
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=j, value=val)
    style_row(ws4, i, 2)

set_col_widths(ws4, [15, 40, 15, 40, 15, 15])

# ========== 시트 5: 필요 도구/API ==========
ws5 = wb.create_sheet("필요 도구 및 API")
ws5.sheet_properties.tabColor = "ED7D31"

ws5.merge_cells('A1:D1')
ws5['A1'] = "필요 도구 및 API"
ws5['A1'].font = TITLE_FONT
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 40

ws5['A3'] = "AI 모델"
ws5['A3'].font = SUBTITLE_FONT
for j, h in enumerate(["용도", "도구", "비고", "월 예상 비용"], 1):
    ws5.cell(row=4, column=j, value=h)
style_header(ws5, 4, 4)

ai_tools = [
    ["텍스트 생성 (기획/대본/분석)", "OpenAI GPT-4 / Claude API", "n8n AI 노드 활용", "$20~50"],
    ["TTS (음성 합성)", "ElevenLabs / OpenAI TTS", "나레이션 자동 생성", "$5~22"],
    ["이미지 생성", "DALL-E / Flux / Midjourney", "썸네일 + 영상 내 이미지", "$10~30"],
    ["영상 생성/편집", "Creatomate / Shotstack / Runway", "API로 영상 자동 조합", "$39~100"],
]
for i, row_data in enumerate(ai_tools, 5):
    for j, val in enumerate(row_data, 1):
        ws5.cell(row=i, column=j, value=val)
    style_row(ws5, i, 4)

ws5['A11'] = "외부 서비스"
ws5['A11'].font = SUBTITLE_FONT
for j, h in enumerate(["서비스", "용도", "인증 방식", "비고"], 1):
    ws5.cell(row=12, column=j, value=h)
style_header(ws5, 12, 4)

services = [
    ["YouTube Data API", "성과 분석 + 업로드", "OAuth2", "일일 10,000 유닛 제한"],
    ["YouTube Analytics API", "조회수/CTR/체류시간/유입소스", "OAuth2", "핵심 데이터"],
    ["Google Sheets", "데이터 저장/관리", "OAuth2", "DB 역할"],
    ["Google Trends / SerpAPI", "트렌드 키워드 수집", "API Key", ""],
    ["Google Drive", "영상/광고소재 저장", "OAuth2", ""],
    ["Slack / 카카오톡", "알림 수신", "Webhook / Bot", ""],
]
for i, row_data in enumerate(services, 13):
    for j, val in enumerate(row_data, 1):
        ws5.cell(row=i, column=j, value=val)
    style_row(ws5, i, 4)

set_col_widths(ws5, [30, 30, 20, 25])

# ========== 시트 6: 구축 로드맵 ==========
ws6 = wb.create_sheet("구축 로드맵")
ws6.sheet_properties.tabColor = "5B9BD5"

ws6.merge_cells('A1:E1')
ws6['A1'] = "구축 로드맵"
ws6['A1'].font = TITLE_FONT
ws6['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 40

ws6.merge_cells('A2:E2')
ws6['A2'] = "* 제작(PR) + 배포(DL) 워크플로우는 기존에 이미 있으므로, 분석/기획/검수/총괄 에이전트 구축에 집중"
ws6['A2'].font = Font(name="맑은 고딕", size=10, color="C00000", bold=True)

rm_headers = ["Phase", "기간", "작업 항목", "상세 내용", "상태"]
for j, h in enumerate(rm_headers, 1):
    ws6.cell(row=4, column=j, value=h)
style_header(ws6, 4, 5)

roadmap = [
    ["Phase 1\n서버 이전", "1주", "Hostinger VPS + n8n 설치", "VPS 구매, n8n 템플릿 설치, 도메인 연결", "대기"],
    ["", "", "기존 워크플로우 이전", "현재 로컬 n8n 워크플로우를 VPS로 이전", "대기"],
    ["", "", "API Credential 등록", "YouTube, OpenAI, TTS, Google 등 API 키 등록", "대기"],
    ["", "", "Google Sheets DB 설계", "영상 DB, 기획안 DB, 성과 DB 시트 구조 설계", "대기"],
    ["Phase 2\n데이터 분석 (DA)", "1~2주", "DA-01: 일일 성과 수집", "YouTube Analytics → 조회수/CTR/체류/클릭/유입소스 수집", "대기"],
    ["", "", "DA-02: 시청자 행동 분석", "어떤 피드? 어디서 유입? 어떤 영상에 오래 머무름?", "대기"],
    ["", "", "DA-03: AI 분석 리포트", "데이터 → AI → '이런 영상 만들어라' 리포트 자동 생성", "대기"],
    ["", "", "프롬프트 튜닝", "분석 리포트 품질 확인 및 최적화", "대기"],
    ["Phase 3\n기획+검수 (PL+QA)", "2~3주", "PL-01: 기획안 자동 생성", "분석 리포트 → AI 기획안 2~3개 자동 생성", "대기"],
    ["", "", "QA-01: 기획 검수 + 토론", "기획안 자동 평가 (점수+피드백), 승인/수정/반려", "대기"],
    ["", "", "PL↔QA 피드백 루프", "수정 → 재검토 반복 (최대 3회) 자동화", "대기"],
    ["", "", "검증", "AI가 만든 기획안이 실제 쓸 수 있는 수준인지 확인", "대기"],
    ["Phase 4\n연결 + 총괄", "1~2주", "기존 PR/DL 워크플로우 연결", "QA 승인 → 기존 제작 워크플로우 자동 트리거", "대기"],
    ["", "", "PR-02: 광고 삽입 규칙 수정", "채널별 광고 규칙 적용 (슬롯=슬롯만, 쇼츠=마지막 광고)", "대기"],
    ["", "", "PM-01: 전체 파이프라인 연결", "DA→PL→QA→PR→DL 완전 자동 실행", "대기"],
    ["", "", "PM-02: 에러 처리", "에러 감지 → 재시도 또는 알림", "대기"],
    ["", "", "E2E 테스트", "전체 파이프라인 통합 테스트", "대기"],
    ["Phase 5\n고도화", "지속", "성과 기반 자동 학습", "잘된 영상 패턴 → 다음 기획에 자동 반영", "대기"],
    ["", "", "A/B 테스트 자동화", "같은 주제, 다른 썸네일/제목으로 테스트", "대기"],
    ["", "", "채널별 스타일 분리", "채널마다 톤/무드/스타일 AI 프롬프트 분리", "대기"],
    ["", "", "롱폼 자동화 확장", "쇼츠 안정화 후 롱폼 영상 자동 제작 추가", "대기"],
]

PHASE_COLORS = {
    "Phase 1": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Phase 2": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "Phase 3": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Phase 4": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "Phase 5": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
}

current_phase = ""
for i, row_data in enumerate(roadmap, 5):
    for j, val in enumerate(row_data, 1):
        ws6.cell(row=i, column=j, value=val)
    if row_data[0]:
        current_phase = row_data[0].split("\n")[0]
    fill = None
    for phase_key, phase_fill in PHASE_COLORS.items():
        if phase_key in current_phase:
            fill = phase_fill
            break
    style_row(ws6, i, 5, fill)

set_col_widths(ws6, [18, 10, 30, 50, 10])

# ========== 시트 7: 비용 및 제약 ==========
ws7 = wb.create_sheet("비용 및 제약")
ws7.sheet_properties.tabColor = "C00000"

ws7.merge_cells('A1:D1')
ws7['A1'] = "비용 및 주의사항"
ws7['A1'].font = TITLE_FONT
ws7['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws7.row_dimensions[1].height = 40

ws7['A3'] = "월간 예상 비용"
ws7['A3'].font = SUBTITLE_FONT

cost_headers = ["항목", "도구", "월 비용 (예상)", "비고"]
for j, h in enumerate(cost_headers, 1):
    ws7.cell(row=4, column=j, value=h)
style_header(ws7, 4, 4)

costs = [
    ["서버", "Hostinger VPS (KVM 2)", "$6.99", "2 vCPU / 8GB RAM"],
    ["AI 텍스트", "OpenAI GPT-4 API", "$20~50", "기획/분석/메타데이터 생성"],
    ["AI 텍스트", "Claude API (선택)", "$10~30", "검수/QA 에이전트"],
    ["TTS", "ElevenLabs", "$5~22", "나레이션 음성"],
    ["영상 제작", "Creatomate / Shotstack", "$39~100", "영상 자동 합성"],
    ["이미지", "DALL-E / Flux", "$10~30", "썸네일/영상 내 이미지"],
    ["합계", "", "$91~239", ""],
]
for i, row_data in enumerate(costs, 5):
    for j, val in enumerate(row_data, 1):
        ws7.cell(row=i, column=j, value=val)
    style_row(ws7, i, 4)
    if row_data[0] == "합계":
        for j in range(1, 5):
            ws7.cell(row=i, column=j).font = BOLD_FONT

ws7['A14'] = "주의사항 및 제약"
ws7['A14'].font = SUBTITLE_FONT

limit_headers = ["항목", "내용", "영향", "대응"]
for j, h in enumerate(limit_headers, 1):
    ws7.cell(row=15, column=j, value=h)
style_header(ws7, 15, 4)

limits = [
    ["YouTube API 할당량", "일일 10,000 유닛 제한", "업로드 1건당 1,600 유닛 (일 최대 6건)", "업로드 스케줄링으로 분산"],
    ["AI 영상 품질", "초기에는 품질이 낮을 수 있음", "시청자 이탈 가능", "쇼츠(짧은 영상)부터 시작"],
    ["저작권", "AI 생성 이미지/음악", "저작권 분쟁 가능", "상업용 라이선스 확인"],
    ["API 비용", "사용량에 따라 변동", "예산 초과 가능", "일일 비용 모니터링 워크플로우 추가"],
]
for i, row_data in enumerate(limits, 16):
    for j, val in enumerate(row_data, 1):
        ws7.cell(row=i, column=j, value=val)
    style_row(ws7, i, 4)

set_col_widths(ws7, [20, 30, 30, 25])

# ========== 시트 8: 피드백 루프 상세 ==========
ws8 = wb.create_sheet("피드백 루프 상세")
ws8.sheet_properties.tabColor = "7030A0"

ws8.merge_cells('A1:D1')
ws8['A1'] = "에이전트 간 피드백 루프 (AI 토론 구조)"
ws8['A1'].font = TITLE_FONT
ws8['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 40

ws8['A3'] = "기획(PL) ↔ 검수(QA) 토론 프로세스"
ws8['A3'].font = SUBTITLE_FONT

loop_headers = ["단계", "에이전트", "행동", "결과"]
for j, h in enumerate(loop_headers, 1):
    ws8.cell(row=4, column=j, value=h)
style_header(ws8, 4, 4)

loop_data = [
    ["1", "PL (기획)", "DA 리포트 기반 기획안 A, B 생성", "기획안 2~3개 + 대본 초안"],
    ["2", "QA (검수)", "기획안 평가 (1~10점 채점)", "참신함/조회수 예측/브랜드 적합도 점수"],
    ["3-a", "QA → PR", "7점 이상 → 승인", "제작팀으로 전달, 영상 제작 시작"],
    ["3-b", "QA → PL", "4~6점 → 수정 요청", "'여기 이렇게 바꿔' 구체적 피드백 전달"],
    ["3-c", "QA → PL", "3점 이하 → 폐기", "새 기획안 처음부터 다시 생성 요청"],
    ["4", "PL (기획)", "QA 피드백 반영하여 수정", "수정된 기획안 재제출"],
    ["5", "QA (검수)", "수정된 기획안 재평가", "다시 점수 매기기 (최대 3라운드 반복)"],
    ["6", "PM (총괄)", "3라운드 후에도 미승인 시 개입", "최종 판단 또는 해당 건 스킵"],
]
for i, row_data in enumerate(loop_data, 5):
    for j, val in enumerate(row_data, 1):
        ws8.cell(row=i, column=j, value=val)
    style_row(ws8, i, 4)

ws8['A15'] = "n8n 구현 방법"
ws8['A15'].font = SUBTITLE_FONT

impl_headers = ["구성요소", "n8n 노드", "설명", ""]
for j, h in enumerate(impl_headers, 1):
    ws8.cell(row=16, column=j, value=h)
style_header(ws8, 16, 3)

impl_data = [
    ["기획안 생성", "AI Agent / OpenAI 노드", "시스템 프롬프트에 '기획 에이전트' 역할 부여"],
    ["검수 평가", "AI Agent / OpenAI 노드", "시스템 프롬프트에 '검수 에이전트' 역할 부여 + JSON 출력 (점수+피드백)"],
    ["점수 분기", "If 노드", "점수 >= 7 → 승인 / 4~6 → 수정 / <= 3 → 폐기"],
    ["반복 루프", "Loop 노드", "PL→QA 최대 3회 반복, 카운터로 관리"],
    ["데이터 전달", "워크플로우 간 Webhook 또는 Sub-workflow", "에이전트 간 기획안/피드백 데이터 전달"],
]
for i, row_data in enumerate(impl_data, 17):
    for j, val in enumerate(row_data, 1):
        ws8.cell(row=i, column=j, value=val)
    style_row(ws8, i, 3)

set_col_widths(ws8, [15, 25, 55, 10])

# ========== 저장 ==========
output_path = "/Users/gimdongseog/n8n-project/n8n_기획서_AI멀티에이전트.xlsx"
wb.save(output_path)
print(f"엑셀 파일 생성 완료: {output_path}")
