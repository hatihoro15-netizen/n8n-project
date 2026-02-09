#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_n8n_planning_document():
    """Create comprehensive n8n AI multi-agent system planning document"""

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")

    # Agent colors
    pm_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
    da_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Blue
    pl_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    qa_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
    pr_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
    dl_fill = PatternFill(start_color="C5A0D5", end_color="C5A0D5", fill_type="solid")  # Purple

    # Priority colors
    high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    med_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green

    # Status colors
    exist_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    new_fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")  # Light blue
    modify_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow

    normal_font = Font(name="맑은 고딕", size=10)
    bold_font = Font(name="맑은 고딕", size=10, bold=True)
    title_font = Font(name="맑은 고딕", size=14, bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ===== Sheet 1: 프로젝트 개요 =====
    ws1 = wb.create_sheet("프로젝트 개요")

    ws1['A1'] = "n8n AI 멀티 에이전트 시스템 구축 기획서"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:D1')

    ws1['A3'] = "프로젝트 목표"
    ws1['A3'].font = bold_font
    ws1['A3'].fill = header_fill
    ws1['A3'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws1.merge_cells('A3:D3')

    ws1['A4'] = "현재 상태: 클릭 1회로 영상 제작 자동화 완료"
    ws1.merge_cells('A4:D4')
    ws1['A5'] = "목표 상태: 클릭 0회 - 완전 자율 운영 AI 멀티 에이전트 시스템 구축"
    ws1.merge_cells('A5:D5')
    ws1['A6'] = "• 유튜브 분석 → 기획 → 검수 → 제작 → 배포 → 성과 추적을 AI가 자동으로 수행"
    ws1.merge_cells('A6:D6')
    ws1['A7'] = "• 사람의 개입 없이 콘텐츠가 자동으로 생성되고 업로드되는 시스템"
    ws1.merge_cells('A7:D7')

    ws1['A9'] = "사업 구조"
    ws1['A9'].font = bold_font
    ws1['A9'].fill = header_fill
    ws1['A9'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws1.merge_cells('A9:D9')

    ws1['A10'] = "마케팅 에이전시 운영 (유튜브 채널 관리)"
    ws1.merge_cells('A10:D10')
    ws1['A11'] = "• 여러 클라이언트의 유튜브 채널을 운영"
    ws1.merge_cells('A11:D11')
    ws1['A12'] = "• 각 채널은 별도의 유튜브 채널로 운영됨"
    ws1.merge_cells('A12:D12')

    ws1['A14'] = "운영 채널 (각각 별도 유튜브 채널)"
    ws1['A14'].font = bold_font
    ws1['A14'].fill = header_fill
    ws1['A14'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws1.merge_cells('A14:D14')

    channels = [
        ["채널명", "콘텐츠 타입", "주요 내용", "광고 정책"],
        ["1. 온카스터디 채널", "롱폼 + 숏폼", "커뮤니티 홍보", "자체 커뮤니티 홍보"],
        ["2. 슬롯 채널", "롱폼 + 숏폼", "슬롯 관련 콘텐츠", "온카스터디 광고만 (타 광고 없음)"],
        ["3. 루믹스 솔루션 채널", "롱폼 + 숏폼", "솔루션 판매/광고", "솔루션 판매 및 광고"],
        ["4. 쇼츠 전용 채널", "숏폼 전용", "다양한 소재", "마지막에 광고 삽입"]
    ]

    for i, row in enumerate(channels, start=15):
        for j, value in enumerate(row, start=1):
            cell = ws1.cell(row=i, column=j, value=value)
            cell.font = normal_font if i > 15 else bold_font
            cell.border = thin_border
            cell.alignment = center_align if i == 15 else left_align
            if i == 15:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    ws1['A21'] = "현재 워크플로우 현황"
    ws1['A21'].font = bold_font
    ws1['A21'].fill = header_fill
    ws1['A21'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws1.merge_cells('A21:D21')

    ws1['A22'] = "• 3개의 n8n 워크플로우 가동 중 (수동 트리거 또는 스케줄)"
    ws1.merge_cells('A22:D22')
    ws1['A23'] = "• Google Sheets 기반 콘텐츠 관리"
    ws1.merge_cells('A23:D23')
    ws1['A24'] = "• 다양한 AI 도구 활용: Gemini, ElevenLabs, Flux Pro, Recraft v3, Kling Video, ESRGAN, Beatoven, Creatomate"
    ws1.merge_cells('A24:D24')
    ws1['A25'] = "• YouTube API 자동 업로드 및 댓글 작성"
    ws1.merge_cells('A25:D25')

    # Set column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 35
    ws1.column_dimensions['D'].width = 30

    # Apply alignment to all cells
    for row in ws1.iter_rows(min_row=4, max_row=25):
        for cell in row:
            if cell.value:
                cell.font = normal_font
                cell.alignment = left_align

    # ===== Sheet 2: 현재 워크플로우 분석 =====
    ws2 = wb.create_sheet("현재 워크플로우 분석")

    ws2['A1'] = "현재 워크플로우 상세 분석"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:E1')

    # Workflow 1
    ws2['A3'] = "워크플로우 1: 전용 (루믹스/LUMIX) - 5장면 쇼츠 제작"
    ws2['A3'].font = bold_font
    ws2['A3'].fill = pr_fill
    ws2.merge_cells('A3:E3')

    workflow1_data = [
        ["단계", "도구/API", "세부 내용", "입력", "출력"],
        ["1. 트리거", "Manual Trigger", "수동 실행", "사용자 클릭", "워크플로우 시작"],
        ["2. 데이터 조회", "Google Sheets", "n8n LUMIX 시트에서 Status=준비인 행 조회", "시트 데이터", "주제, 나레이션"],
        ["3. 나레이션 분할", "Gemini 2.5 Flash", "나레이션을 5개 파트로 분할", "전체 나레이션", "5개 분할 텍스트"],
        ["4-A. TTS 생성", "fal.ai ElevenLabs turbo-v2.5", "한국어 음성 생성\n보이스: FQ3MuLxZh0jHcZmA5vW1", "5개 텍스트", "5개 음성 파일"],
        ["4-B. 이미지 프롬프트", "Gemini 3 Flash Preview", "5개 시네마틱 이미지 프롬프트 생성", "나레이션", "5개 프롬프트"],
        ["4-C. BGM 생성", "fal.ai Beatoven", "배경 음악 생성", "BGM 프롬프트", "음악 파일"],
        ["5. 이미지 생성", "fal.ai Flux Pro", "1080x1920 이미지 생성", "5개 프롬프트", "5개 이미지"],
        ["6. 이미지 업스케일", "ESRGAN", "2배 해상도 향상", "5개 이미지", "고해상도 이미지"],
        ["7. 비디오 생성", "fal.ai Kling Video v1.6", "Image-to-Video\n5초, 9:16 비율", "5개 이미지", "5개 비디오"],
        ["8. 영상 편집", "Creatomate", "템플릿: 056a9082\n5장면 + 나레이션 + 자막 + BGM + 8초 엔딩카드", "비디오+오디오", "완성 영상"],
        ["9. 유튜브 업로드", "YouTube API", "비공개(unlisted) 업로드 + 자동 댓글", "완성 영상", "유튜브 URL"],
        ["10. 시트 업데이트", "Google Sheets", "Status 업데이트, URL 저장\n'루믹스 수동편집' 시트에 저장", "URL, 메타데이터", "시트 갱신"]
    ]

    start_row = 4
    for i, row_data in enumerate(workflow1_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Workflow 2
    ws2['A18'] = "워크플로우 2: 루믹스 솔루션 영상 - 1장면 쇼츠"
    ws2['A18'].font = bold_font
    ws2['A18'].fill = pr_fill
    ws2.merge_cells('A18:E18')

    workflow2_data = [
        ["단계", "도구/API", "세부 내용", "입력", "출력"],
        ["1. 트리거", "Manual Trigger", "수동 실행", "사용자 클릭", "워크플로우 시작"],
        ["2. 데이터 조회", "Google Sheets", "시트에서 데이터 조회", "시트 데이터", "주제, 나레이션"],
        ["3. TTS 생성", "fal.ai ElevenLabs", "음성 생성", "텍스트", "음성 파일"],
        ["4. 이미지 생성", "fal.ai Recraft v3", "스타일: neon_calm\n비율: portrait_16_9", "프롬프트", "이미지"],
        ["5. 비디오 생성", "fal.ai Kling Video v1.6", "Image-to-Video", "이미지", "비디오"],
        ["6. 효과음 생성", "Gemini → ElevenLabs", "SFX 프롬프트 생성 후 음향 효과 생성", "콘텐츠", "효과음"],
        ["7. 영상 편집", "Creatomate", "템플릿: e961afae", "비디오+오디오+SFX", "완성 영상"],
        ["8. 유튜브 업로드", "YouTube API", "업로드 + 자동 댓글", "완성 영상", "유튜브 URL"]
    ]

    start_row = 19
    for i, row_data in enumerate(workflow2_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Workflow 3
    ws2['A29'] = "워크플로우 3: My workflow 2 - 쇼츠 자동 (구형)"
    ws2['A29'].font = bold_font
    ws2['A29'].fill = pr_fill
    ws2.merge_cells('A29:E29')

    workflow3_data = [
        ["단계", "도구/API", "세부 내용", "입력", "출력"],
        ["1. 트리거", "Schedule Trigger", "매일 오전 9시 자동 실행", "스케줄", "워크플로우 시작"],
        ["2. 데이터 조회", "Google Sheets", "'shorts 자동' 스프레드시트 조회", "시트 데이터", "콘텐츠 정보"],
        ["3. 이미지 생성", "Replicate (Flux)", "9:16 비율", "프롬프트", "이미지"],
        ["4. TTS 생성", "ElevenLabs API", "보이스: jB1Cifc2UQbq1gR3wnb0", "텍스트", "음성 파일"],
        ["5. 오디오 저장", "Google Drive", "공개 공유로 업로드", "음성 파일", "공유 URL"],
        ["6. 영상 편집", "Creatomate", "템플릿: e6a55264\n5개 이미지 + 오디오", "이미지+오디오", "완성 영상"],
        ["7. 업로드", "(미구현)", "YouTube 업로드 필요", "완성 영상", "유튜브 URL"]
    ]

    start_row = 30
    for i, row_data in enumerate(workflow3_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Google Sheets structure
    ws2['A39'] = "Google Sheets 데이터 구조"
    ws2['A39'].font = bold_font
    ws2['A39'].fill = header_fill
    ws2['A39'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws2.merge_cells('A39:E39')

    sheets_structure = [
        ["시트명", "주요 컬럼", "용도"],
        ["n8n LUMIX", "Subject, Narration, Status, Caption, BGM_prompt, Publish_date, YouTube_URL 등", "루믹스 5장면 쇼츠 관리"],
        ["루믹스 수동편집", "제작된 영상의 URL 및 메타데이터 저장", "완료된 영상 보관"],
        ["shorts 자동", "콘텐츠 정보, 스케줄 정보", "자동 쇼츠 제작용"]
    ]

    start_row = 40
    for i, row_data in enumerate(sheets_structure, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Set column widths
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 20

    # ===== Sheet 3: 에이전트 구성 =====
    ws3 = wb.create_sheet("에이전트 구성")

    ws3['A1'] = "AI 멀티 에이전트 시스템 구성"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:D1')

    ws3['A3'] = "6개 AI 에이전트 (부서)"
    ws3['A3'].font = bold_font
    ws3['A3'].fill = header_fill
    ws3['A3'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws3.merge_cells('A3:D3')

    agents = [
        ["에이전트", "역할", "주요 업무", "상태"],
        ["PM (총괄 매니저)", "파이프라인 전체 오케스트레이션", "• 전체 프로세스 조율\n• 에이전트 간 통신 관리\n• 일정 관리 및 모니터링", "신규 구축"],
        ["DA (데이터 분석)", "YouTube Analytics 분석", "• 조회수, 시청 지속 시간, CTR 분석\n• 트렌드 분석\n• 분석 리포트 생성", "신규 구축"],
        ["PL (기획)", "콘텐츠 기획", "• DA 리포트 기반 2-3개 기획안 생성\n• 주제, 나레이션, 프롬프트 작성\n• QA 피드백 반영", "신규 구축"],
        ["QA (검수)", "기획안 검수", "• 기획안 1-10점 평가\n• 구체적 피드백 제공\n• PL과 최대 3회까지 피드백 루프", "신규 구축"],
        ["PR (제작)", "영상 제작", "• 기존 워크플로우 활용\n• TTS, 이미지, 영상, 편집 자동화\n• 고품질 콘텐츠 생성", "기존 있음"],
        ["DL (배포)", "유튜브 업로드 및 관리", "• 기존 워크플로우 활용\n• YouTube API 업로드\n• 자동 댓글, 메타데이터 관리", "기존 있음"]
    ]

    start_row = 4
    for i, row_data in enumerate(agents, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif i > start_row:
                # Color code by agent
                agent_fills = [pm_fill, da_fill, pl_fill, qa_fill, pr_fill, dl_fill]
                if i - start_row - 1 < len(agent_fills):
                    cell.fill = agent_fills[i - start_row - 1]

    ws3['A12'] = "현재 vs 목표 비교"
    ws3['A12'].font = bold_font
    ws3['A12'].fill = header_fill
    ws3['A12'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws3.merge_cells('A12:D12')

    comparison = [
        ["항목", "현재 상태", "목표 상태", "개선 사항"],
        ["트리거", "수동 클릭 또는 스케줄", "완전 자동 (YouTube 분석 기반)", "사람 개입 불필요"],
        ["콘텐츠 기획", "사람이 Google Sheets에 입력", "AI가 자동 기획 (DA 분석 기반)", "트렌드 기반 자동 기획"],
        ["품질 검수", "없음 또는 수동", "QA 에이전트 자동 검수", "일관된 품질 보장"],
        ["제작", "자동화 완료", "자동화 완료 (유지)", "기존 시스템 활용"],
        ["배포", "자동화 완료", "자동화 완료 (유지)", "기존 시스템 활용"],
        ["성과 분석", "없음", "DA 에이전트 자동 분석", "데이터 기반 의사결정"],
        ["피드백 루프", "없음", "PL-QA 자동 피드백", "자가 개선 시스템"]
    ]

    start_row = 13
    for i, row_data in enumerate(comparison, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Set column widths
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 35
    ws3.column_dimensions['D'].width = 30

    # ===== Sheet 4: 워크플로우 목록 =====
    ws4 = wb.create_sheet("워크플로우 목록")

    ws4['A1'] = "전체 워크플로우 목록"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:F1')

    workflows = [
        ["워크플로우명", "담당 에이전트", "주요 기능", "상태", "우선순위", "비고"],
        ["전용 (루믹스/LUMIX)", "PR", "5장면 쇼츠 제작 (Flux Pro + Kling)", "기존 있음", "높음", "현재 운영 중"],
        ["루믹스 솔루션 영상", "PR", "1장면 쇼츠 제작 (Recraft v3 + SFX)", "기존 있음", "높음", "현재 운영 중"],
        ["My workflow 2", "PR", "자동 쇼츠 제작 (Replicate)", "기존 있음", "중간", "스케줄 실행 중"],
        ["YouTube Analytics 수집", "DA", "채널별 성과 데이터 수집 및 저장", "신규", "높음", "Phase 2"],
        ["YouTube Analytics 분석", "DA", "데이터 분석 및 인사이트 도출", "신규", "높음", "Phase 2"],
        ["DA 리포트 생성", "DA", "분석 결과를 리포트로 작성", "신규", "높음", "Phase 2"],
        ["콘텐츠 기획 생성", "PL", "DA 리포트 기반 2-3개 기획안 작성", "신규", "높음", "Phase 3"],
        ["기획안 검수", "QA", "기획안 평가 및 피드백 생성", "신규", "높음", "Phase 3"],
        ["PL-QA 피드백 루프", "PL + QA", "최대 3회 반복 개선", "신규", "높음", "Phase 3"],
        ["Google Sheets 기획 저장", "PL", "승인된 기획안을 Sheets에 저장", "신규", "중간", "Phase 3"],
        ["기존 PR 워크플로우 트리거", "PM", "기획안 → 제작 워크플로우 자동 실행", "수정 필요", "높음", "Phase 4"],
        ["PM 오케스트레이터", "PM", "전체 파이프라인 조율 및 모니터링", "신규", "높음", "Phase 4"],
        ["성과 추적 및 학습", "DA + PM", "A/B 테스팅, 자동 학습", "신규", "낮음", "Phase 5"]
    ]

    start_row = 3
    for i, row_data in enumerate(workflows, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws4.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif i > start_row:
                # Color code by agent
                agent_name = row_data[1]
                if "PM" in agent_name:
                    cell.fill = pm_fill
                elif "DA" in agent_name:
                    cell.fill = da_fill
                elif "PL" in agent_name:
                    cell.fill = pl_fill
                elif "QA" in agent_name:
                    cell.fill = qa_fill
                elif "PR" in agent_name:
                    cell.fill = pr_fill
                elif "DL" in agent_name:
                    cell.fill = dl_fill

                # Status color
                status = row_data[3]
                status_cell = ws4.cell(row=i, column=4)
                if status == "기존 있음":
                    status_cell.fill = exist_fill
                elif status == "신규":
                    status_cell.fill = new_fill
                elif status == "수정 필요":
                    status_cell.fill = modify_fill

                # Priority color
                priority = row_data[4]
                priority_cell = ws4.cell(row=i, column=5)
                if priority == "높음":
                    priority_cell.fill = high_fill
                    priority_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
                elif priority == "중간":
                    priority_cell.fill = med_fill
                elif priority == "낮음":
                    priority_cell.fill = low_fill

    # Set column widths
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 40
    ws4.column_dimensions['D'].width = 15
    ws4.column_dimensions['E'].width = 12
    ws4.column_dimensions['F'].width = 20

    # ===== Sheet 5: 콘텐츠 전략 =====
    ws5 = wb.create_sheet("콘텐츠 전략")

    ws5['A1'] = "채널별 콘텐츠 전략"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:E1')

    ws5['A3'] = "채널별 매트릭스"
    ws5['A3'].font = bold_font
    ws5['A3'].fill = header_fill
    ws5['A3'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws5.merge_cells('A3:E3')

    channel_matrix = [
        ["채널명", "콘텐츠 타입", "제작 주기", "광고 삽입 규칙", "사용 워크플로우"],
        ["온카스터디 채널", "롱폼 + 숏폼", "주 3-5회", "커뮤니티 홍보 (자체)", "전용 (루믹스), 루믹스 솔루션"],
        ["슬롯 채널", "롱폼 + 숏폼", "주 3-5회", "온카스터디 광고만\n(타 광고 절대 금지)", "전용 (루믹스), 루믹스 솔루션"],
        ["루믹스 솔루션 채널", "롱폼 + 숏폼", "주 5-7회", "솔루션 판매 및 광고", "루믹스 솔루션 영상 전용"],
        ["쇼츠 전용 채널", "숏폼 전용", "일 1-2회", "마지막 8초 엔딩카드에 광고", "My workflow 2 + 전용"]
    ]

    start_row = 4
    for i, row_data in enumerate(channel_matrix, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws5.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    ws5['A10'] = "쇼츠 자동화 전략"
    ws5['A10'].font = bold_font
    ws5['A10'].fill = header_fill
    ws5['A10'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws5.merge_cells('A10:E10')

    ws5['A11'] = "1단계: 분석 (DA)"
    ws5['A11'].font = bold_font
    ws5.merge_cells('A11:E11')
    ws5['A12'] = "• YouTube Analytics API로 채널별 성과 수집 (일 1회)"
    ws5.merge_cells('A12:E12')
    ws5['A13'] = "• 고성과 콘텐츠 패턴 분석 (주제, 길이, 스타일, 시간대)"
    ws5.merge_cells('A13:E13')
    ws5['A14'] = "• 트렌드 키워드 분석"
    ws5.merge_cells('A14:E14')

    ws5['A16'] = "2단계: 기획 (PL)"
    ws5['A16'].font = bold_font
    ws5.merge_cells('A16:E16')
    ws5['A17'] = "• DA 리포트 기반 2-3개 기획안 생성"
    ws5.merge_cells('A17:E17')
    ws5['A18'] = "• 각 기획안: 주제, 나레이션, 이미지 프롬프트, BGM 프롬프트, 광고 위치"
    ws5.merge_cells('A18:E18')
    ws5['A19'] = "• 채널별 광고 정책 준수"
    ws5.merge_cells('A19:E19')

    ws5['A21'] = "3단계: 검수 (QA)"
    ws5['A21'].font = bold_font
    ws5.merge_cells('A21:E21')
    ws5['A22'] = "• 각 기획안을 1-10점으로 평가"
    ws5.merge_cells('A22:E22')
    ws5['A23'] = "• 평가 기준: 트렌드 적합성, 채널 정체성, 광고 정책 준수, 제작 가능성"
    ws5.merge_cells('A23:E23')
    ws5['A24'] = "• 8점 미만: PL에게 피드백 (최대 3회 반복)"
    ws5.merge_cells('A24:E24')
    ws5['A25'] = "• 8점 이상: 승인 → 제작 단계로"
    ws5.merge_cells('A25:E25')

    ws5['A27'] = "4단계: 제작 (PR)"
    ws5['A27'].font = bold_font
    ws5.merge_cells('A27:E27')
    ws5['A28'] = "• 승인된 기획안을 Google Sheets에 저장"
    ws5.merge_cells('A28:E28')
    ws5['A29'] = "• 기존 워크플로우 자동 트리거"
    ws5.merge_cells('A29:E29')
    ws5['A30'] = "• TTS → 이미지 → 비디오 → 편집 자동 실행"
    ws5.merge_cells('A30:E30')

    ws5['A32'] = "5단계: 배포 (DL)"
    ws5['A32'].font = bold_font
    ws5.merge_cells('A32:E32')
    ws5['A33'] = "• YouTube API 자동 업로드"
    ws5.merge_cells('A33:E33')
    ws5['A34'] = "• 최적 시간대 업로드 (DA 분석 기반)"
    ws5.merge_cells('A34:E34')
    ws5['A35'] = "• 자동 댓글, 설명, 태그 추가"
    ws5.merge_cells('A35:E35')

    ws5['A37'] = "6단계: 추적 (DA)"
    ws5['A37'].font = bold_font
    ws5.merge_cells('A37:E37')
    ws5['A38'] = "• 업로드 후 24시간, 7일, 30일 성과 추적"
    ws5.merge_cells('A38:E38')
    ws5['A39'] = "• 성과 데이터를 다음 기획에 반영"
    ws5.merge_cells('A39:E39')
    ws5['A40'] = "• 자동 학습 시스템 (Phase 5)"
    ws5.merge_cells('A40:E40')

    # Set column widths
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 15
    ws5.column_dimensions['D'].width = 35
    ws5.column_dimensions['E'].width = 30

    # Apply alignment
    for row in ws5.iter_rows(min_row=11, max_row=40):
        for cell in row:
            if cell.value:
                cell.font = normal_font
                cell.alignment = left_align

    # ===== Sheet 6: 구축 로드맵 =====
    ws6 = wb.create_sheet("구축 로드맵")

    ws6['A1'] = "AI 멀티 에이전트 시스템 구축 로드맵"
    ws6['A1'].font = title_font
    ws6.merge_cells('A1:E1')

    roadmap = [
        ["단계", "기간", "주요 작업", "산출물", "우선순위"],
        ["Phase 1:\n서버 구축", "1주", "• Hostinger VPS 설정\n• n8n 인스턴스 구성\n• 데이터베이스 설정\n• 환경 변수 및 보안 설정", "• 운영 서버 환경\n• n8n 접근 가능\n• DB 스키마", "높음"],
        ["Phase 2:\nDA 에이전트", "2주", "• YouTube Analytics API 연동\n• 데이터 수집 워크플로우\n• 데이터 분석 로직 구현\n• 리포트 생성 워크플로우", "• 분석 리포트 자동 생성\n• 트렌드 인사이트\n• 채널별 성과 대시보드", "높음"],
        ["Phase 3:\nPL + QA 에이전트", "2-3주", "• 기획 생성 워크플로우 (Gemini)\n• QA 평가 워크플로우 (Gemini)\n• 피드백 루프 구현 (Loop/If 노드)\n• Google Sheets 연동", "• 자동 기획안 생성\n• 자동 검수 시스템\n• 승인된 기획 저장", "높음"],
        ["Phase 4:\n통합 및 PM", "2주", "• 기존 PR/DL 워크플로우 연결\n• PM 오케스트레이터 구현\n• 전체 파이프라인 자동화\n• 에러 핸들링 및 모니터링", "• 완전 자동 파이프라인\n• 클릭 0회 달성\n• 모니터링 대시보드", "높음"],
        ["Phase 5:\n고급 기능", "4주+", "• A/B 테스팅 시스템\n• 자동 학습 및 최적화\n• 다채널 확장\n• 성과 예측 모델", "• 자가 학습 시스템\n• 예측 분석\n• 확장 가능한 아키텍처", "낮음"]
    ]

    start_row = 3
    for i, row_data in enumerate(roadmap, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws6.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif i > start_row:
                # Priority color
                priority = row_data[4]
                priority_cell = ws6.cell(row=i, column=5)
                if priority == "높음":
                    priority_cell.fill = high_fill
                    priority_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
                elif priority == "낮음":
                    priority_cell.fill = low_fill

    # Set row heights for readability
    for row in range(4, 9):
        ws6.row_dimensions[row].height = 100

    ws6['A10'] = "총 예상 기간: 7-10주 (Phase 1-4), Phase 5는 선택적"
    ws6['A10'].font = bold_font
    ws6.merge_cells('A10:E10')

    ws6['A12'] = "주요 마일스톤"
    ws6['A12'].font = bold_font
    ws6['A12'].fill = header_fill
    ws6['A12'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws6.merge_cells('A12:E12')

    milestones = [
        ["마일스톤", "완료 조건", "검증 방법"],
        ["서버 준비 완료", "n8n 접속 가능, DB 연결 확인", "테스트 워크플로우 실행"],
        ["DA 에이전트 가동", "YouTube 데이터 수집 및 리포트 생성", "리포트 품질 검증"],
        ["PL+QA 에이전트 가동", "자동 기획 및 검수 완료", "승인률 80% 이상"],
        ["클릭 0회 달성", "전체 파이프라인 자동 실행", "7일간 무개입 테스트 성공"],
        ["고급 기능 구현", "A/B 테스팅 및 학습 시스템 작동", "성과 개선 확인"]
    ]

    start_row = 13
    for i, row_data in enumerate(milestones, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws6.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Set column widths
    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 12
    ws6.column_dimensions['C'].width = 45
    ws6.column_dimensions['D'].width = 35
    ws6.column_dimensions['E'].width = 12

    # ===== Sheet 7: 필요 도구 및 API =====
    ws7 = wb.create_sheet("필요 도구 및 API")

    ws7['A1'] = "필요 도구 및 API 목록"
    ws7['A1'].font = title_font
    ws7.merge_cells('A1:F1')

    ws7['A3'] = "현재 사용 중인 도구"
    ws7['A3'].font = bold_font
    ws7['A3'].fill = exist_fill
    ws7.merge_cells('A3:F3')

    existing_tools = [
        ["도구/API", "용도", "사용 워크플로우", "비용", "상태", "비고"],
        ["Google Sheets API", "콘텐츠 관리", "모든 워크플로우", "무료", "사용 중", "필수"],
        ["Gemini 2.5 Flash", "나레이션 분할", "전용 (루믹스)", "종량제", "사용 중", ""],
        ["Gemini 3 Flash Preview", "이미지 프롬프트 생성", "전용 (루믹스)", "종량제", "사용 중", ""],
        ["Gemini", "SFX 프롬프트, 기획", "루믹스 솔루션, 신규", "종량제", "사용 중", ""],
        ["fal.ai ElevenLabs", "TTS 음성 생성", "전용, 루믹스 솔루션", "종량제", "사용 중", "한국어 지원"],
        ["ElevenLabs API", "TTS, 효과음", "My workflow 2, 루믹스 솔루션", "종량제", "사용 중", "직접 API"],
        ["fal.ai Flux Pro", "이미지 생성 (1080x1920)", "전용 (루믹스)", "종량제", "사용 중", "고품질"],
        ["fal.ai Recraft v3", "이미지 생성 (neon_calm)", "루믹스 솔루션", "종량제", "사용 중", "스타일 특화"],
        ["Replicate (Flux)", "이미지 생성 (9:16)", "My workflow 2", "종량제", "사용 중", ""],
        ["ESRGAN", "이미지 업스케일 (2x)", "전용 (루믹스)", "종량제", "사용 중", ""],
        ["fal.ai Beatoven", "BGM 생성", "전용 (루믹스)", "종량제", "사용 중", ""],
        ["fal.ai Kling Video v1.6", "Image-to-Video (5s, 9:16)", "전용, 루믹스 솔루션", "종량제", "사용 중", "고품질 영상"],
        ["Creatomate", "영상 편집", "모든 워크플로우", "종량제", "사용 중", "템플릿 다수"],
        ["YouTube API", "업로드, 댓글", "전용, 루믹스 솔루션", "무료", "사용 중", "필수"],
        ["Google Drive API", "오디오 저장", "My workflow 2", "무료", "사용 중", ""]
    ]

    start_row = 4
    for i, row_data in enumerate(existing_tools, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws7.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    ws7['A21'] = "신규 필요 도구"
    ws7['A21'].font = bold_font
    ws7['A21'].fill = new_fill
    ws7.merge_cells('A21:F21')

    new_tools = [
        ["도구/API", "용도", "담당 에이전트", "예상 비용", "우선순위", "비고"],
        ["Hostinger VPS", "n8n 서버 호스팅", "인프라", "$20-50/월", "높음", "Phase 1"],
        ["YouTube Analytics API", "채널 성과 데이터 수집", "DA", "무료 (할당량)", "높음", "Phase 2 필수"],
        ["YouTube Data API v3", "영상 메타데이터 조회", "DA", "무료 (할당량)", "높음", "Phase 2"],
        ["PostgreSQL/MongoDB", "분석 데이터 저장", "DA", "무료 (VPS 포함)", "중간", "Phase 2"],
        ["n8n Webhook", "에이전트 간 통신", "PM", "무료", "높음", "Phase 4"],
        ["n8n Loop Node", "PL-QA 피드백 반복", "PL+QA", "무료", "높음", "Phase 3"],
        ["n8n Schedule Trigger", "자동 실행", "PM", "무료", "높음", "Phase 4"]
    ]

    start_row = 22
    for i, row_data in enumerate(new_tools, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws7.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif i > start_row:
                # Priority color
                priority = row_data[4]
                priority_cell = ws7.cell(row=i, column=5)
                if priority == "높음":
                    priority_cell.fill = high_fill
                    priority_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
                elif priority == "중간":
                    priority_cell.fill = med_fill

    ws7['A31'] = "예상 월간 비용"
    ws7['A31'].font = bold_font
    ws7['A31'].fill = header_fill
    ws7['A31'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws7.merge_cells('A31:F31')

    cost_estimate = [
        ["항목", "예상 비용", "비고"],
        ["Hostinger VPS", "$20-50/월", "서버 호스팅"],
        ["Gemini API", "$50-100/월", "기획, 분석, 프롬프트 생성"],
        ["fal.ai (TTS, 이미지, 비디오)", "$200-400/월", "영상당 비용, 제작량에 따라 변동"],
        ["ElevenLabs API", "$50-100/월", "TTS 및 효과음"],
        ["Creatomate", "$100-200/월", "영상 편집"],
        ["기타 API", "$20-50/월", "Replicate 등"],
        ["총합", "$440-900/월", "제작량에 따라 변동"]
    ]

    start_row = 32
    for i, row_data in enumerate(cost_estimate, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws7.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row or i == start_row + 7 else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif i == start_row + 7:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # Set column widths
    ws7.column_dimensions['A'].width = 30
    ws7.column_dimensions['B'].width = 30
    ws7.column_dimensions['C'].width = 25
    ws7.column_dimensions['D'].width = 20
    ws7.column_dimensions['E'].width = 12
    ws7.column_dimensions['F'].width = 25

    # ===== Sheet 8: 피드백 루프 상세 =====
    ws8 = wb.create_sheet("피드백 루프 상세")

    ws8['A1'] = "PL ↔ QA 피드백 루프 상세 설계"
    ws8['A1'].font = title_font
    ws8.merge_cells('A1:D1')

    ws8['A3'] = "프로세스 흐름"
    ws8['A3'].font = bold_font
    ws8['A3'].fill = header_fill
    ws8['A3'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws8.merge_cells('A3:D3')

    process_flow = [
        ["단계", "담당", "작업 내용", "출력"],
        ["1. DA 리포트 입력", "DA", "YouTube 분석 리포트 생성", "트렌드 인사이트, 고성과 패턴"],
        ["2. 초기 기획", "PL", "2-3개 기획안 생성 (Gemini)", "기획안 A, B, C"],
        ["3. 검수 (1차)", "QA", "각 기획안 평가 (1-10점) + 피드백", "점수, 구체적 피드백"],
        ["4. 판단", "n8n If Node", "최고 점수 ≥ 8?", "YES → 승인 / NO → 다음"],
        ["5. 피드백 반영", "PL", "QA 피드백 기반 개선 (Gemini)", "개선된 기획안"],
        ["6. 검수 (2차)", "QA", "재평가", "점수, 피드백"],
        ["7. 판단", "n8n If Node", "최고 점수 ≥ 8 OR 반복 횟수 = 3?", "YES → 승인 / NO → 반복"],
        ["8. 최종 승인", "PM", "최고 점수 기획안 선택", "Google Sheets 저장"],
        ["9. 제작 트리거", "PM", "PR 워크플로우 실행", "영상 제작 시작"]
    ]

    start_row = 4
    for i, row_data in enumerate(process_flow, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws8.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif i > start_row:
                # Color by agent
                agent = row_data[1]
                if "DA" in agent:
                    cell.fill = da_fill
                elif "PL" in agent:
                    cell.fill = pl_fill
                elif "QA" in agent:
                    cell.fill = qa_fill
                elif "PM" in agent:
                    cell.fill = pm_fill

    ws8['A15'] = "QA 평가 기준"
    ws8['A15'].font = bold_font
    ws8['A15'].fill = header_fill
    ws8['A15'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws8.merge_cells('A15:D15')

    qa_criteria = [
        ["평가 항목", "배점", "세부 기준", "예시"],
        ["트렌드 적합성", "3점", "DA 리포트의 인사이트 반영 여부\n현재 트렌드 키워드 활용", "고성과 주제 반영 시 만점"],
        ["채널 정체성", "2점", "채널의 콘텐츠 방향성 일치\n기존 성공 콘텐츠와의 일관성", "채널 톤앤매너 일치 시 만점"],
        ["광고 정책 준수", "2점", "채널별 광고 삽입 규칙 준수\n슬롯 채널: 온카스터디만", "정책 위반 시 0점"],
        ["제작 가능성", "2점", "현재 워크플로우로 제작 가능 여부\n리소스 및 시간 현실성", "제작 불가 시 0점"],
        ["창의성 및 차별성", "1점", "기존 콘텐츠와 차별화\n독창적 아이디어", "혁신적 아이디어 시 만점"]
    ]

    start_row = 16
    for i, row_data in enumerate(qa_criteria, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws8.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    ws8['A23'] = "n8n 구현 방법"
    ws8['A23'].font = bold_font
    ws8['A23'].fill = header_fill
    ws8['A23'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws8.merge_cells('A23:D23')

    implementation = [
        ["노드 타입", "용도", "설정", "비고"],
        ["Loop Over Items", "최대 3회 반복", "Max Iterations: 3", "PL-QA 루프"],
        ["HTTP Request (Gemini)", "PL 기획 생성", "Prompt: DA 리포트 + 이전 피드백", "Gemini API 호출"],
        ["HTTP Request (Gemini)", "QA 평가", "Prompt: 기획안 + 평가 기준", "점수 및 피드백"],
        ["If Node", "점수 판단", "Condition: 최고점수 >= 8", "승인/거부 분기"],
        ["Set Variable", "반복 카운터", "Iteration count 증가", "Loop 제어"],
        ["If Node", "최대 반복 체크", "Condition: Iteration >= 3", "강제 종료 조건"],
        ["Code Node", "최고 점수 선택", "JavaScript로 최고 점수 기획안 선택", "승인 시"],
        ["Google Sheets", "기획안 저장", "승인된 기획안 저장", "제작 대기 상태로"]
    ]

    start_row = 24
    for i, row_data in enumerate(implementation, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws8.cell(row=i, column=j, value=value)
            cell.font = bold_font if i == start_row else normal_font
            cell.border = thin_border
            cell.alignment = center_align if i == start_row else left_align
            if i == start_row:
                cell.fill = header_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    ws8['A34'] = "피드백 예시"
    ws8['A34'].font = bold_font
    ws8['A34'].fill = header_fill
    ws8['A34'].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    ws8.merge_cells('A34:D34')

    ws8['A35'] = "낮은 점수 (6점) 피드백:"
    ws8['A35'].font = bold_font
    ws8.merge_cells('A35:D35')
    ws8['A36'] = "\"트렌드 적합성: 2/3 - DA 리포트에서 '슬롯 전략'이 고성과 키워드였으나 기획안에 반영되지 않음\n채널 정체성: 2/2 - 슬롯 채널 톤 일치\n광고 정책: 0/2 - 외부 광고 언급, 온카스터디 광고만 허용\n제작 가능성: 2/2 - 제작 가능\n창의성: 0/1 - 기존 콘텐츠와 유사\n\n개선 제안: 1) '슬롯 전략' 키워드 포함, 2) 광고는 온카스터디만, 3) 차별화된 접근 필요\""
    ws8.merge_cells('A36:D36')
    ws8['A36'].alignment = left_align

    ws8['A38'] = "높은 점수 (9점) 피드백:"
    ws8['A38'].font = bold_font
    ws8.merge_cells('A38:D38')
    ws8['A39'] = "\"트렌드 적합성: 3/3 - DA 리포트의 '슬롯 전략' 키워드 완벽 반영\n채널 정체성: 2/2 - 슬롯 채널 톤 완벽 일치\n광고 정책: 2/2 - 온카스터디 광고만 삽입\n제작 가능성: 2/2 - 제작 가능\n창의성: 0/1 - 좋은 아이디어이나 혁신적이지는 않음\n\n승인합니다. 제작 진행하세요.\""
    ws8.merge_cells('A39:D39')
    ws8['A39'].alignment = left_align

    # Set column widths
    ws8.column_dimensions['A'].width = 25
    ws8.column_dimensions['B'].width = 20
    ws8.column_dimensions['C'].width = 45
    ws8.column_dimensions['D'].width = 30

    # Set row heights for feedback examples
    ws8.row_dimensions[36].height = 100
    ws8.row_dimensions[39].height = 80

    # Apply normal font to feedback cells
    ws8['A36'].font = normal_font
    ws8['A39'].font = normal_font

    # Save workbook
    output_file = "/Users/gimdongseog/n8n-project/n8n_기획서_v2.xlsx"
    wb.save(output_file)
    print(f"✅ Excel file created successfully: {output_file}")

if __name__ == "__main__":
    create_n8n_planning_document()
