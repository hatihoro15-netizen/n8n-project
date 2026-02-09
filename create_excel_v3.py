#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def set_cell_style(cell, bg_color=None, font_color="000000", bold=False, align_h="left", align_v="top", wrap=True):
    """셀 스타일 설정 헬퍼 함수"""
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.font = Font(name="맑은 고딕", size=10, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    cell.border = thin_border

def create_header(ws, row, columns, bg_color="2F5496", font_color="FFFFFF"):
    """헤더 행 생성"""
    for col_idx, header_text in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        set_cell_style(cell, bg_color=bg_color, font_color=font_color, bold=True, align_h="center", align_v="center")

def auto_adjust_column_width(ws, column, width):
    """컬럼 너비 조정"""
    ws.column_dimensions[get_column_letter(column)].width = width

# 워크북 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# ========== Sheet 1: 프로젝트 개요 ==========
ws1 = wb.create_sheet("프로젝트 개요")

# 제목
ws1.merge_cells('A1:F1')
cell = ws1['A1']
cell.value = "n8n AI 멀티 에이전트 영상 제작 시스템 — 최종 기획서"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center", align_v="center")
cell.font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
ws1.row_dimensions[1].height = 30

# 개요 테이블
row = 3
overview_data = [
    ("목적", "4개 YouTube 채널에 대한 완전 자동화된 영상 제작 및 배포 시스템 구축"),
    ("핵심개념", "6개 AI 에이전트(PM, DA, PL, QA, PR, DL)가 협업하여 기획-제작-검수-배포 전 과정 자동화"),
    ("최종목표", "클릭 0번으로 매일 자동으로 채널별 최적화된 영상이 업로드되고 성과 분석까지 완료"),
    ("현재상태", "숏폼 자동화 완료 (클릭 1번) - '전용' 워크플로우로 5씬 쇼츠 자동 생성"),
    ("다음단계", "다중 채널 확장 + AI 기획/검수 에이전트 추가 → 클릭 0번 달성")
]

create_header(ws1, row, ["항목", "내용"], bg_color="2F5496")
row += 1
for item, content in overview_data:
    ws1.cell(row=row, column=1, value=item)
    ws1.cell(row=row, column=2, value=content)
    set_cell_style(ws1.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws1.cell(row=row, column=2))
    ws1.merge_cells(f'B{row}:F{row}')
    row += 1

# 4개 채널
row += 2
ws1.merge_cells(f'A{row}:F{row}')
cell = ws1.cell(row=row, column=1, value="4개 YouTube 채널 구성")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

channels = [
    ("채널명", "콘텐츠 유형", "목적"),
    ("온카스터디 채널", "롱폼 + 숏폼", "커뮤니티 홍보, 교육 콘텐츠, 다양한 주제"),
    ("슬롯 채널", "롱폼 + 숏폼", "슬롯 전문 콘텐츠 + 온카스터디 광고만 (다른 광고 없음)"),
    ("루믹스 솔루션 채널", "롱폼 + 숏폼", "솔루션 판매/광고, 제품 데모, 비즈니스 콘텐츠"),
    ("쇼츠 전용 채널", "숏폼 only", "다양한 주제 + 끝에 광고, 바이럴 콘텐츠")
]

create_header(ws1, row, channels[0], bg_color="2F5496")
row += 1
for channel_data in channels[1:]:
    for col_idx, val in enumerate(channel_data, 1):
        cell = ws1.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell)
    row += 1

# 비즈니스 구조
row += 2
ws1.merge_cells(f'A{row}:F{row}')
cell = ws1.cell(row=row, column=1, value="비즈니스 구조")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1
ws1.merge_cells(f'A{row}:F{row}')
cell = ws1.cell(row=row, column=1, value="핵심 = 온카스터디와 루믹스 솔루션을 알리고 판매하는 것")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")

# 컬럼 너비
auto_adjust_column_width(ws1, 1, 20)
auto_adjust_column_width(ws1, 2, 50)
auto_adjust_column_width(ws1, 3, 25)

# ========== Sheet 2: 현재 워크플로우 분석 ==========
ws2 = wb.create_sheet("현재 워크플로우 분석")

# 제목
ws2.merge_cells('A1:E1')
cell = ws2['A1']
cell.value = "현재 '전용' 워크플로우 분석 (5-Scene Shorts)"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws2.row_dimensions[1].height = 25

# 워크플로우 단계
row = 3
workflow_steps = [
    ("단계", "노드/프로세스", "사용 도구/API", "입력", "출력"),
    ("1", "Manual Trigger → Google Sheets 읽기", "Google Sheets API", "수동 트리거", "Subject, Narration, Status=준비 행 데이터"),
    ("2", "나레이션 5개로 분할", "Google Gemini 2.5 Flash", "전체 나레이션", "5개 분할된 나레이션 (배열)"),
    ("3a", "TTS 생성 (병렬 처리)", "fal.ai ElevenLabs TTS turbo-v2.5", "각 나레이션 부분 (5개)", "5개 음성 파일 URL"),
    ("3b-1", "이미지 프롬프트 생성", "Google Gemini 3 Flash Preview", "전체 나레이션 (현재 문제: 분할 전 사용)", "5개 이미지 프롬프트"),
    ("3b-2", "이미지 생성", "fal.ai Flux Pro", "5개 프롬프트", "5개 이미지 URL"),
    ("3b-3", "이미지 업스케일", "fal.ai ESRGAN", "5개 이미지", "5개 고해상도 이미지 URL"),
    ("3b-4", "비디오 생성", "fal.ai Kling Video v1.6", "5개 이미지", "5개 비디오 URL"),
    ("3c", "BGM 생성 (병렬 처리)", "fal.ai Beatoven", "BGM 프롬프트", "BGM URL"),
    ("4", "최종 합성", "Creatomate", "5개 비디오 + 5개 음성 + BGM + 자막 + 8초 엔딩카드", "최종 쇼츠 영상 URL"),
    ("5a", "YouTube 업로드", "YouTube Data API v3", "영상 URL + 제목 + 설명 + 태그", "YouTube 영상 URL"),
    ("5b", "자동 댓글 작성", "YouTube Data API v3", "댓글 내용", "댓글 ID"),
    ("6", "Sheets 상태 업데이트", "Google Sheets API", "업로드 URL, 상태", "Status=생성완료, Publish=발행완료, 업로드 URL 저장")
]

create_header(ws2, row, workflow_steps[0], bg_color="2F5496")
row += 1
for step_data in workflow_steps[1:]:
    for col_idx, val in enumerate(step_data, 1):
        cell = ws2.cell(row=row, column=col_idx, value=val)
        # 문제 있는 단계 강조
        if "현재 문제" in str(val):
            set_cell_style(cell, bg_color="FFEB9C")
        else:
            set_cell_style(cell)
    row += 1

# 플로우 다이어그램
row += 2
ws2.merge_cells(f'A{row}:E{row}')
cell = ws2.cell(row=row, column=1, value="플로우 다이어그램")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

flow_diagram = [
    "📌 Manual Trigger",
    "  ↓",
    "📊 Google Sheets (Status=준비 행 읽기)",
    "  ↓",
    "🤖 Gemini 2.5 Flash (나레이션 → 5개 분할)",
    "  ↓",
    "┌─────────────────┼─────────────────┐",
    "│ 3a. TTS         │ 3b. 이미지/비디오  │ 3c. BGM",
    "│ ElevenLabs (5개) │ Gemini → Flux Pro│ Beatoven",
    "│                 │ → ESRGAN         │",
    "│                 │ → Kling (5개)    │",
    "└─────────────────┴─────────────────┘",
    "  ↓",
    "🎬 Creatomate (5씬 합성 + 자막 + 엔딩카드)",
    "  ↓",
    "📺 YouTube 업로드 + 댓글",
    "  ↓",
    "✅ Google Sheets 상태 업데이트"
]

for line in flow_diagram:
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2.cell(row=row, column=1, value=line)
    set_cell_style(cell, align_h="left")
    cell.font = Font(name="맑은 고딕", size=9)
    row += 1

# Google Sheets DB 구조
row += 2
ws2.merge_cells(f'A{row}:E{row}')
cell = ws2.cell(row=row, column=1, value="Google Sheets DB 구조 (현재)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

sheets_columns = [
    ("컬럼명", "설명", "예시 값"),
    ("Subject", "영상 제목", "도박 중독에서 벗어나는 3가지 방법"),
    ("Narration", "전체 나레이션 대본", "여러분, 혹시 도박 때문에 힘드신가요?..."),
    ("Status", "제작 상태", "준비 / 생성완료"),
    ("Caption", "YouTube 설명", "이 영상은..."),
    ("댓글", "자동 댓글 내용", "더 많은 정보는..."),
    ("Publish", "발행 상태", "대기중 / 발행완료"),
    ("BGM prompt", "BGM 생성 프롬프트", "hopeful, inspiring, gentle piano"),
    ("imagePrompt", "이미지 프롬프트 (사용 안 함)", ""),
    ("업로드 URL", "YouTube URL", "https://youtube.com/watch?v=..."),
    ("row_number", "행 번호", "2")
]

create_header(ws2, row, sheets_columns[0], bg_color="2F5496")
row += 1
for col_data in sheets_columns[1:]:
    for col_idx, val in enumerate(col_data, 1):
        cell = ws2.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell)
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws2, 1, 12)
auto_adjust_column_width(ws2, 2, 35)
auto_adjust_column_width(ws2, 3, 35)
auto_adjust_column_width(ws2, 4, 25)
auto_adjust_column_width(ws2, 5, 25)

# ========== Sheet 3: 워크플로우 개선안 ==========
ws3 = wb.create_sheet("워크플로우 개선안")

# 제목
ws3.merge_cells('A1:F1')
cell = ws3['A1']
cell.value = "워크플로우 개선안 (CRITICAL)"
set_cell_style(cell, bg_color="C00000", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws3.row_dimensions[1].height = 25

# 개선 1: 이미지-나레이션 매칭
row = 3
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="개선 1: 이미지-나레이션 1:1 매칭 (가장 중요)")
set_cell_style(cell, bg_color="FF6B6B", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
row += 1

ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="문제점")
set_cell_style(cell, bg_color="FFC7CE", bold=True)
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="현재: 전체 나레이션이 Gemini로 한 번에 전달되어 5개 이미지 프롬프트 생성 → 각 이미지가 특정 나레이션 부분과 매칭되지 않음")
set_cell_style(cell, bg_color="FFEB9C")
row += 1

ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="해결 방안")
set_cell_style(cell, bg_color="C6EFCE", bold=True)
row += 1

improvement1_flow = [
    ("현재 플로우", "개선된 플로우"),
    ("1. 전체 나레이션 → Gemini 2.5 Flash (5개 분할)", "1. 전체 나레이션 → Gemini 2.5 Flash (5개 분할)"),
    ("2. 전체 나레이션 → Gemini 3 Flash Preview (5개 이미지 프롬프트 한번에 생성) ❌", "2. 각 분할된 나레이션 → Loop (5번 반복) → Gemini 3 Flash Preview (1개씩 이미지 프롬프트 생성) ✅"),
    ("3. 5개 프롬프트 → Flux Pro (5개 이미지)", "3. 5개 프롬프트 → Flux Pro (5개 이미지)"),
    ("결과: 이미지와 나레이션이 안 맞음", "결과: 각 이미지가 해당 나레이션 부분과 완벽히 매칭됨")
]

create_header(ws3, row, improvement1_flow[0], bg_color="2F5496")
row += 1
for flow_data in improvement1_flow[1:]:
    for col_idx, val in enumerate(flow_data, 1):
        ws3.merge_cells(f'{get_column_letter(col_idx*3-2)}{row}:{get_column_letter(col_idx*3)}{row}')
        cell = ws3.cell(row=row, column=col_idx*3-2, value=val)
        if "❌" in val:
            set_cell_style(cell, bg_color="FFC7CE")
        elif "✅" in val:
            set_cell_style(cell, bg_color="C6EFCE")
        else:
            set_cell_style(cell)
    row += 1

# 구현 방법
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="n8n 구현 방법")
set_cell_style(cell, bg_color="BDD7EE", bold=True)
row += 1

implementation1 = [
    "1. Gemini 2.5 Flash로 나레이션 5개 분할 (기존과 동일)",
    "2. Loop Over Items 노드 추가 (5번 반복)",
    "3. Loop 내부에서 각 나레이션 부분을 Gemini 3 Flash Preview에 전달",
    "   - System prompt: '이 나레이션 부분에 맞는 이미지 1개를 생성하기 위한 프롬프트를 작성하세요'",
    "   - Input: 나레이션 부분 1개 + 선택된 장르/무드 스타일",
    "4. Loop 종료 후 5개 프롬프트 배열 생성",
    "5. Flux Pro로 이미지 생성 (기존과 동일)",
    "",
    "핵심: 전체 나레이션을 한번에 보내지 말고, 분할 후 각각 개별적으로 프롬프트 생성"
]

for impl_line in implementation1:
    ws3.merge_cells(f'A{row}:F{row}')
    cell = ws3.cell(row=row, column=1, value=impl_line)
    set_cell_style(cell)
    row += 1

# 개선 2: 비주얼 다양성
row += 2
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="개선 2: 비주얼 다양성 강화")
set_cell_style(cell, bg_color="FF6B6B", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
row += 1

ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="문제점")
set_cell_style(cell, bg_color="FFC7CE", bold=True)
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="현재: 고정된 15개 스타일 레퍼런스를 매번 동일하게 사용 → 모든 영상이 비슷한 비주얼 → 반복적이고 지루함")
set_cell_style(cell, bg_color="FFEB9C")
row += 1

ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="해결 방안: 랜덤 장르/무드 선택 시스템")
set_cell_style(cell, bg_color="C6EFCE", bold=True)
row += 1

genres_list = [
    ("장르/무드", "스타일 설명", "프롬프트 키워드"),
    ("사이버펑크 네온", "미래 도시, 네온 조명, 어두운 배경", "cyberpunk, neon lights, dark futuristic city"),
    ("자연 다큐멘터리", "자연 풍경, 야생동물, 생생한 색감", "nature documentary, wildlife, vivid natural colors"),
    ("미니멀 흑백", "단순함, 흑백, 고대비", "minimalist, black and white, high contrast"),
    ("레트로 필름그레인", "1970-80년대, 필름 질감, 빈티지", "retro 70s 80s, film grain, vintage"),
    ("하이테크 미래", "깨끗한 미래, 흰색/은색, 첨단 기술", "high-tech future, clean white silver, advanced technology"),
    ("다크 시네마틱", "영화 같은, 어두운 톤, 드라마틱 조명", "dark cinematic, dramatic lighting, moody"),
    ("팝아트 컬러풀", "밝은 색상, 만화 스타일, 대담한 색", "pop art, colorful, bold colors, comic style"),
    ("수중/우주", "물속 또는 우주, 떠다니는 느낌, 신비로움", "underwater or space, floating, mysterious"),
    ("동양풍 무협/사극", "한국/중국 전통, 수묵화, 고풍스러움", "oriental martial arts, ink painting, traditional"),
    ("고딕 다크판타지", "중세 판타지, 어두운 마법, 고딕 건축", "gothic dark fantasy, medieval magic, gothic architecture")
]

create_header(ws3, row, genres_list[0], bg_color="2F5496")
row += 1
for genre_data in genres_list[1:]:
    for col_idx, val in enumerate(genre_data, 1):
        cell = ws3.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell)
    row += 1

# 구현 방법
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="n8n 구현 방법")
set_cell_style(cell, bg_color="BDD7EE", bold=True)
row += 1

implementation2 = [
    "1. Code 노드 추가 (워크플로우 시작 부분)",
    "2. JavaScript로 장르 배열 정의 후 랜덤 선택:",
    "",
    "const genres = [",
    '  {name: "cyberpunk", keywords: "cyberpunk, neon lights, dark futuristic city"},',
    '  {name: "nature", keywords: "nature documentary, wildlife, vivid natural colors"},',
    '  {name: "minimal_bw", keywords: "minimalist, black and white, high contrast"},',
    "  // ... 나머지 장르들",
    "];",
    "const selected = genres[Math.floor(Math.random() * genres.length)];",
    "return {selected_genre: selected.name, style_keywords: selected.keywords};",
    "",
    "3. Gemini 이미지 프롬프트 생성 시 style_keywords를 함께 전달",
    "4. Google Sheets에 '장르/무드' 컬럼 추가하여 어떤 스타일 사용했는지 기록",
    "",
    "효과: 매번 완전히 다른 비주얼 스타일로 영상 생성 → 채널 콘텐츠 다양성 확보"
]

for impl_line in implementation2:
    ws3.merge_cells(f'A{row}:F{row}')
    cell = ws3.cell(row=row, column=1, value=impl_line)
    set_cell_style(cell)
    row += 1

# 개선 3: Flux Pro 프롬프트 최적화
row += 2
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="개선 3: Flux Pro 프롬프트 최적화")
set_cell_style(cell, bg_color="FFD93D", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="문제점")
set_cell_style(cell, bg_color="FFC7CE", bold=True)
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="현재: Gemini가 생성한 프롬프트에 'ultra-detailed, sharp focus, 8k, professional...' 같은 범용 suffix 추가 → 이미 Gemini 프롬프트에 스타일 포함되어 있어 불필요한 중복")
set_cell_style(cell, bg_color="FFEB9C")
row += 1

ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="해결 방안")
set_cell_style(cell, bg_color="C6EFCE", bold=True)
row += 1
ws3.merge_cells(f'A{row}:F{row}')
cell = ws3.cell(row=row, column=1, value="Gemini가 생성한 프롬프트를 그대로 사용. 범용 suffix 제거. Gemini 프롬프트 자체가 이미 장르/무드 스타일을 포함하고 있으므로 추가 수식어 불필요.")
set_cell_style(cell, bg_color="C6EFCE")
row += 1

# 컬럼 너비
auto_adjust_column_width(ws3, 1, 35)
auto_adjust_column_width(ws3, 2, 35)
auto_adjust_column_width(ws3, 3, 35)
auto_adjust_column_width(ws3, 4, 35)
auto_adjust_column_width(ws3, 5, 35)
auto_adjust_column_width(ws3, 6, 35)

# ========== Sheet 4: 채널별 워크플로우 설계 ==========
ws4 = wb.create_sheet("채널별 워크플로우 설계")

# 제목
ws4.merge_cells('A1:E1')
cell = ws4['A1']
cell.value = "채널별 워크플로우 설계"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws4.row_dimensions[1].height = 25

# 옵션 비교
row = 3
ws4.merge_cells(f'A{row}:E{row}')
cell = ws4.cell(row=row, column=1, value="다중 채널 처리 방식")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

options = [
    ("방식", "설명", "장점", "단점", "권장"),
    ("Option A", "단일 워크플로우 + Switch 분기", "유지보수 용이, 코드 중복 없음", "복잡도 증가", "✅ 숏폼 권장"),
    ("Option B", "채널별 별도 워크플로우", "독립적 관리, 명확한 구조", "코드 중복, 업데이트 번거로움", "롱폼 전용")
]

create_header(ws4, row, options[0], bg_color="2F5496")
row += 1
for opt_data in options[1:]:
    for col_idx, val in enumerate(opt_data, 1):
        cell = ws4.cell(row=row, column=col_idx, value=val)
        if "✅" in val:
            set_cell_style(cell, bg_color="C6EFCE", bold=True)
        else:
            set_cell_style(cell)
    row += 1

# 채널별 설정
row += 2
ws4.merge_cells(f'A{row}:E{row}')
cell = ws4.cell(row=row, column=1, value="채널별 세부 설정")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

channel_config = [
    ("채널", "콘텐츠 특성", "광고 정책", "Creatomate 템플릿", "추가 설정"),
    ("쇼츠 전용", "다양한 주제, 바이럴 콘텐츠", "8초 엔딩카드에 광고 (온카/루믹스/외부)", "기본 5씬 템플릿", "광고 랜덤 선택 로직"),
    ("루믹스 솔루션", "제품 데모, 솔루션 설명", "루믹스 제품 광고만", "루믹스 브랜딩 템플릿", "제품 이미지 오버레이"),
    ("온카스터디", "교육 콘텐츠, 커뮤니티", "온카 홍보 + 커뮤니티 링크", "교육용 템플릿 (깔끔한 폰트)", "자막 스타일 차별화"),
    ("슬롯", "슬롯 전문 콘텐츠", "온카스터디 광고만 (다른 광고 금지)", "슬롯 테마 템플릿", "광고 = 온카만 고정")
]

create_header(ws4, row, channel_config[0], bg_color="2F5496")
row += 1
for channel_data in channel_config[1:]:
    for col_idx, val in enumerate(channel_data, 1):
        cell = ws4.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell)
    row += 1

# Google Sheets 구조 확장
row += 2
ws4.merge_cells(f'A{row}:E{row}')
cell = ws4.cell(row=row, column=1, value="Google Sheets 구조 확장")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

sheets_expansion = [
    ("추가 컬럼", "타입", "설명", "예시 값"),
    ("채널", "Text", "어느 채널용 영상인지", "쇼츠 전용 / 루믹스 솔루션 / 온카스터디 / 슬롯"),
    ("장르/무드", "Text", "선택된 비주얼 스타일", "cyberpunk / nature / minimal_bw / ..."),
    ("광고유형", "Text", "어떤 광고를 넣을지", "온카 / 루믹스 / 외부광고 / 없음"),
    ("템플릿ID", "Text", "Creatomate 템플릿 ID", "abc123-shorts / def456-lumix / ..."),
    ("YouTube채널ID", "Text", "업로드할 YouTube 채널", "UCxxxxx... (채널별 다름)")
]

create_header(ws4, row, sheets_expansion[0], bg_color="2F5496")
row += 1
for sheet_data in sheets_expansion[1:]:
    for col_idx, val in enumerate(sheet_data, 1):
        cell = ws4.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="E7E6E6" if col_idx == 1 else None, bold=(col_idx==1))
    row += 1

# Switch 분기 로직
row += 2
ws4.merge_cells(f'A{row}:E{row}')
cell = ws4.cell(row=row, column=1, value="n8n Switch 분기 로직 (Option A 구현)")
set_cell_style(cell, bg_color="BDD7EE", bold=True, align_h="center")
row += 1

switch_logic = [
    "1. Google Sheets에서 '채널' 컬럼 읽기",
    "2. Switch 노드로 4가지 경로 분기:",
    "   - 케이스 1: 채널 == '쇼츠 전용' → 쇼츠 전용 브랜치",
    "   - 케이스 2: 채널 == '루믹스 솔루션' → 루믹스 브랜치",
    "   - 케이스 3: 채널 == '온카스터디' → 온카 브랜치",
    "   - 케이스 4: 채널 == '슬롯' → 슬롯 브랜치",
    "3. 각 브랜치에서:",
    "   - Creatomate 템플릿ID 설정",
    "   - 광고 소재 선택 (광고유형 컬럼 기반)",
    "   - YouTube 채널ID 설정",
    "4. 각 브랜치가 공통 업로드 노드로 합류",
    "",
    "장점: 공통 로직(TTS, 이미지 생성)은 한번만 작성, 채널별 차이만 분기 처리"
]

for logic_line in switch_logic:
    ws4.merge_cells(f'A{row}:E{row}')
    cell = ws4.cell(row=row, column=1, value=logic_line)
    set_cell_style(cell)
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws4, 1, 18)
auto_adjust_column_width(ws4, 2, 30)
auto_adjust_column_width(ws4, 3, 30)
auto_adjust_column_width(ws4, 4, 25)
auto_adjust_column_width(ws4, 5, 25)

# ========== Sheet 5: 에이전트 구성 ==========
ws5 = wb.create_sheet("에이전트 구성(상세)")

# 제목
ws5.merge_cells('A1:F1')
cell = ws5['A1']
cell.value = "6개 AI 에이전트 상세 설계"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws5.row_dimensions[1].height = 25

# PM 에이전트
row = 3
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="PM (총괄 매니저) - 전체 파이프라인 오케스트레이션")
set_cell_style(cell, bg_color="E2EFDA", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

pm_details = [
    ("항목", "내용"),
    ("트리거", "Schedule Trigger - 매일 08:00 (Cron: 0 8 * * *)"),
    ("주요 기능", "1. 오늘 제작할 영상 수량/채널 결정 (설정 시트에서 읽기)\n2. DA 에이전트 호출 → 성과 리포트 수집\n3. PL 에이전트 호출 → 기획안 생성 요청\n4. QA 에이전트 호출 → 기획안 검수\n5. 승인된 기획안 → PR 에이전트 트리거\n6. 전체 프로세스 모니터링 및 에러 처리"),
    ("입력", "없음 (자동 스케줄)"),
    ("출력", "전체 파이프라인 실행 결과 (성공/실패 리포트)"),
    ("n8n 구현", "- Schedule Trigger 노드\n- Execute Workflow 노드 (DA, PL, QA, PR 순차 호출)\n- If 노드로 에러 체크\n- Error Workflow로 실패 시 알림"),
    ("설정 예시", "Google Sheets '설정' 탭:\n- 쇼츠 전용: 3개/일\n- 루믹스: 1개/일\n- 온카: 1개/일\n- 슬롯: 2개/일"),
    ("에러 처리", "- 각 서브 워크플로우 실패 시 Slack/이메일 알림\n- 최대 3회 재시도\n- 실패 로그를 Google Sheets '에러 로그' 탭에 기록")
]

create_header(ws5, row, pm_details[0], bg_color="70AD47", font_color="FFFFFF")
row += 1
for pm_data in pm_details[1:]:
    ws5.cell(row=row, column=1, value=pm_data[0])
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=pm_data[1])
    set_cell_style(ws5.cell(row=row, column=1), bg_color="E2EFDA", bold=True)
    set_cell_style(ws5.cell(row=row, column=2))
    row += 1

# DA 에이전트
row += 1
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="DA (데이터 분석) - 성과 분석 및 인사이트 도출")
set_cell_style(cell, bg_color="D6E4F0", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

da_details = [
    ("항목", "내용"),
    ("트리거", "1. PM이 호출하면 실행 (Webhook)\n2. 매일 06:00 자동 실행 (독립 스케줄)"),
    ("주요 기능", "1. YouTube Analytics API로 채널별 데이터 수집:\n   - 조회수, 평균 시청 시간, CTR, 유입 소스\n   - 최근 7일/30일 데이터\n2. Google Sheets '성과 DB'에 저장\n3. Gemini로 AI 분석:\n   - 어떤 주제/스타일이 성과 좋았는지\n   - 채널별 트렌드 파악\n4. 분석 리포트 생성 (텍스트 형식)"),
    ("입력", "YouTube 채널ID 목록 (4개 채널)"),
    ("출력", "분석 리포트 (Google Sheets '분석 리포트' 탭에 저장)"),
    ("n8n 구현", "- HTTP Request 노드 (YouTube Analytics API)\n- Google Sheets 노드 (데이터 저장)\n- Gemini API 노드 (AI 분석)\n- Google Sheets 노드 (리포트 저장)"),
    ("분석 예시", "'쇼츠 전용 채널에서 사이버펑크 스타일 영상이 평균 조회수 15% 높음'\n'루믹스 채널에서 제품 데모 형식이 CTR 8% 더 좋음'\n'슬롯 채널에서 짧은 제목(30자 이하)이 성과 좋음'"),
    ("API 설정", "YouTube Analytics API:\n- Scope: youtube.readonly, yt-analytics.readonly\n- Metrics: views, averageViewDuration, ctr, subscribersGained\n- Dimensions: video, day")
]

create_header(ws5, row, da_details[0], bg_color="5B9BD5", font_color="FFFFFF")
row += 1
for da_data in da_details[1:]:
    ws5.cell(row=row, column=1, value=da_data[0])
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=da_data[1])
    set_cell_style(ws5.cell(row=row, column=1), bg_color="D6E4F0", bold=True)
    set_cell_style(ws5.cell(row=row, column=2))
    row += 1

# PL 에이전트
row += 1
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="PL (기획) - 채널별 영상 기획안 자동 생성")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

pl_details = [
    ("항목", "내용"),
    ("트리거", "PM이 호출 (DA 리포트 완료 후)"),
    ("주요 기능", "1. DA 리포트 읽기\n2. 채널별 가이드라인 읽기 (Google Sheets '채널 가이드' 탭)\n3. Gemini로 기획안 생성:\n   - 영상 제목 (2-3개 옵션)\n   - 나레이션 대본\n   - 톤 앤 매너\n   - 타겟 오디언스\n4. Google Sheets '기획안 DB'에 저장 (Status=기획완료)"),
    ("입력", "- DA 리포트\n- 채널명\n- 생성 개수"),
    ("출력", "기획안 (Google Sheets에 저장)"),
    ("n8n 구현", "- Google Sheets 노드 (DA 리포트, 채널 가이드 읽기)\n- Gemini API 노드 (기획안 생성)\n- Loop 노드 (채널별/개수별 반복)\n- Google Sheets 노드 (기획안 저장)"),
    ("Gemini Prompt", "System: '당신은 YouTube 숏폼 콘텐츠 기획 전문가입니다.'\nUser: '채널: {채널명}\n채널 가이드: {가이드}\nDA 리포트: {리포트}\n\n위 정보를 바탕으로 영상 기획안을 작성하세요.\n- 제목 3개 (각 50자 이내, 클릭 유도)\n- 나레이션 대본 (60초 분량)\n- 톤: 친근함/전문적/유머 등\n- 타겟: 20대 남성/30대 직장인 등'"),
    ("채널 가이드", "각 채널별로 정의:\n- 쇼츠 전용: 다양한 주제, 자극적 제목, 젊은 층\n- 루믹스: 비즈니스 톤, 솔루션 강조, 30-50대\n- 온카: 교육적, 커뮤니티 친화, 전 연령\n- 슬롯: 슬롯 전문, 온카 자연스럽게 언급, 20-40대")
]

create_header(ws5, row, pl_details[0], bg_color="FFD966", font_color="000000")
row += 1
for pl_data in pl_details[1:]:
    ws5.cell(row=row, column=1, value=pl_data[0])
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=pl_data[1])
    set_cell_style(ws5.cell(row=row, column=1), bg_color="FFF2CC", bold=True)
    set_cell_style(ws5.cell(row=row, column=2))
    row += 1

# QA 에이전트
row += 1
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="QA (검수) - 기획안 평가 및 피드백")
set_cell_style(cell, bg_color="F8CBAD", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

qa_details = [
    ("항목", "내용"),
    ("트리거", "PL 기획안 완료 후 자동 실행"),
    ("주요 기능", "1. PL이 생성한 기획안 읽기\n2. Gemini로 기획안 평가 (1-10점 채점 + 구체적 피드백)\n3. 평가 기준:\n   - 트렌드 적합성 (20%)\n   - 채널 정체성 부합 (25%)\n   - 광고 정책 준수 (15%)\n   - 제작 가능성 (20%)\n   - 참신함/독창성 (20%)\n4. 점수에 따른 처리:\n   - 7-10점: 승인 → PR 에이전트로 전달\n   - 4-6점: 수정 요청 → PL에게 피드백 전달\n   - 1-3점: 폐기 → 새 기획안 요청\n5. 중복 체크: 최근 3일 내 비슷한 주제 있는지 확인"),
    ("입력", "PL 기획안"),
    ("출력", "평가 점수 + 피드백 (Google Sheets에 저장)"),
    ("n8n 구현", "- Google Sheets 노드 (기획안 읽기)\n- Gemini API 노드 (평가 및 피드백)\n- If 노드 (점수별 분기)\n- Loop 노드 (최대 3회 피드백 루프)\n- Google Sheets 노드 (평가 결과 저장)"),
    ("Gemini Prompt", "System: '당신은 엄격한 콘텐츠 품질 검수자입니다.'\nUser: '기획안:\n제목: {제목}\n나레이션: {나레이션}\n채널: {채널}\n\n다음 기준으로 1-10점 채점 및 피드백:\n1. 트렌드 적합성 (20점)\n2. 채널 정체성 (25점)\n3. 광고 정책 준수 (15점)\n4. 제작 가능성 (20점)\n5. 참신함 (20점)\n\n형식:\n총점: X/10\n피드백: [구체적인 개선 사항]'"),
    ("피드백 루프", "1차: PL 생성 → QA 평가 6점 → 피드백 '제목이 평범함'\n2차: PL 수정 → QA 평가 8점 → 승인\n\n최대 3회까지 반복, 3회 후에도 7점 미만이면 폐기"),
    ("중복 체크", "Google Sheets에서 최근 3일 영상 제목/키워드 검색\n유사도 70% 이상이면 '중복 가능성' 경고")
]

create_header(ws5, row, qa_details[0], bg_color="F4B084", font_color="000000")
row += 1
for qa_data in qa_details[1:]:
    ws5.cell(row=row, column=1, value=qa_data[0])
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=qa_data[1])
    set_cell_style(ws5.cell(row=row, column=1), bg_color="F8CBAD", bold=True)
    set_cell_style(ws5.cell(row=row, column=2))
    row += 1

# PR 에이전트
row += 1
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="PR (제작) - 승인된 기획안을 영상으로 제작")
set_cell_style(cell, bg_color="D9E2F3", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

pr_details = [
    ("항목", "내용"),
    ("트리거", "QA 승인 후 자동 실행"),
    ("주요 기능", "1. 승인된 기획안을 Google Sheets에 Status=준비로 저장\n2. 기존 '전용' 워크플로우 트리거 (개선 버전)\n3. 워크플로우 실행:\n   - 나레이션 5개 분할\n   - 장르/무드 랜덤 선택\n   - 이미지-나레이션 1:1 매칭 (Loop)\n   - TTS, 이미지, BGM, 비디오 생성\n   - 채널별 Creatomate 템플릿 적용\n   - 최종 영상 합성\n4. Status=생성완료로 업데이트"),
    ("입력", "승인된 기획안 (제목, 나레이션, 채널)"),
    ("출력", "완성된 영상 URL (Creatomate)"),
    ("n8n 구현", "- Google Sheets 노드 (기획안 → 메인 시트로 복사)\n- Execute Workflow 노드 ('전용' 워크플로우 호출)\n- 또는 직접 통합 (하나의 워크플로우로)"),
    ("개선 사항", "Sheet 3의 개선안 모두 적용:\n- 개선 1: 이미지-나레이션 매칭\n- 개선 2: 장르/무드 랜덤화\n- 개선 3: Flux Pro 프롬프트 최적화\n- 채널별 분기 처리"),
    ("기존 워크플로우", "Sheet 2의 '전용' 워크플로우 = PR 에이전트의 핵심\n현재는 수동 트리거, 향후 QA 승인 시 자동 트리거로 변경")
]

create_header(ws5, row, pr_details[0], bg_color="8EA9DB", font_color="FFFFFF")
row += 1
for pr_data in pr_details[1:]:
    ws5.cell(row=row, column=1, value=pr_data[0])
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=pr_data[1])
    set_cell_style(ws5.cell(row=row, column=1), bg_color="D9E2F3", bold=True)
    set_cell_style(ws5.cell(row=row, column=2))
    row += 1

# DL 에이전트
row += 1
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="DL (배포) - 채널별 YouTube 업로드 및 최적화")
set_cell_style(cell, bg_color="E2D9F3", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

dl_details = [
    ("항목", "내용"),
    ("트리거", "PR 완료 후 자동 실행 (Status=생성완료)"),
    ("주요 기능", "1. 완성된 영상 다운로드 (Creatomate URL)\n2. 채널별 YouTube 계정 선택\n3. Gemini로 최적화된 메타데이터 생성:\n   - 제목 (채널 특성 반영)\n   - 설명 (키워드 포함)\n   - 태그 (SEO 최적화)\n4. YouTube Data API로 업로드\n5. 자동 댓글 작성 (채널별 다른 내용)\n6. Google Sheets 업데이트:\n   - 업로드 URL\n   - Status=발행완료\n   - Publish=발행완료"),
    ("입력", "영상 URL, 채널명, 기획안"),
    ("출력", "YouTube URL + Sheets 업데이트"),
    ("n8n 구현", "- HTTP Request 노드 (영상 다운로드)\n- Switch 노드 (채널별 분기)\n- Gemini API 노드 (메타데이터 최적화)\n- YouTube API 노드 (업로드 + 댓글)\n- Google Sheets 노드 (상태 업데이트)"),
    ("채널별 최적화", "- 쇼츠 전용: 자극적 제목, 해시태그 많이\n- 루믹스: 비즈니스 키워드, 전문적 설명\n- 온카: 커뮤니티 링크, 교육 태그\n- 슬롯: 슬롯 관련 태그, 온카 링크"),
    ("댓글 예시", "- 쇼츠 전용: '더 많은 정보는 프로필 링크에서!'\n- 루믹스: '루믹스 솔루션 문의: [링크]'\n- 온카: '온카스터디 커뮤니티 가입: [링크]'\n- 슬롯: '안전한 슬롯은 온카스터디에서: [링크]'"),
    ("기존 워크플로우", "Sheet 2의 단계 5-6 = DL 에이전트의 핵심\n채널별 분기 및 메타데이터 최적화 추가 필요")
]

create_header(ws5, row, dl_details[0], bg_color="B4A7D6", font_color="FFFFFF")
row += 1
for dl_data in dl_details[1:]:
    ws5.cell(row=row, column=1, value=dl_data[0])
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=dl_data[1])
    set_cell_style(ws5.cell(row=row, column=1), bg_color="E2D9F3", bold=True)
    set_cell_style(ws5.cell(row=row, column=2))
    row += 1

# 미래 에이전트
row += 2
ws5.merge_cells(f'A{row}:F{row}')
cell = ws5.cell(row=row, column=1, value="미래 확장 에이전트 (선택적)")
set_cell_style(cell, bg_color="A9D08E", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=12, bold=True)
row += 1

future_agents = [
    ("에이전트", "목적", "기능"),
    ("AD (광고관리)", "광고 소재 관리 및 최적 배치", "- 광고 소재 DB 관리\n- 채널/영상별 최적 광고 선택\n- 광고 성과 트래킹"),
    ("TR (트렌드)", "외부 트렌드 모니터링", "- Google Trends API 연동\n- 경쟁 채널 분석\n- 실시간 트렌드 키워드 수집\n- PL에게 트렌드 정보 제공")
]

create_header(ws5, row, future_agents[0], bg_color="70AD47", font_color="FFFFFF")
row += 1
for agent_data in future_agents[1:]:
    for col_idx, val in enumerate(agent_data, 1):
        if col_idx == 3:
            ws5.merge_cells(f'C{row}:F{row}')
        ws5.cell(row=row, column=col_idx if col_idx < 3 else 3, value=val)
        set_cell_style(ws5.cell(row=row, column=col_idx if col_idx < 3 else 3))
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws5, 1, 18)
auto_adjust_column_width(ws5, 2, 60)

# ========== Sheet 6: 피드백 루프 상세 ==========
ws6 = wb.create_sheet("피드백 루프 상세")

# 제목
ws6.merge_cells('A1:E1')
cell = ws6['A1']
cell.value = "PL ↔ QA 피드백 루프 상세 설계"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws6.row_dimensions[1].height = 25

# 프로세스 플로우
row = 3
ws6.merge_cells(f'A{row}:E{row}')
cell = ws6.cell(row=row, column=1, value="피드백 루프 프로세스")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

feedback_flow = [
    ("단계", "담당", "작업 내용", "출력", "다음 단계"),
    ("1", "PL", "기획안 생성 (초안)", "제목, 나레이션, 톤", "→ QA"),
    ("2", "QA", "기획안 평가 (채점)", "점수 (1-10) + 피드백", "분기"),
    ("3a", "QA", "점수 ≥7 → 승인", "Status=승인", "→ PR"),
    ("3b", "QA", "점수 4-6 → 수정 요청", "구체적 피드백", "→ PL (4단계)"),
    ("3c", "QA", "점수 ≤3 → 폐기", "Status=폐기", "→ PL 새 기획안"),
    ("4", "PL", "피드백 반영하여 수정", "수정된 기획안", "→ QA (2단계)"),
    ("5", "시스템", "최대 3회 반복 확인", "루프 카운트", "3회 초과 시 폐기")
]

create_header(ws6, row, feedback_flow[0], bg_color="2F5496")
row += 1
for flow_data in feedback_flow[1:]:
    for col_idx, val in enumerate(flow_data, 1):
        cell = ws6.cell(row=row, column=col_idx, value=val)
        if "승인" in val:
            set_cell_style(cell, bg_color="C6EFCE")
        elif "폐기" in val:
            set_cell_style(cell, bg_color="FFC7CE")
        elif "수정" in val:
            set_cell_style(cell, bg_color="FFEB9C")
        else:
            set_cell_style(cell)
    row += 1

# 평가 기준
row += 2
ws6.merge_cells(f'A{row}:E{row}')
cell = ws6.cell(row=row, column=1, value="QA 평가 기준 (가중치)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

scoring_criteria = [
    ("평가 항목", "가중치", "평가 내용", "점수 산정 예시"),
    ("트렌드 적합성", "20%", "현재 인기 있는 주제/형식인가?", "DA 리포트와 비교, 최근 트렌드 키워드 포함 여부"),
    ("채널 정체성", "25%", "채널 가이드라인에 맞는가?", "쇼츠=자극적, 루믹스=전문적, 온카=교육적, 슬롯=슬롯 전문"),
    ("광고 정책 준수", "15%", "채널별 광고 정책 위반 없는가?", "슬롯 채널에 온카 외 광고 없는지, 광고 시간 적절한지"),
    ("제작 가능성", "20%", "현재 워크플로우로 제작 가능한가?", "너무 복잡한 비주얼 요구 없는지, 5씬 구조에 맞는지"),
    ("참신함/독창성", "20%", "기존 콘텐츠와 차별화되는가?", "최근 3일 내 비슷한 주제 없는지, 새로운 시각 있는지")
]

create_header(ws6, row, scoring_criteria[0], bg_color="2F5496")
row += 1
for criteria_data in scoring_criteria[1:]:
    for col_idx, val in enumerate(criteria_data, 1):
        cell = ws6.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell)
    row += 1

# n8n 구현
row += 2
ws6.merge_cells(f'A{row}:E{row}')
cell = ws6.cell(row=row, column=1, value="n8n 구현 방법")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

implementation = [
    "【워크플로우 구조】",
    "PL Workflow (기획안 생성) ↔ QA Workflow (검수) 분리",
    "",
    "【PL Workflow】",
    "1. Webhook Trigger (PM 또는 QA에서 호출)",
    "2. Input: {request_type: 'new' or 'revise', feedback: '...', loop_count: 0}",
    "3. If 노드로 분기:",
    "   - request_type == 'new': DA 리포트 기반 새 기획안",
    "   - request_type == 'revise': QA 피드백 기반 수정",
    "4. Gemini 노드 (기획안 생성/수정)",
    "5. Google Sheets 저장 (기획안 DB)",
    "6. HTTP Request로 QA Workflow 호출 (기획안 전달)",
    "",
    "【QA Workflow】",
    "1. Webhook Trigger (PL에서 호출)",
    "2. Input: {plan_id: '...', loop_count: 0}",
    "3. Google Sheets 읽기 (기획안, 최근 3일 영상)",
    "4. Gemini 노드 (평가 + 채점)",
    "5. Code 노드로 점수 파싱",
    "6. If 노드로 3방향 분기:",
    "   - 점수 ≥7: Google Sheets 업데이트 (Status=승인) → Execute Workflow (PR 호출)",
    "   - 점수 4-6 AND loop_count < 3: HTTP Request (PL Workflow 호출, request_type='revise')",
    "   - 점수 ≤3 OR loop_count ≥3: Google Sheets 업데이트 (Status=폐기) → PM에 알림",
    "7. loop_count++ 하여 다음 루프에 전달",
    "",
    "【중복 체크 로직】",
    "- Google Sheets에서 최근 3일 영상 제목 가져오기",
    "- Code 노드로 유사도 계산 (간단한 키워드 매칭 또는 Gemini로 유사도 판단)",
    "- 유사도 70% 이상이면 QA 점수에 -2점 페널티"
]

for impl_line in implementation:
    ws6.merge_cells(f'A{row}:E{row}')
    cell = ws6.cell(row=row, column=1, value=impl_line)
    if "【" in impl_line:
        set_cell_style(cell, bg_color="E7E6E6", bold=True)
    else:
        set_cell_style(cell)
    row += 1

# 예시 시나리오
row += 2
ws6.merge_cells(f'A{row}:E{row}')
cell = ws6.cell(row=row, column=1, value="구체적 예시 시나리오")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

scenario = [
    ("라운드", "담당", "내용", "점수", "결과"),
    ("1차", "PL", "제목: '도박 중독 탈출하는 법'\n나레이션: 평범한 설명형", "-", "→ QA"),
    ("1차", "QA", "평가: 제목이 너무 평범, 자극성 부족\n채널 정체성(쇼츠) 불일치", "5/10", "수정 요청"),
    ("2차", "PL", "제목: '도박으로 3천만원 날린 내가 벗어난 방법'\n나레이션: 스토리텔링 추가", "-", "→ QA"),
    ("2차", "QA", "평가: 제목 개선됨, 스토리 좋음\n하지만 비슷한 주제 2일 전 업로드됨", "6/10", "수정 요청"),
    ("3차", "PL", "제목: '도박 중독자들이 숨기는 3가지 심리'\n나레이션: 심리 분석 각도로 변경", "-", "→ QA"),
    ("3차", "QA", "평가: 참신한 각도, 중복 없음\nDA 리포트와 부합, 제작 가능", "8/10", "✅ 승인 → PR")
]

create_header(ws6, row, scenario[0], bg_color="2F5496")
row += 1
for scenario_data in scenario[1:]:
    for col_idx, val in enumerate(scenario_data, 1):
        cell = ws6.cell(row=row, column=col_idx, value=val)
        if "승인" in str(val):
            set_cell_style(cell, bg_color="C6EFCE", bold=True)
        elif "수정 요청" in str(val):
            set_cell_style(cell, bg_color="FFEB9C")
        else:
            set_cell_style(cell)
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws6, 1, 12)
auto_adjust_column_width(ws6, 2, 12)
auto_adjust_column_width(ws6, 3, 40)
auto_adjust_column_width(ws6, 4, 12)
auto_adjust_column_width(ws6, 5, 18)

# ========== Sheet 7: 구축 로드맵 ==========
ws7 = wb.create_sheet("구축 로드맵")

# 제목
ws7.merge_cells('A1:F1')
cell = ws7['A1']
cell.value = "단계별 구축 로드맵"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws7.row_dimensions[1].height = 25

# 로드맵
row = 3
roadmap = [
    ("Phase", "기간", "주요 작업", "세부 내용", "완료 조건", "우선순위"),
    ("Phase 1", "1주", "Hostinger VPS 이전", "- VPS KVM 2 구매 및 설정\n- n8n Docker 설치\n- 기존 워크플로우 Export/Import\n- Credentials 재설정 (모든 API 키)", "- n8n 정상 작동\n- 기존 '전용' 워크플로우 실행 성공", "높음"),
    ("Phase 2", "1주", "'전용' 워크플로우 개선", "- 이미지-나레이션 1:1 매칭 구현 (Loop)\n- 장르/무드 랜덤 선택 Code 노드\n- Flux Pro 프롬프트 최적화\n- 테스트 실행 (최소 5회)", "- 이미지가 나레이션과 정확히 매칭됨\n- 매번 다른 비주얼 스타일 확인", "높음"),
    ("Phase 3", "1-2주", "채널별 확장", "- Google Sheets 구조 확장 (채널, 광고유형 등)\n- Switch 노드로 채널 분기\n- 채널별 Creatomate 템플릿 생성 (4개)\n- YouTube 채널별 Credential 설정\n- 테스트 (각 채널 1개씩)", "- 4개 채널 모두 정상 업로드\n- 채널별 브랜딩 적용 확인", "높음"),
    ("Phase 4", "1-2주", "DA 에이전트 구축", "- YouTube Analytics API 설정\n- 데이터 수집 워크플로우 (DA-01)\n- Gemini 분석 워크플로우 (DA-02)\n- Google Sheets '성과 DB' 구조 설계\n- 일주일 데이터 수집 후 분석 테스트", "- 채널별 성과 데이터 자동 수집\n- AI 분석 리포트 생성 확인", "중간"),
    ("Phase 5", "2-3주", "PL + QA 에이전트", "- PL 워크플로우 (기획안 생성)\n- QA 워크플로우 (검수 + 채점)\n- 피드백 루프 구현 (최대 3회)\n- 중복 체크 로직\n- Google Sheets '기획안 DB' 구조\n- 테스트 (10개 기획안)", "- 기획안 자동 생성\n- QA 검수 및 피드백 정상 작동\n- 승인된 기획안 PR로 전달", "중간"),
    ("Phase 6", "1주", "PM 에이전트", "- PM 워크플로우 (오케스트레이션)\n- Schedule Trigger 설정 (매일 08:00)\n- 에러 처리 워크플로우 (PM-02)\n- 전체 파이프라인 통합 테스트\n- 설정 시트 구조 설계", "- 클릭 0번으로 전체 파이프라인 실행\n- 에러 발생 시 자동 알림", "중간"),
    ("Phase 7", "지속적", "고도화 및 확장", "- A/B 테스트 시스템\n- 롱폼 콘텐츠 워크플로우\n- AD 에이전트 (광고 관리)\n- TR 에이전트 (트렌드 모니터링)\n- 성과 기반 자동 학습\n- 비용 최적화", "- 지속적 개선\n- 새로운 기능 추가", "낮음")
]

create_header(ws7, row, roadmap[0], bg_color="2F5496")
row += 1
for phase_data in roadmap[1:]:
    for col_idx, val in enumerate(phase_data, 1):
        cell = ws7.cell(row=row, column=col_idx, value=val)
        # 우선순위 색상
        if col_idx == 6:
            if val == "높음":
                set_cell_style(cell, bg_color="FF6B6B", font_color="FFFFFF", bold=True, align_h="center")
            elif val == "중간":
                set_cell_style(cell, bg_color="FFD93D", bold=True, align_h="center")
            else:
                set_cell_style(cell, bg_color="6BCB77", font_color="FFFFFF", bold=True, align_h="center")
        # Phase 색상
        elif col_idx == 1:
            set_cell_style(cell, bg_color="D6E4F0", bold=True, align_h="center")
        else:
            set_cell_style(cell)
    ws7.row_dimensions[row].height = 80
    row += 1

# 타임라인
row += 2
ws7.merge_cells(f'A{row}:F{row}')
cell = ws7.cell(row=row, column=1, value="전체 타임라인 (약 7-10주)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

timeline = [
    "Week 1: Phase 1 (VPS 이전)",
    "Week 2: Phase 2 (워크플로우 개선)",
    "Week 3-4: Phase 3 (채널별 확장)",
    "Week 5-6: Phase 4 (DA 에이전트)",
    "Week 7-9: Phase 5 (PL + QA 에이전트)",
    "Week 10: Phase 6 (PM 에이전트 + 통합 테스트)",
    "Week 11+: Phase 7 (고도화)",
    "",
    "🎯 Week 10 완료 시점 = 클릭 0번 자동화 시스템 완성"
]

for timeline_line in timeline:
    ws7.merge_cells(f'A{row}:F{row}')
    cell = ws7.cell(row=row, column=1, value=timeline_line)
    if "🎯" in timeline_line:
        set_cell_style(cell, bg_color="C6EFCE", bold=True, align_h="center")
        cell.font = Font(name="맑은 고딕", size=11, bold=True)
    else:
        set_cell_style(cell)
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws7, 1, 12)
auto_adjust_column_width(ws7, 2, 10)
auto_adjust_column_width(ws7, 3, 25)
auto_adjust_column_width(ws7, 4, 40)
auto_adjust_column_width(ws7, 5, 30)
auto_adjust_column_width(ws7, 6, 12)

# ========== Sheet 8: 워크플로우 전체 목록 ==========
ws8 = wb.create_sheet("워크플로우 전체 목록")

# 제목
ws8.merge_cells('A1:F1')
cell = ws8['A1']
cell.value = "전체 워크플로우 목록"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws8.row_dimensions[1].height = 25

# 워크플로우 목록
row = 3
workflows = [
    ("워크플로우명", "상태", "설명", "트리거", "주요 노드", "담당 에이전트"),
    ("전용 (기존)", "수정필요", "5-Scene 쇼츠 자동 생성 (개선 적용 필요)", "Manual / Webhook", "Google Sheets, Gemini, fal.ai, Creatomate, YouTube", "PR"),
    ("DA-01 일일 성과 수집", "신규", "YouTube Analytics API로 채널별 성과 데이터 수집", "Schedule (06:00) / Webhook", "HTTP Request, Google Sheets", "DA"),
    ("DA-02 AI 분석 리포트", "신규", "수집된 데이터를 Gemini로 분석하여 인사이트 도출", "DA-01 완료 후", "Google Sheets, Gemini, Google Sheets", "DA"),
    ("PL-01 기획안 자동 생성", "신규", "DA 리포트 기반 채널별 영상 기획안 생성", "Webhook (PM 또는 QA)", "Google Sheets, Gemini, Loop, Google Sheets", "PL"),
    ("QA-01 기획 검수", "신규", "PL 기획안 평가 및 채점 (1-10점)", "PL-01 완료 후", "Google Sheets, Gemini, If, Google Sheets", "QA"),
    ("PL-02 기획안 수정 루프", "신규", "QA 피드백 기반 기획안 수정 (최대 3회)", "QA-01 수정 요청 시", "Google Sheets, Gemini, Loop, Google Sheets", "PL"),
    ("PR-01 영상 제작", "수정필요", "'전용' 워크플로우 = PR 에이전트 (개선 적용)", "QA 승인 후", "전용 워크플로우와 동일", "PR"),
    ("DL-01 채널별 업로드", "수정필요", "완성 영상을 채널별 YouTube 계정에 업로드", "PR-01 완료 후", "HTTP Request, Switch, Gemini, YouTube, Google Sheets", "DL"),
    ("PM-01 일일 자동 파이프라인", "신규", "매일 자동으로 전체 파이프라인 실행 (DA → PL → QA → PR → DL)", "Schedule (08:00)", "Execute Workflow (서브워크플로우 호출)", "PM"),
    ("PM-02 에러 처리", "신규", "파이프라인 에러 발생 시 재시도 및 알림", "에러 발생 시", "If, Loop, HTTP Request (Slack/Email)", "PM"),
    ("AD-01 광고 소재 관리", "미래", "광고 소재 DB 관리 및 최적 배치 (선택적)", "-", "-", "AD (미래)"),
    ("TR-01 트렌드 모니터링", "미래", "Google Trends 및 경쟁 채널 분석 (선택적)", "Schedule (매일)", "HTTP Request, Gemini", "TR (미래)")
]

create_header(ws8, row, workflows[0], bg_color="2F5496")
row += 1
for wf_data in workflows[1:]:
    for col_idx, val in enumerate(wf_data, 1):
        cell = ws8.cell(row=row, column=col_idx, value=val)
        # 상태별 색상
        if col_idx == 2:
            if val == "신규":
                set_cell_style(cell, bg_color="BDD7EE", bold=True, align_h="center")
            elif val == "수정필요":
                set_cell_style(cell, bg_color="FFEB9C", bold=True, align_h="center")
            elif val == "기존있음":
                set_cell_style(cell, bg_color="C6EFCE", bold=True, align_h="center")
            elif val == "미래":
                set_cell_style(cell, bg_color="E7E6E6", align_h="center")
            else:
                set_cell_style(cell)
        else:
            set_cell_style(cell)
    row += 1

# 범례
row += 2
ws8.merge_cells(f'A{row}:F{row}')
cell = ws8.cell(row=row, column=1, value="상태 범례")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

legend = [
    ("상태", "의미", "설명"),
    ("신규", "새로 만들어야 함", "Phase 4-6에서 구축"),
    ("수정필요", "기존 것을 개선", "Phase 2-3에서 개선"),
    ("기존있음", "이미 작동 중", "그대로 사용 또는 약간 수정"),
    ("미래", "선택적 확장", "Phase 7 이후 필요 시 구축")
]

create_header(ws8, row, legend[0], bg_color="2F5496")
row += 1
for legend_data in legend[1:]:
    for col_idx, val in enumerate(legend_data, 1):
        cell = ws8.cell(row=row, column=col_idx, value=val)
        if col_idx == 1:
            if val == "신규":
                set_cell_style(cell, bg_color="BDD7EE", bold=True, align_h="center")
            elif val == "수정필요":
                set_cell_style(cell, bg_color="FFEB9C", bold=True, align_h="center")
            elif val == "기존있음":
                set_cell_style(cell, bg_color="C6EFCE", bold=True, align_h="center")
            elif val == "미래":
                set_cell_style(cell, bg_color="E7E6E6", bold=True, align_h="center")
        else:
            set_cell_style(cell)
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws8, 1, 25)
auto_adjust_column_width(ws8, 2, 15)
auto_adjust_column_width(ws8, 3, 40)
auto_adjust_column_width(ws8, 4, 20)
auto_adjust_column_width(ws8, 5, 35)
auto_adjust_column_width(ws8, 6, 18)

# ========== Sheet 9: 필요 도구 및 API + 비용 ==========
ws9 = wb.create_sheet("필요 도구 및 API + 비용")

# 제목
ws9.merge_cells('A1:F1')
cell = ws9['A1']
cell.value = "필요 도구, API 및 월간 비용 예상"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws9.row_dimensions[1].height = 25

# 기존 도구
row = 3
ws9.merge_cells(f'A{row}:F{row}')
cell = ws9.cell(row=row, column=1, value="기존 도구 (이미 사용 중)")
set_cell_style(cell, bg_color="70AD47", font_color="FFFFFF", bold=True, align_h="center")
row += 1

existing_tools = [
    ("도구/서비스", "용도", "현재 상태", "가격", "예상 월 비용", "비고"),
    ("Google Gemini 2.5 Flash", "나레이션 분할", "사용 중", "Free tier / Pay-as-you-go", "$0 (Free tier 내)", "월 1500 requests (무료)"),
    ("Google Gemini 3 Flash Preview", "이미지 프롬프트 생성", "사용 중", "Free tier / Pay-as-you-go", "$0 (Free tier 내)", "월 1500 requests (무료)"),
    ("fal.ai - ElevenLabs TTS", "음성 생성", "사용 중", "Pay-per-use", "$5-10", "영상당 5개 × 30개 = 150개/월"),
    ("fal.ai - Flux Pro", "이미지 생성", "사용 중", "Pay-per-use", "$15-25", "영상당 5개 × 30개 = 150개/월"),
    ("fal.ai - ESRGAN", "이미지 업스케일", "사용 중", "Pay-per-use", "$3-5", "150개/월"),
    ("fal.ai - Kling Video v1.6", "비디오 생성", "사용 중", "Pay-per-use", "$30-50", "150개/월 (비용 높음)"),
    ("fal.ai - Beatoven", "BGM 생성", "사용 중", "Pay-per-use", "$2-5", "30-50개/월"),
    ("Creatomate", "영상 합성", "사용 중", "Subscription", "$49/월", "무제한 렌더링 플랜"),
    ("YouTube Data API v3", "업로드, 댓글", "사용 중", "Free (할당량 내)", "$0", "일일 할당량 10,000 units"),
    ("Google Sheets API", "데이터베이스", "사용 중", "Free", "$0", ""),
    ("Google Drive API", "파일 저장", "사용 중", "Free", "$0", "")
]

create_header(ws9, row, existing_tools[0], bg_color="70AD47", font_color="FFFFFF")
row += 1
for tool_data in existing_tools[1:]:
    for col_idx, val in enumerate(tool_data, 1):
        cell = ws9.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="E2EFDA")
    row += 1

# 신규 필요
row += 2
ws9.merge_cells(f'A{row}:F{row}')
cell = ws9.cell(row=row, column=1, value="신규 필요 도구/서비스")
set_cell_style(cell, bg_color="5B9BD5", font_color="FFFFFF", bold=True, align_h="center")
row += 1

new_tools = [
    ("도구/서비스", "용도", "필요 시점", "가격", "예상 월 비용", "비고"),
    ("Hostinger VPS KVM 2", "n8n 호스팅", "Phase 1", "$6.99/월", "$6.99", "4GB RAM, 2 Core, 100GB SSD"),
    ("YouTube Analytics API", "성과 데이터 수집", "Phase 4", "Free (OAuth)", "$0", "YouTube Data API와 동일 OAuth"),
    ("Gemini API (추가 호출)", "DA/PL/QA 에이전트", "Phase 4-5", "Pay-as-you-go", "$5-15", "Free tier 초과 시 (월 3000+ requests)")
]

create_header(ws9, row, new_tools[0], bg_color="5B9BD5", font_color="FFFFFF")
row += 1
for new_data in new_tools[1:]:
    for col_idx, val in enumerate(new_data, 1):
        cell = ws9.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="D6E4F0")
    row += 1

# 월간 비용 총합
row += 2
ws9.merge_cells(f'A{row}:F{row}')
cell = ws9.cell(row=row, column=1, value="월간 비용 총 예상")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

cost_summary = [
    ("카테고리", "예상 비용 (최소)", "예상 비용 (최대)", "비고"),
    ("인프라 (VPS)", "$6.99", "$6.99", "고정 비용"),
    ("fal.ai (TTS, 이미지, 비디오, BGM)", "$50", "$85", "사용량 기반 (영상 30개/월 가정)"),
    ("Creatomate", "$49", "$49", "고정 비용 (무제한 플랜)"),
    ("Gemini API", "$0", "$15", "Free tier 초과 시"),
    ("기타 (YouTube, Sheets 등)", "$0", "$0", "무료"),
    ("", "", "", ""),
    ("월간 총 비용", "$105.99", "$155.99", "영상 30개/월 기준")
]

for cost_data in cost_summary:
    if cost_data[0] == "월간 총 비용":
        ws9.merge_cells(f'A{row}:A{row}')
        cell = ws9.cell(row=row, column=1, value=cost_data[0])
        set_cell_style(cell, bg_color="FFD966", bold=True, align_h="right")
        for col_idx, val in enumerate(cost_data[1:], 2):
            cell = ws9.cell(row=row, column=col_idx, value=val)
            set_cell_style(cell, bg_color="FFD966", bold=True, align_h="center" if col_idx < 4 else "left")
    elif cost_data[0] == "":
        row += 1
        continue
    else:
        ws9.merge_cells(f'A{row}:B{row}')
        ws9.cell(row=row, column=1, value=cost_data[0])
        ws9.cell(row=row, column=3, value=cost_data[1])
        ws9.cell(row=row, column=4, value=cost_data[2])
        ws9.merge_cells(f'E{row}:F{row}')
        ws9.cell(row=row, column=5, value=cost_data[3])
        for col in [1, 3, 4, 5]:
            set_cell_style(ws9.cell(row=row, column=col))
    row += 1

# 비용 최적화 팁
row += 1
ws9.merge_cells(f'A{row}:F{row}')
cell = ws9.cell(row=row, column=1, value="비용 최적화 팁")
set_cell_style(cell, bg_color="70AD47", font_color="FFFFFF", bold=True, align_h="center")
row += 1

optimization_tips = [
    "1. Gemini Free Tier 최대 활용: 월 1500 requests까지 무료이므로, 캐싱 및 배치 처리로 요청 수 줄이기",
    "2. fal.ai 비용 관리: Kling Video가 가장 비싸므로, 필요 시 대안 탐색 (Runway, Pika 등)",
    "3. VPS 업그레이드: 트래픽 증가 시 KVM 4로 업그레이드 ($13.99/월)",
    "4. Creatomate 최적화: 템플릿 재사용으로 렌더링 시간 단축",
    "5. YouTube API 할당량: 일일 10,000 units 초과 시 Google Cloud 유료 플랜 필요 (현재는 충분)"
]

for tip in optimization_tips:
    ws9.merge_cells(f'A{row}:F{row}')
    cell = ws9.cell(row=row, column=1, value=tip)
    set_cell_style(cell)
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws9, 1, 25)
auto_adjust_column_width(ws9, 2, 25)
auto_adjust_column_width(ws9, 3, 18)
auto_adjust_column_width(ws9, 4, 20)
auto_adjust_column_width(ws9, 5, 18)
auto_adjust_column_width(ws9, 6, 30)

# ========== Sheet 10: Google Sheets DB 설계 ==========
ws10 = wb.create_sheet("Google Sheets DB 설계")

# 제목
ws10.merge_cells('A1:E1')
cell = ws10['A1']
cell.value = "Google Sheets 데이터베이스 설계"
set_cell_style(cell, bg_color="2F5496", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws10.row_dimensions[1].height = 25

# 메인 시트 (확장)
row = 3
ws10.merge_cells(f'A{row}:E{row}')
cell = ws10.cell(row=row, column=1, value="메인 시트 (영상 제작 DB) - 확장 구조")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

main_sheet = [
    ("컬럼명", "타입", "설명", "예시 값", "필수 여부"),
    ("Subject", "Text", "영상 제목", "도박 중독 탈출하는 3가지 방법", "필수"),
    ("Narration", "Text", "전체 나레이션 대본", "여러분, 혹시 도박 때문에...", "필수"),
    ("Status", "Text", "제작 상태", "준비 / 생성완료", "필수"),
    ("Caption", "Text", "YouTube 설명", "이 영상은 도박 중독에서...", "필수"),
    ("댓글", "Text", "자동 댓글 내용", "더 많은 정보는...", "필수"),
    ("Publish", "Text", "발행 상태", "대기중 / 발행완료", "필수"),
    ("BGM prompt", "Text", "BGM 생성 프롬프트", "hopeful, inspiring, gentle piano", "필수"),
    ("imagePrompt", "Text", "(사용 안 함)", "", "선택"),
    ("업로드 URL", "Text", "YouTube URL", "https://youtube.com/watch?v=...", "자동"),
    ("row_number", "Number", "행 번호", "2", "자동"),
    ("채널", "Text", "어느 채널용인지", "쇼츠 전용 / 루믹스 솔루션 / 온카스터디 / 슬롯", "필수 (신규)"),
    ("장르/무드", "Text", "선택된 비주얼 스타일", "cyberpunk / nature / minimal_bw / ...", "자동 (신규)"),
    ("광고유형", "Text", "어떤 광고 넣을지", "온카 / 루믹스 / 외부광고 / 없음", "필수 (신규)"),
    ("템플릿ID", "Text", "Creatomate 템플릿 ID", "abc123-shorts / def456-lumix / ...", "자동 (신규)"),
    ("YouTube채널ID", "Text", "업로드할 YouTube 채널", "UCxxxxx...", "자동 (신규)"),
    ("기획안점수", "Number", "QA 평가 점수", "8", "자동 (신규)"),
    ("QA피드백", "Text", "QA 피드백 내용", "제목 개선 필요: ...", "자동 (신규)")
]

create_header(ws10, row, main_sheet[0], bg_color="2F5496")
row += 1
for sheet_data in main_sheet[1:]:
    for col_idx, val in enumerate(sheet_data, 1):
        cell = ws10.cell(row=row, column=col_idx, value=val)
        if "신규" in str(val):
            set_cell_style(cell, bg_color="FFEB9C")
        else:
            set_cell_style(cell)
    row += 1

# 성과 DB
row += 2
ws10.merge_cells(f'A{row}:E{row}')
cell = ws10.cell(row=row, column=1, value="성과 DB 시트 (DA 에이전트용) - 신규")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

performance_db = [
    ("컬럼명", "타입", "설명", "예시 값"),
    ("날짜", "Date", "데이터 수집 날짜", "2026-02-08"),
    ("채널", "Text", "YouTube 채널", "쇼츠 전용"),
    ("영상ID", "Text", "YouTube 영상 ID", "dQw4w9WgXcQ"),
    ("제목", "Text", "영상 제목", "도박 중독 탈출..."),
    ("조회수", "Number", "조회수", "15234"),
    ("평균 시청 시간", "Number", "초 단위", "42"),
    ("CTR", "Number", "클릭률 (%)", "5.8"),
    ("좋아요", "Number", "좋아요 수", "456"),
    ("댓글수", "Number", "댓글 수", "23"),
    ("유입소스", "Text", "주요 유입 경로", "YouTube 검색 / 추천"),
    ("장르/무드", "Text", "사용한 비주얼 스타일", "cyberpunk")
]

create_header(ws10, row, performance_db[0], bg_color="2F5496")
row += 1
for perf_data in performance_db[1:]:
    for col_idx, val in enumerate(perf_data, 1):
        cell = ws10.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="D6E4F0")
    row += 1

# 기획안 DB
row += 2
ws10.merge_cells(f'A{row}:E{row}')
cell = ws10.cell(row=row, column=1, value="기획안 DB 시트 (PL/QA 에이전트용) - 신규")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

plan_db = [
    ("컬럼명", "타입", "설명", "예시 값"),
    ("기획안ID", "Text", "고유 ID", "PLAN-20260208-001"),
    ("생성일시", "DateTime", "기획안 생성 시간", "2026-02-08 09:23:15"),
    ("채널", "Text", "대상 채널", "쇼츠 전용"),
    ("제목옵션1", "Text", "제목 1안", "도박 중독 탈출하는 법"),
    ("제목옵션2", "Text", "제목 2안", "도박으로 3천만원 날린 내가..."),
    ("제목옵션3", "Text", "제목 3안", "도박 중독자들이 숨기는 심리"),
    ("나레이션", "Text", "대본", "여러분, 혹시..."),
    ("톤앤매너", "Text", "콘텐츠 톤", "친근하고 공감적"),
    ("타겟오디언스", "Text", "타겟", "20-30대 남성"),
    ("Status", "Text", "기획 상태", "기획완료 / QA중 / 승인 / 폐기"),
    ("QA점수", "Number", "QA 평가 점수", "8"),
    ("QA피드백", "Text", "피드백 내용", "제목 2안 채택 권장"),
    ("루프횟수", "Number", "피드백 루프 횟수", "2"),
    ("최종승인일시", "DateTime", "승인 시간", "2026-02-08 10:45:30")
]

create_header(ws10, row, plan_db[0], bg_color="2F5496")
row += 1
for plan_data in plan_db[1:]:
    for col_idx, val in enumerate(plan_data, 1):
        cell = ws10.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="FFF2CC")
    row += 1

# 분석 리포트 시트
row += 2
ws10.merge_cells(f'A{row}:E{row}')
cell = ws10.cell(row=row, column=1, value="분석 리포트 시트 (DA 에이전트용) - 신규")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

report_db = [
    ("컬럼명", "타입", "설명", "예시 값"),
    ("리포트ID", "Text", "고유 ID", "REPORT-20260208"),
    ("생성일시", "DateTime", "리포트 생성 시간", "2026-02-08 06:30:00"),
    ("분석기간", "Text", "분석한 기간", "최근 7일 / 최근 30일"),
    ("채널", "Text", "대상 채널", "쇼츠 전용 / 전체"),
    ("주요인사이트", "Text", "AI 분석 결과", "사이버펑크 스타일이 15% 높은 조회수..."),
    ("권장사항", "Text", "다음 기획 권장", "쇼츠 채널에서 사이버펑크 스타일 2개 더 제작..."),
    ("통계요약", "Text", "통계 데이터", "평균 조회수: 12,345 / 평균 CTR: 5.2%")
]

create_header(ws10, row, report_db[0], bg_color="2F5496")
row += 1
for report_data in report_db[1:]:
    for col_idx, val in enumerate(report_data, 1):
        cell = ws10.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="E7E6E6")
    row += 1

# 설정 시트
row += 2
ws10.merge_cells(f'A{row}:E{row}')
cell = ws10.cell(row=row, column=1, value="설정 시트 (PM 에이전트용) - 신규")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

config_db = [
    ("컬럼명", "타입", "설명", "예시 값"),
    ("채널", "Text", "채널명", "쇼츠 전용"),
    ("일일목표개수", "Number", "하루 생성 목표", "3"),
    ("활성화", "Boolean", "활성화 여부", "TRUE / FALSE"),
    ("YouTube채널ID", "Text", "YouTube 채널 ID", "UCxxxxx..."),
    ("CreatomateTemplateID", "Text", "Creatomate 템플릿", "abc123-shorts"),
    ("기본광고유형", "Text", "기본 광고", "온카 / 루믹스 / 외부광고")
]

create_header(ws10, row, config_db[0], bg_color="2F5496")
row += 1
for config_data in config_db[1:]:
    for col_idx, val in enumerate(config_data, 1):
        cell = ws10.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="C6EFCE")
    row += 1

# 채널 가이드 시트
row += 2
ws10.merge_cells(f'A{row}:E{row}')
cell = ws10.cell(row=row, column=1, value="채널 가이드 시트 (PL 에이전트용) - 신규")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="center")
row += 1

guide_db = [
    ("컬럼명", "타입", "설명", "예시 값"),
    ("채널", "Text", "채널명", "쇼츠 전용"),
    ("콘텐츠방향", "Text", "콘텐츠 특성", "다양한 주제, 자극적 제목, 바이럴 콘텐츠"),
    ("톤앤매너", "Text", "말투/분위기", "친근하고 에너지 넘치는"),
    ("타겟오디언스", "Text", "타겟층", "20-30대 남성, 직장인"),
    ("금지사항", "Text", "하면 안 되는 것", "정치/종교/혐오 표현"),
    ("광고정책", "Text", "광고 관련", "8초 엔딩카드에 온카/루믹스/외부 광고 가능"),
    ("제목스타일", "Text", "제목 가이드", "50자 이내, 숫자 포함, 궁금증 유발")
]

create_header(ws10, row, guide_db[0], bg_color="2F5496")
row += 1
for guide_data in guide_db[1:]:
    for col_idx, val in enumerate(guide_data, 1):
        cell = ws10.cell(row=row, column=col_idx, value=val)
        set_cell_style(cell, bg_color="F8CBAD")
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws10, 1, 20)
auto_adjust_column_width(ws10, 2, 15)
auto_adjust_column_width(ws10, 3, 35)
auto_adjust_column_width(ws10, 4, 35)
auto_adjust_column_width(ws10, 5, 15)

# ========== Sheet 11: 지침서 - 온카스터디 채널 ==========
ws11 = wb.create_sheet("지침서 - 온카스터디 채널")
ws11.sheet_properties.tabColor = "00B050"  # 녹색

# 제목
ws11.merge_cells('A1:B1')
cell = ws11['A1']
cell.value = "온카스터디 채널 지침서 (AI 에이전트용)"
set_cell_style(cell, bg_color="00B050", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws11.row_dimensions[1].height = 25

# 노트
row = 2
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="⚠️ 이 지침서는 각 AI 에이전트의 시스템 프롬프트에 삽입됩니다")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")
ws11.row_dimensions[row].height = 20

# Section 1: 채널 기본 정보
row += 2
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="📌 Section 1: 채널 기본 정보")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section1_data = [
    ("채널명", "온카스터디"),
    ("목적", "온카스터디 커뮤니티 홍보 및 알리기"),
    ("타겟", "20-40대 남녀"),
    ("톤", "신뢰감 있는 교육적 톤, 친근하면서도 전문적"),
    ("분위기", "커뮤니티의 가치와 성과를 보여주는 느낌, '같이 성장하자'"),
    ("콘텐츠 형식", "숏폼 + 롱폼")
]

for item, value in section1_data:
    ws11.cell(row=row, column=1, value=item)
    ws11.cell(row=row, column=2, value=value)
    set_cell_style(ws11.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws11.cell(row=row, column=2))
    row += 1

# Section 2: 콘텐츠 규칙
row += 1
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="📌 Section 2: 콘텐츠 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section2_data = [
    ("다뤄야 할 주제", "온카스터디 커뮤니티 소개, 멤버 성공 사례, 스터디 방법론, 커뮤니티 혜택, 후기/리뷰"),
    ("금지 주제", "정치, 종교, 혐오표현, 폭력, 불법 콘텐츠"),
    ("중복 기준", "최근 7일 내 같은 키워드 주제 → 패스"),
    ("콘텐츠 비율", "교육/정보 60%, 홍보/CTA 30%, 트렌드/재미 10%")
]

for item, value in section2_data:
    ws11.cell(row=row, column=1, value=item)
    ws11.cell(row=row, column=2, value=value)
    set_cell_style(ws11.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws11.cell(row=row, column=2))
    row += 1

# Section 3: 영상 제작 규칙
row += 1
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="📌 Section 3: 영상 제작 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section3_data = [
    ("숏폼 길이", "30-60초"),
    ("롱폼 길이", "5-15분"),
    ("자막 스타일", "깔끔한 화이트/블랙, 핵심 키워드 강조"),
    ("BGM 분위기", "밝고 희망적, 업비트, 동기부여"),
    ("이미지 스타일", "밝은 조명, 깨끗한 배경, 사람/커뮤니티 느낌"),
    ("나레이션 톤", "차분하고 신뢰감 있는, 설명하는 느낌")
]

for item, value in section3_data:
    ws11.cell(row=row, column=1, value=item)
    ws11.cell(row=row, column=2, value=value)
    set_cell_style(ws11.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws11.cell(row=row, column=2))
    row += 1

# Section 4: 광고 삽입 규칙
row += 1
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="📌 Section 4: 광고 삽입 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section4_data = [
    ("광고 유형", "자체 홍보 (콘텐츠 자체가 홍보)"),
    ("외부 광고", "없음"),
    ("CTA", "'온카스터디 커뮤니티 가입', '링크 확인' 등"),
    ("CTA 위치", "영상 마지막 5초 또는 설명란")
]

for item, value in section4_data:
    ws11.cell(row=row, column=1, value=item)
    ws11.cell(row=row, column=2, value=value)
    set_cell_style(ws11.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws11.cell(row=row, column=2))
    row += 1

# Section 5: 업로드 규칙
row += 1
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="📌 Section 5: 업로드 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section5_data = [
    ("제목 형식", "[온카스터디] + 핵심 키워드 + 호기심 유발"),
    ("제목 길이", "30-50자"),
    ("태그", "온카스터디, 커뮤니티, 스터디, 자기개발, 성장 + 영상 관련 태그"),
    ("설명", "영상 요약 2-3줄 + 커뮤니티 링크 + 해시태그"),
    ("공개 설정", "공개"),
    ("카테고리", "교육 (Education)")
]

for item, value in section5_data:
    ws11.cell(row=row, column=1, value=item)
    ws11.cell(row=row, column=2, value=value)
    set_cell_style(ws11.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws11.cell(row=row, column=2))
    row += 1

# Section 6: QA 채점 기준
row += 1
ws11.merge_cells(f'A{row}:B{row}')
cell = ws11.cell(row=row, column=1, value="📌 Section 6: QA 채점 기준 (이 채널 전용)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section6_data = [
    ("교육적 가치", "스터디/성장 관련 내용인가 (1-10점)"),
    ("커뮤니티 홍보력", "온카스터디를 알리는 데 효과적인가 (1-10점)"),
    ("클릭 유도력", "제목/썸네일이 클릭하고 싶은가 (1-10점)"),
    ("채널 톤 적합성", "신뢰감+친근함이 느껴지는가 (1-10점)"),
    ("참신함", "최근 영상과 차별화되는가 (1-10점)"),
    ("합격 기준", "평균 7점 이상"),
    ("자동 탈락", "교육적 가치 3점 이하 또는 커뮤니티 홍보력 3점 이하")
]

for item, value in section6_data:
    ws11.cell(row=row, column=1, value=item)
    ws11.cell(row=row, column=2, value=value)
    set_cell_style(ws11.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws11.cell(row=row, column=2))
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws11, 1, 25)
auto_adjust_column_width(ws11, 2, 70)

# ========== Sheet 12: 지침서 - 슬롯 채널 ==========
ws12 = wb.create_sheet("지침서 - 슬롯 채널")
ws12.sheet_properties.tabColor = "FF0000"  # 빨강

# 제목
ws12.merge_cells('A1:B1')
cell = ws12['A1']
cell.value = "슬롯 채널 지침서 (AI 에이전트용)"
set_cell_style(cell, bg_color="C00000", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws12.row_dimensions[1].height = 25

# 노트
row = 2
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="⚠️ 이 지침서는 각 AI 에이전트의 시스템 프롬프트에 삽입됩니다")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")
ws12.row_dimensions[row].height = 20

# Section 1: 채널 기본 정보
row += 2
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="📌 Section 1: 채널 기본 정보")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section1_data = [
    ("채널명", "슬롯 채널"),
    ("목적", "슬롯 관련 콘텐츠 + 온카스터디 광고 목적"),
    ("타겟", "20-40대 남녀"),
    ("톤", "흥미 유발, 자극적이면서 정보성 있는, 게임/엔터테인먼트 느낌"),
    ("분위기", "화려하고 다이나믹, 승리/스릴의 느낌"),
    ("콘텐츠 형식", "숏폼 + 롱폼")
]

for item, value in section1_data:
    ws12.cell(row=row, column=1, value=item)
    ws12.cell(row=row, column=2, value=value)
    set_cell_style(ws12.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws12.cell(row=row, column=2))
    row += 1

# Section 2: 콘텐츠 규칙
row += 1
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="📌 Section 2: 콘텐츠 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section2_data = [
    ("다뤄야 할 주제", "슬롯 게임 관련 정보, 전략, 리뷰, 뉴스, 재미있는 슬롯 콘텐츠"),
    ("금지 주제", "정치, 종교, 혐오표현, 폭력, 불법 콘텐츠, 도박 조장/중독 조장 표현, 확정 수익 보장 표현"),
    ("중복 기준", "최근 7일 내 같은 키워드 주제 → 패스"),
    ("콘텐츠 비율", "엔터테인먼트/정보 70%, 온카광고 30%")
]

for item, value in section2_data:
    ws12.cell(row=row, column=1, value=item)
    ws12.cell(row=row, column=2, value=value)
    set_cell_style(ws12.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws12.cell(row=row, column=2))
    row += 1

# Section 3: 영상 제작 규칙
row += 1
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="📌 Section 3: 영상 제작 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section3_data = [
    ("숏폼 길이", "30-60초"),
    ("롱폼 길이", "5-15분"),
    ("자막 스타일", "화려한 네온 스타일, 강렬한 색상, 핵심 키워드 강조"),
    ("BGM 분위기", "다이나믹하고 긴장감 있는, 카지노/게임 분위기"),
    ("이미지 스타일", "화려한 네온, 카지노 분위기, 다이나믹한 이펙트"),
    ("나레이션 톤", "흥분되고 에너지 넘치는, 호기심 자극하는 느낌")
]

for item, value in section3_data:
    ws12.cell(row=row, column=1, value=item)
    ws12.cell(row=row, column=2, value=value)
    set_cell_style(ws12.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws12.cell(row=row, column=2))
    row += 1

# Section 4: 광고 삽입 규칙
row += 1
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="📌 Section 4: 광고 삽입 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section4_data = [
    ("광고 유형", "슬롯 관련 콘텐츠만 + 온카스터디 광고 삽입 (영상 마지막)"),
    ("외부 광고", "없음 (온카스터디 외 타 광고 절대 금지)"),
    ("CTA", "'온카스터디 커뮤니티 가입', '더 많은 정보' 등"),
    ("CTA 위치", "영상 마지막 5초 또는 설명란")
]

for item, value in section4_data:
    ws12.cell(row=row, column=1, value=item)
    ws12.cell(row=row, column=2, value=value)
    set_cell_style(ws12.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws12.cell(row=row, column=2))
    row += 1

# Section 5: 업로드 규칙
row += 1
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="📌 Section 5: 업로드 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section5_data = [
    ("제목 형식", "호기심+자극적 키워드 (슬롯 관련)"),
    ("제목 길이", "30-50자"),
    ("태그", "슬롯, 카지노, 게임 + 영상 관련 태그"),
    ("설명", "영상 요약 2-3줄 + 온카스터디 링크 (자연스럽게) + 해시태그"),
    ("공개 설정", "공개"),
    ("카테고리", "게임 (Gaming)")
]

for item, value in section5_data:
    ws12.cell(row=row, column=1, value=item)
    ws12.cell(row=row, column=2, value=value)
    set_cell_style(ws12.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws12.cell(row=row, column=2))
    row += 1

# Section 6: QA 채점 기준
row += 1
ws12.merge_cells(f'A{row}:B{row}')
cell = ws12.cell(row=row, column=1, value="📌 Section 6: QA 채점 기준 (이 채널 전용)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section6_data = [
    ("엔터테인먼트 가치", "슬롯/게임 관련으로 재미있는가 (1-10점)"),
    ("온카광고 자연스러움", "온카스터디 광고가 자연스럽게 삽입되었는가 (1-10점)"),
    ("클릭 유도력", "제목/썸네일이 클릭하고 싶은가 (1-10점)"),
    ("채널 톤 적합성", "화려하고 다이나믹한 느낌이 있는가 (1-10점)"),
    ("참신함", "최근 영상과 차별화되는가 (1-10점)"),
    ("합격 기준", "평균 7점 이상"),
    ("자동 탈락", "엔터테인먼트 가치 3점 이하 또는 온카광고 자연스러움 3점 이하")
]

for item, value in section6_data:
    ws12.cell(row=row, column=1, value=item)
    ws12.cell(row=row, column=2, value=value)
    set_cell_style(ws12.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws12.cell(row=row, column=2))
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws12, 1, 25)
auto_adjust_column_width(ws12, 2, 70)

# ========== Sheet 13: 지침서 - 루믹스 솔루션 채널 ==========
ws13 = wb.create_sheet("지침서 - 루믹스 솔루션 채널")
ws13.sheet_properties.tabColor = "0070C0"  # 파랑

# 제목
ws13.merge_cells('A1:B1')
cell = ws13['A1']
cell.value = "루믹스 솔루션 채널 지침서 (AI 에이전트용)"
set_cell_style(cell, bg_color="0070C0", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws13.row_dimensions[1].height = 25

# 노트
row = 2
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="⚠️ 이 지침서는 각 AI 에이전트의 시스템 프롬프트에 삽입됩니다")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")
ws13.row_dimensions[row].height = 20

# Section 1: 채널 기본 정보
row += 2
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="📌 Section 1: 채널 기본 정보")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section1_data = [
    ("채널명", "루믹스 솔루션 채널"),
    ("목적", "루믹스 솔루션 판매 및 광고, 알리기"),
    ("타겟", "20-40대 남녀"),
    ("톤", "전문적이고 설득력 있는, 솔루션의 가치를 강조"),
    ("분위기", "깔끔하고 프로페셔널, 기술/혁신 느낌"),
    ("콘텐츠 형식", "숏폼 + 롱폼")
]

for item, value in section1_data:
    ws13.cell(row=row, column=1, value=item)
    ws13.cell(row=row, column=2, value=value)
    set_cell_style(ws13.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws13.cell(row=row, column=2))
    row += 1

# Section 2: 콘텐츠 규칙
row += 1
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="📌 Section 2: 콘텐츠 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section2_data = [
    ("다뤄야 할 주제", "루믹스 솔루션 기능 소개, 사용 사례, 도입 효과, 비교 분석, 고객 후기"),
    ("금지 주제", "정치, 종교, 혐오표현, 폭력, 불법 콘텐츠"),
    ("중복 기준", "최근 7일 내 같은 키워드 주제 → 패스"),
    ("콘텐츠 비율", "솔루션 설명/판매 80%, 업계 정보/트렌드 20%")
]

for item, value in section2_data:
    ws13.cell(row=row, column=1, value=item)
    ws13.cell(row=row, column=2, value=value)
    set_cell_style(ws13.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws13.cell(row=row, column=2))
    row += 1

# Section 3: 영상 제작 규칙
row += 1
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="📌 Section 3: 영상 제작 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section3_data = [
    ("숏폼 길이", "30-60초"),
    ("롱폼 길이", "5-15분"),
    ("자막 스타일", "깔끔한 화이트/블랙, 핵심 기능/혜택 강조"),
    ("BGM 분위기", "세련되고 모던한, 테크/혁신 분위기"),
    ("이미지 스타일", "깔끔한 UI/대시보드, 테크 느낌, 미래지향적"),
    ("나레이션 톤", "전문적이고 신뢰감 있는, 설득하는 느낌")
]

for item, value in section3_data:
    ws13.cell(row=row, column=1, value=item)
    ws13.cell(row=row, column=2, value=value)
    set_cell_style(ws13.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws13.cell(row=row, column=2))
    row += 1

# Section 4: 광고 삽입 규칙
row += 1
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="📌 Section 4: 광고 삽입 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section4_data = [
    ("광고 유형", "자체 판매/홍보 (콘텐츠 자체가 광고)"),
    ("외부 광고", "없음"),
    ("CTA", "'루믹스 솔루션 도입 문의', '무료 체험', '상담 신청'"),
    ("CTA 위치", "영상 마지막 5초 또는 설명란")
]

for item, value in section4_data:
    ws13.cell(row=row, column=1, value=item)
    ws13.cell(row=row, column=2, value=value)
    set_cell_style(ws13.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws13.cell(row=row, column=2))
    row += 1

# Section 5: 업로드 규칙
row += 1
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="📌 Section 5: 업로드 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section5_data = [
    ("제목 형식", "[루믹스] + 솔루션 키워드 + 혜택 강조"),
    ("제목 길이", "30-50자"),
    ("태그", "루믹스, 솔루션, 비즈니스, 자동화 + 영상 관련 태그"),
    ("설명", "솔루션 설명 2-3줄 + 도입 문의 링크 + 해시태그"),
    ("공개 설정", "공개"),
    ("카테고리", "과학과 기술 (Science & Technology)")
]

for item, value in section5_data:
    ws13.cell(row=row, column=1, value=item)
    ws13.cell(row=row, column=2, value=value)
    set_cell_style(ws13.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws13.cell(row=row, column=2))
    row += 1

# Section 6: QA 채점 기준
row += 1
ws13.merge_cells(f'A{row}:B{row}')
cell = ws13.cell(row=row, column=1, value="📌 Section 6: QA 채점 기준 (이 채널 전용)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section6_data = [
    ("판매 설득력", "솔루션 도입을 원하게 만드는가 (1-10점)"),
    ("솔루션 설명력", "기능과 혜택이 명확히 전달되는가 (1-10점)"),
    ("클릭 유도력", "제목/썸네일이 클릭하고 싶은가 (1-10점)"),
    ("전문적 톤", "전문적이고 신뢰감 있는가 (1-10점)"),
    ("참신함", "최근 영상과 차별화되는가 (1-10점)"),
    ("합격 기준", "평균 7점 이상"),
    ("자동 탈락", "판매 설득력 3점 이하 또는 솔루션 설명력 3점 이하")
]

for item, value in section6_data:
    ws13.cell(row=row, column=1, value=item)
    ws13.cell(row=row, column=2, value=value)
    set_cell_style(ws13.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws13.cell(row=row, column=2))
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws13, 1, 25)
auto_adjust_column_width(ws13, 2, 70)

# ========== Sheet 14: 지침서 - 쇼츠 전용 채널 ==========
ws14 = wb.create_sheet("지침서 - 쇼츠 전용 채널")
ws14.sheet_properties.tabColor = "FFC000"  # 오렌지

# 제목
ws14.merge_cells('A1:B1')
cell = ws14['A1']
cell.value = "쇼츠 전용 채널 지침서 (AI 에이전트용)"
set_cell_style(cell, bg_color="C65911", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws14.row_dimensions[1].height = 25

# 노트
row = 2
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="⚠️ 이 지침서는 각 AI 에이전트의 시스템 프롬프트에 삽입됩니다")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")
ws14.row_dimensions[row].height = 20

# Section 1: 채널 기본 정보
row += 2
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="📌 Section 1: 채널 기본 정보")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section1_data = [
    ("채널명", "쇼츠 전용 채널"),
    ("목적", "데이터 수집(조회수/체류/클릭 분석) + 마지막에 광고 삽입"),
    ("타겟", "20-40대 남녀"),
    ("톤", "다양함 — 주제에 따라 유연하게 변경 (재미/정보/자극/감동 등)"),
    ("분위기", "첫 1초에 시선 잡기, 끝까지 보게 만들기"),
    ("콘텐츠 형식", "숏폼 only")
]

for item, value in section1_data:
    ws14.cell(row=row, column=1, value=item)
    ws14.cell(row=row, column=2, value=value)
    set_cell_style(ws14.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws14.cell(row=row, column=2))
    row += 1

# Section 2: 콘텐츠 규칙
row += 1
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="📌 Section 2: 콘텐츠 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section2_data = [
    ("다뤄야 할 주제", "다양한 소재 (트렌드, 흥미로운 사실, 리스트, 비교, 꿀팁 등)"),
    ("금지 주제", "정치, 종교, 혐오표현, 폭력, 불법 콘텐츠"),
    ("중복 기준", "최근 7일 내 같은 키워드 주제 → 패스"),
    ("특수 규칙", "영상은 다양한 장르/주제 (DA 분석 데이터 기반), 마지막에 온카스터디 또는 루믹스 광고 삽입, 3개 중 1개는 광고 포함, 광고가 자연스럽게 느껴져야 함")
]

for item, value in section2_data:
    ws14.cell(row=row, column=1, value=item)
    ws14.cell(row=row, column=2, value=value)
    set_cell_style(ws14.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws14.cell(row=row, column=2))
    row += 1

# Section 3: 영상 제작 규칙
row += 1
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="📌 Section 3: 영상 제작 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section3_data = [
    ("숏폼 길이", "30-60초"),
    ("자막 스타일", "주제에 따라 다양하게 (재미있는 영상은 캐주얼, 정보성은 깔끔)"),
    ("BGM 분위기", "주제에 따라 다양하게"),
    ("이미지 스타일", "주제에 따라 다양하게 (장르 랜덤 시스템 활용)"),
    ("나레이션 톤", "주제에 따라 다양하게")
]

for item, value in section3_data:
    ws14.cell(row=row, column=1, value=item)
    ws14.cell(row=row, column=2, value=value)
    set_cell_style(ws14.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws14.cell(row=row, column=2))
    row += 1

# Section 4: 광고 삽입 규칙
row += 1
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="📌 Section 4: 광고 삽입 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section4_data = [
    ("광고 유형", "영상 마지막 5-8초에 온카스터디 or 루믹스 광고"),
    ("광고 로테이션", "온카 → 루믹스 → 온카 → 루믹스 (교대)"),
    ("외부 광고", "없음"),
    ("CTA", "광고에 따라 변경 (온카: 커뮤니티 가입 / 루믹스: 도입 문의)"),
    ("CTA 위치", "영상 마지막 5-8초")
]

for item, value in section4_data:
    ws14.cell(row=row, column=1, value=item)
    ws14.cell(row=row, column=2, value=value)
    set_cell_style(ws14.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws14.cell(row=row, column=2))
    row += 1

# Section 5: 업로드 규칙
row += 1
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="📌 Section 5: 업로드 규칙")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section5_data = [
    ("제목 형식", "호기심 극대화, 클릭 유도"),
    ("제목 길이", "30-50자"),
    ("태그", "주제 관련 태그 + 트렌드 태그"),
    ("설명", "영상 요약 1-2줄 + 광고 링크 (자연스럽게) + 해시태그"),
    ("공개 설정", "공개"),
    ("카테고리", "주제에 따라 변경 (엔터테인먼트, 교육, 과학 등)")
]

for item, value in section5_data:
    ws14.cell(row=row, column=1, value=item)
    ws14.cell(row=row, column=2, value=value)
    set_cell_style(ws14.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws14.cell(row=row, column=2))
    row += 1

# Section 6: QA 채점 기준
row += 1
ws14.merge_cells(f'A{row}:B{row}')
cell = ws14.cell(row=row, column=1, value="📌 Section 6: QA 채점 기준 (이 채널 전용)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

section6_data = [
    ("시선 잡기 (훅)", "첫 1초에 시선을 잡는가 (1-10점)"),
    ("체류시간 예측", "끝까지 볼 만한 내용인가 (1-10점)"),
    ("클릭 유도력", "제목/썸네일이 클릭하고 싶은가 (1-10점)"),
    ("광고 자연스러움", "광고가 억지스럽지 않고 자연스러운가 (1-10점)"),
    ("참신함", "최근 영상과 차별화되는가 (1-10점)"),
    ("합격 기준", "평균 7점 이상"),
    ("자동 탈락", "시선 잡기 3점 이하 또는 체류시간 예측 3점 이하"),
    ("특수", "DA 분석 결과 반영 필수 (어떤 주제가 잘 되는지)")
]

for item, value in section6_data:
    ws14.cell(row=row, column=1, value=item)
    ws14.cell(row=row, column=2, value=value)
    set_cell_style(ws14.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
    set_cell_style(ws14.cell(row=row, column=2))
    row += 1

# 컬럼 너비
auto_adjust_column_width(ws14, 1, 25)
auto_adjust_column_width(ws14, 2, 70)

# ========== Sheet 15: 에이전트별 시스템 프롬프트 (예시) ==========
ws15 = wb.create_sheet("에이전트별 시스템 프롬프트")
ws15.sheet_properties.tabColor = "7030A0"  # 보라

# 제목
ws15.merge_cells('A1:B1')
cell = ws15['A1']
cell.value = "에이전트별 시스템 프롬프트 (예시)"
set_cell_style(cell, bg_color="7030A0", font_color="FFFFFF", bold=True, align_h="center")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
ws15.row_dimensions[1].height = 25

# 노트
row = 2
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="💡 각 에이전트의 시스템 프롬프트에 채널별 지침서를 자동으로 삽입하여 사용합니다")
set_cell_style(cell, bg_color="FFF2CC", bold=True, align_h="center")
ws15.row_dimensions[row].height = 20

# DA 에이전트
row += 2
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="🤖 DA 에이전트 (Data Analyst)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

ws15.cell(row=row, column=1, value="역할")
ws15.cell(row=row, column=2, value="유튜브 채널 데이터 분석 전문가")
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
row += 1

ws15.cell(row=row, column=1, value="프롬프트 예시")
prompt_text = """당신은 유튜브 채널 데이터 분석 전문가입니다. 4개 채널(온카스터디, 슬롯, 루믹스, 쇼츠전용)의 성과 데이터를 분석하고, 각 채널별로 다음 영상 방향을 제안하는 리포트를 작성하세요.

리포트에는 다음을 포함하세요:
1) 최근 7일 성과 요약
2) 잘 된 영상 패턴
3) 채널별 추천 주제 3개씩
4) 주의사항

각 채널의 특성과 타겟을 고려하여 분석하세요."""
ws15.cell(row=row, column=2, value=prompt_text)
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
ws15.row_dimensions[row].height = 120
row += 1

# PL 에이전트
row += 1
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="🤖 PL 에이전트 (Planner)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

ws15.cell(row=row, column=1, value="역할")
ws15.cell(row=row, column=2, value="유튜브 영상 기획 전문가")
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
row += 1

ws15.cell(row=row, column=1, value="프롬프트 예시")
prompt_text = """당신은 유튜브 영상 기획 전문가입니다. DA 분석 리포트를 바탕으로 [채널명] 채널의 영상 기획안을 생성하세요.

[채널 지침서 내용 자동 삽입]

기획안에는 다음을 포함하세요:
1) 제목 (30-50자)
2) 나레이션 대본 (30-60초 분량)
3) 분위기/톤
4) 예상 타겟 반응

서로 다른 방향의 기획안 2개를 생성하세요."""
ws15.cell(row=row, column=2, value=prompt_text)
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
ws15.row_dimensions[row].height = 120
row += 1

# QA 에이전트
row += 1
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="🤖 QA 에이전트 (Quality Assurance)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

ws15.cell(row=row, column=1, value="역할")
ws15.cell(row=row, column=2, value="유튜브 콘텐츠 품질 검수관")
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
row += 1

ws15.cell(row=row, column=1, value="프롬프트 예시")
prompt_text = """당신은 유튜브 콘텐츠 품질 검수관입니다. [채널명] 채널의 기획안을 평가하세요.

[채널별 QA 채점 기준 자동 삽입]

JSON 형식으로 출력:
{
  "score1": X,
  "score2": X,
  "score3": X,
  "score4": X,
  "score5": X,
  "average": X,
  "verdict": "승인/수정/폐기",
  "feedback": "구체적 피드백"
}

각 점수는 1-10점으로 평가하며, 평균 7점 이상이면 승인, 5-7점이면 수정, 5점 미만이면 폐기입니다."""
ws15.cell(row=row, column=2, value=prompt_text)
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
ws15.row_dimensions[row].height = 150
row += 1

# PR 에이전트 (이미지)
row += 1
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="🤖 PR 에이전트 (Production - Image)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

ws15.cell(row=row, column=1, value="역할")
ws15.cell(row=row, column=2, value="유튜브 숏폼 영상의 비주얼 디렉터")
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
row += 1

ws15.cell(row=row, column=1, value="프롬프트 예시")
prompt_text = """당신은 유튜브 숏폼 영상의 비주얼 디렉터입니다. 다음 나레이션 파트에 정확히 매칭되는 시네마틱 이미지 프롬프트를 생성하세요.

[채널별 이미지 스타일 지침 자동 삽입]

이번 영상의 비주얼 장르: [랜덤 선택된 장르]
나레이션: [분할된 나레이션 파트]

각 나레이션 파트마다 하나의 이미지 프롬프트를 생성하세요. 프롬프트는 영어로 작성하며, 채널의 톤과 분위기를 반영해야 합니다."""
ws15.cell(row=row, column=2, value=prompt_text)
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
ws15.row_dimensions[row].height = 130
row += 1

# DL 에이전트
row += 1
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="🤖 DL 에이전트 (Deploy & Launch)")
set_cell_style(cell, bg_color="4472C4", font_color="FFFFFF", bold=True, align_h="left")
row += 1

ws15.cell(row=row, column=1, value="역할")
ws15.cell(row=row, column=2, value="유튜브 SEO 전문가")
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
row += 1

ws15.cell(row=row, column=1, value="프롬프트 예시")
prompt_text = """당신은 유튜브 SEO 전문가입니다. [채널명] 채널에 업로드할 영상의 메타데이터를 최적화하세요.

[채널별 업로드 규칙 자동 삽입]

영상 내용: [기획안 요약]

JSON 형식으로 출력:
{
  "title": "최적화된 제목 (30-50자)",
  "description": "최적화된 설명 (영상 요약 + CTA + 해시태그)",
  "tags": ["태그1", "태그2", "태그3", ...],
  "category": "카테고리명"
}

제목은 클릭을 유도하면서도 채널 지침을 따라야 하며, 설명은 SEO에 최적화되어야 합니다."""
ws15.cell(row=row, column=2, value=prompt_text)
set_cell_style(ws15.cell(row=row, column=1), bg_color="E7E6E6", bold=True)
set_cell_style(ws15.cell(row=row, column=2))
ws15.row_dimensions[row].height = 150
row += 1

# 사용 방법
row += 2
ws15.merge_cells(f'A{row}:B{row}')
cell = ws15.cell(row=row, column=1, value="📚 사용 방법")
set_cell_style(cell, bg_color="70AD47", font_color="FFFFFF", bold=True, align_h="left")
row += 1

usage_text = """1. 각 에이전트의 시스템 프롬프트에 해당 채널의 지침서 내용을 자동으로 삽입합니다.
2. 예를 들어, 온카스터디 채널의 영상을 기획할 때는 PL 에이전트의 프롬프트에 '지침서 - 온카스터디 채널'의 내용을 삽입합니다.
3. QA 에이전트는 각 채널의 QA 채점 기준을 참고하여 기획안을 평가합니다.
4. PR 에이전트는 각 채널의 이미지 스타일 지침을 참고하여 이미지 프롬프트를 생성합니다.
5. DL 에이전트는 각 채널의 업로드 규칙을 참고하여 메타데이터를 최적화합니다.
6. 이렇게 하면 각 채널의 특성에 맞는 콘텐츠가 자동으로 생성됩니다."""
ws15.merge_cells(f'A{row}:B{row}')
ws15.cell(row=row, column=1, value=usage_text)
set_cell_style(ws15.cell(row=row, column=1))
ws15.row_dimensions[row].height = 100

# 컬럼 너비
auto_adjust_column_width(ws15, 1, 25)
auto_adjust_column_width(ws15, 2, 80)

# 파일 저장
output_path = "/Users/gimdongseog/n8n-project/n8n_기획서_v3.xlsx"
wb.save(output_path)
print(f"✅ Excel file created successfully: {output_path}")
print(f"📊 Total sheets: {len(wb.sheetnames)}")
print(f"📋 Sheets: {', '.join(wb.sheetnames)}")
